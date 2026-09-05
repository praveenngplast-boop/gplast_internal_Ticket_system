from django.utils import timezone
from openpyxl.styles import Alignment, Font, PatternFill


def add_replies_sheet(workbook, tickets_queryset):
    worksheet = workbook.create_sheet('Ticket Replies')
    headers = ['Ticket Number', 'Timestamp', 'Sender', 'Role', 'Reply', 'Attachment']
    worksheet.append(headers)

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for reply in tickets_queryset.prefetch_related('replies').order_by('created_at').values_list(
        'id', 'ticket_number'
    ).iterator():
        ticket_id, ticket_number = reply
        for ticket_reply in tickets_queryset.model.objects.get(pk=ticket_id).replies.all():
            worksheet.append([
                ticket_number,
                timezone.localtime(ticket_reply.created_at).strftime('%d-%b-%Y %I:%M %p'),
                ticket_reply.author_name,
                ticket_reply.author_role,
                ticket_reply.body,
                ticket_reply.attachment.name if ticket_reply.attachment else '',
            ])

    widths = [20, 23, 24, 16, 60, 35]
    for index, width in enumerate(widths, 1):
        worksheet.column_dimensions[chr(64 + index)].width = width
    worksheet.freeze_panes = 'A2'


def append_replies_section(worksheet, row, ticket, section_font, header_font, data_font, section_fill, header_fill, border):
    worksheet.merge_cells(f'A{row}:F{row}')
    worksheet[f'A{row}'] = 'REPLIES'
    worksheet[f'A{row}'].font = section_font
    worksheet[f'A{row}'].fill = section_fill
    worksheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    row += 1

    headers = ['Timestamp', 'Sender', 'Role', 'Reply', 'Attachment']
    for column, header in enumerate(headers, 1):
        cell = worksheet.cell(row=row, column=column, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    for reply in ticket.replies.all().order_by('created_at'):
        values = [
            timezone.localtime(reply.created_at).strftime('%d-%b-%Y %I:%M %p'),
            reply.author_name,
            reply.author_role,
            reply.body,
            reply.attachment.name if reply.attachment else '',
        ]
        for column, value in enumerate(values, 1):
            cell = worksheet.cell(row=row, column=column, value=value)
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        row += 1

    return row + 1
