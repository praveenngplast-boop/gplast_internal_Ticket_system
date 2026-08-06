from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import logout, update_session_auth_hash
from django.contrib.auth.models import User
from django.contrib.auth.views import LoginView
from django.contrib.auth import login as auth_login
from django.urls import reverse_lazy
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Count, Q, DateField
from django.db.models.functions import TruncMonth, Cast
from django.db import transaction, IntegrityError
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.template.loader import render_to_string
from django.template import TemplateDoesNotExist
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.cache import never_cache

from datetime import datetime, timedelta
import json
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
import logging

from tickets.models import Unit, Department, AdminContact, AdminNotificationEmail, Ticket, TicketHistory, EmployeeMaster, DepartmentCredential
from tickets.forms import TicketForm, AdminTicketForm, AdminContactForm, UnitForm, DepartmentForm, AdminNotificationEmailForm, AdminPasswordChangeForm, AdminSetUserPasswordForm, UserSelectionForm
from tickets.utils import generate_ticket_number, send_ticket_email

logger = logging.getLogger(__name__)


def is_admin(user):
    return user.is_authenticated and user.is_staff


def format_timedelta_display(td):
    if not td:
        return ""
    days = td.days
    hours, remainder = divmod(td.seconds, 3600)
    minutes, _ = divmod(remainder, 60)
    parts = []
    if days > 0:
        parts.append(f"{days} day{'s' if days > 1 else ''}")
    if hours > 0:
        parts.append(f"{hours} hour{'s' if hours > 1 else ''}")
    if minutes > 0:
        parts.append(f"{minutes} min{'s' if minutes > 1 else ''}")
    if not parts:
        return "< 1 minute"
    return ", ".join(parts)


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def reopen_ticket_logic(ticket, performed_by, remarks):
    with transaction.atomic():
        ticket.status = 'Open'
        ticket.closed_by = None
        ticket.closed_at = None
        ticket.closing_remarks = None
        ticket.save()
        TicketHistory.objects.create(
            ticket=ticket,
            action="Ticket Reopened",
            remarks=remarks,
            performed_by=performed_by
        )
    send_ticket_email(ticket, 'Reopened', remarks=remarks)


def generate_ticket_list_html(tickets, status):
    if not tickets:
        return """
        <div class="empty-state text-center py-5">
            <i class="fa-solid fa-receipt fa-3x mb-3 d-block opacity-25" style="color: var(--accent);"></i>
            <h6 style="color: var(--text-secondary); font-weight: 600;">No Tickets Found</h6>
            <p style="color: var(--text-muted); font-size: 0.85rem;">
                No tickets found.
            </p>
            <a href="javascript:void(0)" onclick="window.location.href='/create-ticket/'" class="btn btn-primary-custom btn-sm mt-2">
                <i class="fa-solid fa-circle-plus me-1"></i>Create New Ticket
            </a>
        </div>
        """
    
    html = """
    <div class="table-responsive">
        <table class="table table-hover align-middle mb-0">
            <thead>
                <tr>
                    <th style="font-size: 0.6rem; text-transform: uppercase; color: var(--text-secondary);">Ticket</th>
                    <th style="font-size: 0.6rem; text-transform: uppercase; color: var(--text-secondary);">Created</th>
                    <th style="font-size: 0.6rem; text-transform: uppercase; color: var(--text-secondary);">Subject</th>
                    <th style="font-size: 0.6rem; text-transform: uppercase; color: var(--text-secondary);">Status</th>
                    <th class="text-center" style="font-size: 0.6rem; text-transform: uppercase; color: var(--text-secondary);">Action</th>
                </tr>
            </thead>
            <tbody>
    """
    
    for ticket in tickets:
        status_class = ticket.status.lower()
        badge_class = {
            'open': 'badge-status-open',
            'assigned': 'badge-status-assigned',
            'hold': 'badge-status-hold',
            'escalated': 'badge-status-escalated',
            'closed': 'badge-status-closed',
        }.get(status_class, 'badge-status-open')
        
        html += f"""
                <tr>
                    <td class="fw-bold" style="color: var(--accent-light); font-size: 0.7rem;">
                        {ticket.ticket_number}
                    </td>
                    <td style="font-size: 0.65rem; color: var(--text-secondary);">
                        {ticket.created_at.strftime('%d-%m-%Y') if ticket.created_at else '-'}
                        <span class="d-block" style="font-size: 0.55rem; color: var(--text-muted);">{ticket.created_at.strftime('%I:%M %p') if ticket.created_at else ''}</span>
                    </td>
                    <td class="text-truncate" style="max-width: 120px; font-size: 0.7rem;" title="{ticket.subject}">
                        {ticket.subject}
                    </td>
                    <td>
                        <span class="badge-custom {badge_class}" style="font-size: 0.5rem; padding: 0.15rem 0.5rem;">
                            {ticket.status}
                        </span>
                    </td>
                    <td class="text-center">
                        <a href="/ticket/{ticket.id}/" class="btn-view" style="font-size: 0.6rem; padding: 0.15rem 0.6rem;">
                            <i class="fa-solid fa-eye"></i> View
                        </a>
                    </td>
                </tr>
        """
    
    html += """
            </tbody>
        </table>
    </div>
    """
    
    return html


def generate_admin_ticket_list_html(tickets, status):
    if not tickets:
        return """
        <div class="empty-state text-center py-5">
            <i class="fa-solid fa-receipt fa-3x mb-3 d-block opacity-25" style="color: var(--accent);"></i>
            <h6 style="color: var(--text-secondary); font-weight: 600;">No Tickets Found</h6>
            <p style="color: var(--text-muted); font-size: 0.85rem;">
                No tickets found.
            </p>
        </div>
        """
    
    html = """
    <div class="table-responsive">
        <table class="table table-hover align-middle mb-0">
            <thead>
                <tr>
                    <th style="font-size: 0.6rem; text-transform: uppercase; color: var(--text-secondary);">Ticket</th>
                    <th style="font-size: 0.6rem; text-transform: uppercase; color: var(--text-secondary);">Created</th>
                    <th style="font-size: 0.6rem; text-transform: uppercase; color: var(--text-secondary);">Unit</th>
                    <th style="font-size: 0.6rem; text-transform: uppercase; color: var(--text-secondary);">Subject</th>
                    <th style="font-size: 0.6rem; text-transform: uppercase; color: var(--text-secondary);">Status</th>
                    <th style="font-size: 0.6rem; text-transform: uppercase; color: var(--text-secondary);">Priority</th>
                    <th class="text-center" style="font-size: 0.6rem; text-transform: uppercase; color: var(--text-secondary);">Action</th>
                </tr>
            </thead>
            <tbody>
    """
    
    badge_class_map = {
        'open': 'badge-status-open',
        'assigned': 'badge-status-assigned',
        'hold': 'badge-status-hold',
        'escalated': 'badge-status-escalated',
        'closed': 'badge-status-closed',
    }
    
    priority_class_map = {
        'critical': 'badge-priority-critical',
        'high': 'badge-priority-high',
        'medium': 'badge-priority-medium',
        'low': 'badge-priority-low',
    }
    
    for ticket in tickets:
        status_class = ticket.status.lower()
        badge_class = badge_class_map.get(status_class, 'badge-status-open')
        priority_class = priority_class_map.get(ticket.priority.lower(), 'badge-priority-low')
        
        html += f"""
                <tr>
                    <td class="fw-bold" style="color: var(--accent-light); font-size: 0.7rem;">
                        {ticket.ticket_number}
                    </td>
                    <td style="font-size: 0.65rem; color: var(--text-secondary);">
                        {ticket.created_at.strftime('%d-%m-%Y') if ticket.created_at else '-'}
                        <span class="d-block" style="font-size: 0.55rem; color: var(--text-muted);">{ticket.created_at.strftime('%I:%M %p') if ticket.created_at else ''}</span>
                    </td>
                    <td style="font-size: 0.65rem; color: var(--text-secondary);">
                        <span class="unit-badge" style="background: rgba(233,69,96,0.1); color: var(--accent); border: 1px solid rgba(233,69,96,0.15); padding: 0.15rem 0.5rem; border-radius: 50px; font-size: 0.6rem; font-weight: 600;">{ticket.unit.code if ticket.unit else '-'}</span>
                    </td>
                    <td class="text-truncate" style="max-width: 100px; font-size: 0.7rem;" title="{ticket.subject}">
                        {ticket.subject}
                    </td>
                    <td>
                        <span class="badge-custom {badge_class}" style="font-size: 0.5rem; padding: 0.15rem 0.5rem;">
                            {ticket.status}
                        </span>
                    </td>
                    <td>
                        <span class="badge-priority {priority_class}" style="font-size: 0.5rem; padding: 0.15rem 0.5rem;">
                            <i class="fa-solid fa-flag" style="font-size: 0.3rem;"></i>
                            {ticket.priority}
                        </span>
                    </td>
                    <td class="text-center">
                        <a href="/admin/ticket/{ticket.id}/" class="btn-view" style="font-size: 0.6rem; padding: 0.15rem 0.6rem; background: var(--accent-gradient); color: white; border-radius: 50px; text-decoration: none; display: inline-block;">
                            <i class="fa-solid fa-eye"></i> View
                        </a>
                    </td>
                </tr>
        """
    
    html += """
            </tbody>
        </table>
    </div>
    """
    
    return html


class CustomLoginView(LoginView):
    template_name = 'login.html'
    redirect_authenticated_user = True
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contact = AdminContact.objects.first()
        context['contact'] = contact
        return context
    
    def get_success_url(self):
        if self.request.user.is_staff:
            return reverse_lazy('admin_dashboard')
        return reverse_lazy('employee_dashboard')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f"Welcome back, {self.request.user.username}!")
        return response
    
    def form_invalid(self, form):
        messages.error(self.request, "Invalid username or password. Please try again.")
        return super().form_invalid(form)
    
    @never_cache
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            if request.user.is_staff:
                return redirect('admin_dashboard')
            return redirect('employee_dashboard')
        return super().dispatch(request, *args, **kwargs)


@login_required
def role_redirect(request):
    if request.user.is_staff:
        return redirect('admin_dashboard')
    return redirect('employee_dashboard')


def custom_logout(request):
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect('login')


@login_required
@user_passes_test(lambda u: not u.is_staff, login_url='login')
def employee_dashboard(request):
    user_credential = DepartmentCredential.objects.filter(
        username=request.user.username, is_active=True
    ).first()
    
    contact = AdminContact.objects.first()
    
    if user_credential:
        department_tickets = Ticket.objects.filter(
            unit=user_credential.unit,
            department=user_credential.department
        )
        personal_tickets = Ticket.objects.filter(created_by_user=request.user)
        tickets_qs = department_tickets
        
        kpis = {
            'total': tickets_qs.count(),
            'open': tickets_qs.filter(status='Open').count(),
            'assigned': tickets_qs.filter(status='Assigned').count(),
            'hold': tickets_qs.filter(status='Hold').count(),
            'escalated': tickets_qs.filter(status='Escalated').count(),
            'closed': tickets_qs.filter(status='Closed').count(),
            'critical': tickets_qs.filter(priority='Critical').count(),
            'my_total': personal_tickets.count(),
            'my_open': personal_tickets.filter(status='Open').count(),
            'my_closed': personal_tickets.filter(status='Closed').count(),
        }
        
        latest_tickets = tickets_qs.order_by('-created_at')[:5]
        
        dept_status_counts = department_tickets.values('status').annotate(count=Count('id'))
        chart_dept_status = {item['status']: item['count'] for item in dept_status_counts}
        
        dept_priority_counts = department_tickets.values('priority').annotate(count=Count('id'))
        chart_dept_priority = {item['priority']: item['count'] for item in dept_priority_counts}
        
        charts_data = {
            'dept_status': chart_dept_status,
            'dept_priority': chart_dept_priority,
        }
        
        context = {
            'kpis': kpis,
            'latest_tickets': latest_tickets,
            'contact': contact,
            'user_credential': user_credential,
            'department_name': user_credential.department.name if user_credential else None,
            'unit_name': user_credential.unit.code if user_credential else None,
            'show_department_tickets': True,
            'charts_data': charts_data,
        }
    else:
        user_tickets = Ticket.objects.filter(created_by_user=request.user)
        kpis = {
            'total': user_tickets.count(),
            'open': user_tickets.filter(status='Open').count(),
            'assigned': user_tickets.filter(status='Assigned').count(),
            'hold': user_tickets.filter(status='Hold').count(),
            'escalated': user_tickets.filter(status='Escalated').count(),
            'closed': user_tickets.filter(status='Closed').count(),
            'critical': user_tickets.filter(priority='Critical').count(),
            'my_total': user_tickets.count(),
            'my_open': user_tickets.filter(status='Open').count(),
            'my_closed': user_tickets.filter(status='Closed').count(),
        }
        latest_tickets = user_tickets.order_by('-created_at')[:5]
        
        personal_status_counts = user_tickets.values('status').annotate(count=Count('id'))
        chart_dept_status = {item['status']: item['count'] for item in personal_status_counts}
        
        personal_priority_counts = user_tickets.values('priority').annotate(count=Count('id'))
        chart_dept_priority = {item['priority']: item['count'] for item in personal_priority_counts}
        
        charts_data = {
            'dept_status': chart_dept_status,
            'dept_priority': chart_dept_priority,
        }
        
        context = {
            'kpis': kpis,
            'latest_tickets': latest_tickets,
            'contact': contact,
            'user_credential': None,
            'department_name': None,
            'unit_name': None,
            'show_department_tickets': False,
            'charts_data': charts_data,
        }
    
    return render(request, 'employee/dashboard.html', context)


@login_required
@user_passes_test(lambda u: not u.is_staff, login_url='login')
def create_ticket(request):
    user_credential = DepartmentCredential.objects.filter(
        username=request.user.username, is_active=True
    ).first()
    
    if not user_credential:
        messages.error(request, "Your account is not associated with any department. Please contact admin.")
        return redirect('employee_dashboard')
    
    if request.method == 'POST':
        form = TicketForm(request.POST, request.FILES)
        if form.is_valid():
            employee_id = form.cleaned_data.get('employee_id', '').strip().upper()
            
            if employee_id:
                try:
                    employee = EmployeeMaster.objects.get(employee_id=employee_id, is_active=True)
                    
                    if employee.unit_id != user_credential.unit_id:
                        messages.error(request, f'Employee "{employee_id}" belongs to {employee.unit.code if employee.unit else "Unknown"} unit. You can only create tickets for {user_credential.unit.code} unit employees.')
                        return render(request, 'employee/create_ticket.html', {
                            'form': form, 
                            'user_credential': user_credential,
                            'validation_error': True
                        })
                    
                    if employee.department_id != user_credential.department_id:
                        messages.error(request, f'Employee "{employee_id}" belongs to {employee.department.name if employee.department else "Unknown"} department. You can only create tickets for {user_credential.department.name} department employees.')
                        return render(request, 'employee/create_ticket.html', {
                            'form': form, 
                            'user_credential': user_credential,
                            'validation_error': True
                        })
                    
                except EmployeeMaster.DoesNotExist:
                    messages.error(request, f'Employee "{employee_id}" not found. Please check the Employee ID.')
                    return render(request, 'employee/create_ticket.html', {
                        'form': form, 
                        'user_credential': user_credential,
                        'validation_error': True
                    })
                except EmployeeMaster.MultipleObjectsReturned:
                    messages.error(request, f'Multiple employees found with ID "{employee_id}". Please contact admin.')
                    return render(request, 'employee/create_ticket.html', {
                        'form': form, 
                        'user_credential': user_credential,
                        'validation_error': True
                    })
            
            with transaction.atomic():
                ticket = form.save(commit=False)
                ticket.created_by_user = request.user
                ticket.created_by_role = 'Employee'
                ticket.status = 'Open'
                ticket.save()
                TicketHistory.objects.create(
                    ticket=ticket, 
                    action="Ticket Created",
                    remarks="Ticket Created by Employee", 
                    performed_by=f"Employee {request.user.username}"
                )
            send_ticket_email(ticket, 'Created')
            messages.success(request, f'Ticket {ticket.ticket_number} created successfully!')
            return redirect('employee_dashboard')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = TicketForm()
        if user_credential:
            form.fields['unit'].initial = user_credential.unit_id
            form.fields['department'].initial = user_credential.department_id
    
    return render(request, 'employee/create_ticket.html', {
        'form': form, 
        'user_credential': user_credential,
        'unit_name': user_credential.unit.code if user_credential else None,
        'department_name': user_credential.department.name if user_credential else None,
    })


@login_required
@user_passes_test(lambda u: not u.is_staff, login_url='login')
def my_tickets(request):
    status = request.GET.get('status')
    priority = request.GET.get('priority')
    ticket_number = request.GET.get('ticket_number', '').strip()
    search = request.GET.get('search', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    is_ajax = request.GET.get('ajax', False)
    filter_type = request.GET.get('filter_type')
    
    if isinstance(is_ajax, str):
        is_ajax = is_ajax.lower() in ['true', '1', 'yes']
    
    user_credential = DepartmentCredential.objects.filter(
        username=request.user.username, is_active=True
    ).first()
    
    if user_credential:
        tickets_qs = Ticket.objects.filter(
            unit=user_credential.unit,
            department=user_credential.department
        ).order_by('-created_at')
        personal_tickets = Ticket.objects.filter(created_by_user=request.user)
        tickets_qs = tickets_qs | personal_tickets
        tickets_qs = tickets_qs.distinct().order_by('-created_at')
    else:
        tickets_qs = Ticket.objects.filter(created_by_user=request.user).order_by('-created_at')
    
    if status: 
        tickets_qs = tickets_qs.filter(status=status)
    if priority: 
        tickets_qs = tickets_qs.filter(priority=priority)
    if ticket_number: 
        tickets_qs = tickets_qs.filter(ticket_number__icontains=ticket_number)
    if search: 
        tickets_qs = tickets_qs.filter(
            Q(ticket_number__icontains=search) | 
            Q(subject__icontains=search) | 
            Q(description__icontains=search)
        )
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            from_datetime = timezone.make_aware(
                datetime.combine(date_from_obj, datetime.min.time())
            )
            tickets_qs = tickets_qs.filter(created_at__gte=from_datetime)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            to_datetime = timezone.make_aware(
                datetime.combine(date_to_obj, datetime.max.time())
            )
            tickets_qs = tickets_qs.filter(created_at__lte=to_datetime)
        except ValueError:
            pass
    
    if is_ajax:
        filter_value = status or priority or 'All'
        if filter_type == 'priority':
            filter_value = priority
        elif filter_type == 'status':
            filter_value = status

        try:
            try:
                html = render_to_string('employee/_ticket_list_modal.html', {
                    'tickets': tickets_qs[:50],
                    'status_label': filter_value,
                }, request=request)
                return JsonResponse({'html': html, 'success': True, 'count': tickets_qs.count()})
            except TemplateDoesNotExist:
                html = generate_ticket_list_html(tickets_qs[:50], status or priority)
                return JsonResponse({'html': html, 'success': True, 'count': tickets_qs.count()})
            except Exception as e:
                logger.error(f"Template rendering error in my_tickets: {str(e)}")
                html = generate_ticket_list_html(tickets_qs[:50], status or priority)
                return JsonResponse({'html': html, 'success': True, 'count': tickets_qs.count()})
        except Exception as e:
            logger.error(f"AJAX error in my_tickets: {str(e)}")
            return JsonResponse({
                'success': False, 
                'error': str(e),
                'message': 'Error loading tickets. Please try again.'
            }, status=500)
    
    paginator = Paginator(tickets_qs, 10)
    page_number = request.GET.get('page')
    try: 
        tickets_page = paginator.page(page_number)
    except PageNotAnInteger: 
        tickets_page = paginator.page(1)
    except EmptyPage: 
        tickets_page = paginator.page(paginator.num_pages)
    
    return render(request, 'employee/my_tickets.html', {
        'tickets': tickets_page, 
        'status_choices': Ticket.STATUS_CHOICES,
        'priority_choices': Ticket.PRIORITY_CHOICES, 
        'selected_status': status,
        'selected_priority': priority, 
        'selected_ticket_number': ticket_number,
        'search_query': search, 
        'date_from': date_from, 
        'date_to': date_to,
    })


@login_required
@user_passes_test(lambda u: not u.is_staff, login_url='login')
def ticket_detail(request, pk):
    user_credential = DepartmentCredential.objects.filter(
        username=request.user.username, is_active=True
    ).first()
    
    if user_credential:
        tickets_qs = Ticket.objects.filter(
            Q(created_by_user=request.user) |
            Q(unit=user_credential.unit, department=user_credential.department)
        ).distinct()
    else:
        tickets_qs = Ticket.objects.filter(created_by_user=request.user)
    
    ticket = get_object_or_404(tickets_qs, pk=pk)
    
    history = ticket.history.all().order_by('timestamp')
    can_reopen = False
    reopen_time_left = None
    reopen_deadline = None
    
    if ticket.status == 'Closed' and ticket.closed_at:
        reopen_deadline = ticket.closed_at + timedelta(hours=48)
        if timezone.now() < reopen_deadline:
            can_reopen = True
            reopen_time_left = reopen_deadline - timezone.now()
    
    time_to_close_str = ""
    if ticket.status == 'Closed' and ticket.created_at and ticket.closed_at:
        time_to_close = ticket.closed_at - ticket.created_at
        time_to_close_str = format_timedelta_display(time_to_close)
    
    if request.method == 'POST' and can_reopen:
        remarks = request.POST.get('remarks', '').strip()
        if remarks:
            reopen_ticket_logic(ticket, f"Employee {request.user.username}", remarks)
            messages.success(request, f"Ticket {ticket.ticket_number} has been reopened.")
            return redirect('ticket_detail', pk=pk)
        else:
            messages.error(request, "Please provide a reason for reopening.")
    
    context = {
        'ticket': ticket,
        'history': history,
        'can_reopen': can_reopen,
        'time_to_close': time_to_close_str,
        'reopen_time_left': reopen_time_left,
        'reopen_deadline_iso': reopen_deadline.isoformat() if reopen_deadline else None,
    }
    return render(request, 'employee/ticket_detail.html', context)


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def admin_dashboard(request):
    all_tickets = Ticket.objects.all()
    
    kpis = {
        'total': all_tickets.count(), 
        'open': all_tickets.filter(status='Open').count(),
        'assigned': all_tickets.filter(status='Assigned').count(), 
        'hold': all_tickets.filter(status='Hold').count(),
        'escalated': all_tickets.filter(status='Escalated').count(), 
        'closed': all_tickets.filter(status='Closed').count(),
        'critical': all_tickets.filter(priority='Critical').count(),
    }
    
    status_counts = list(all_tickets.values('status').annotate(count=Count('id')))
    chart_status = {item['status']: item['count'] for item in status_counts}
    
    unit_counts = list(all_tickets.filter(unit__isnull=False).values('unit_id', 'unit__code').annotate(count=Count('id')).order_by('unit__code'))
    chart_units = [{'id': item['unit_id'], 'label': item['unit__code'], 'value': item['count']} for item in unit_counts]
    
    prio_counts = list(all_tickets.values('priority').annotate(count=Count('id')))
    chart_priority = {item['priority']: item['count'] for item in prio_counts}
    
    closed_tickets = Ticket.objects.filter(status='Closed')
    error_type_counts = (
        closed_tickets
        .exclude(error_type__isnull=True)
        .exclude(error_type__exact='')
        .values('error_type')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    chart_error_type = {item['error_type']: item['count'] for item in error_type_counts}
    
    twelve_months_ago = timezone.now() - timedelta(days=365)
    monthly_counts_qs = Ticket.objects.filter(created_at__gte=twelve_months_ago).annotate(month=TruncMonth(Cast('created_at', output_field=DateField()))).values('month').annotate(count=Count('id')).order_by('month')
    chart_monthly = [{'label': item['month'].strftime('%b %Y'), 'value': item['count']} for item in monthly_counts_qs]
    
    charts_data = {
        'status': chart_status,
        'units': chart_units,
        'priority': chart_priority,
        'errorType': chart_error_type,
        'monthly': chart_monthly,
    }
    
    context = {'kpis': kpis, 'charts_data': charts_data}
    return render(request, 'admin_panel/dashboard.html', context)


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def create_ticket_admin(request):
    if request.method == 'POST':
        form = AdminTicketForm(request.POST, request.FILES)
        if form.is_valid():
            with transaction.atomic():
                ticket = form.save(commit=False)
                if ticket.created_by_role == 'Admin':
                    hist_remark = f"Ticket created by Admin on behalf of employee (Reason: {ticket.admin_creation_reason})"
                    hist_perf = f"Admin {request.user.username}"
                    ticket.created_by_user = request.user
                else:
                    hist_remark = "Ticket Created by Employee (logged by Admin)"
                    hist_perf = "Employee"
                    employee_user, _ = User.objects.get_or_create(username='GPLERPUSERS', defaults={'is_staff': False, 'password': 'pbkdf2_sha256$720000$j5xL6pS0LpGvLq3sRjVbWk$V/Hq7aYt2x531enqYm5d9f2uZdtsJ7MLd2y221C+L9s='})
                    ticket.created_by_user = employee_user
                ticket.save()
                TicketHistory.objects.create(ticket=ticket, action="Ticket Created", remarks=hist_remark, performed_by=hist_perf)
            send_ticket_email(ticket, 'Created')
            messages.success(request, f'Ticket {ticket.ticket_number} created successfully by Admin!')
            return redirect('admin_dashboard')
    else:
        form = AdminTicketForm()
    return render(request, 'admin_panel/create_ticket.html', {'form': form})


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def all_tickets(request):
    is_ajax = request.GET.get('ajax', False)
    
    if isinstance(is_ajax, str):
        is_ajax = is_ajax.lower() in ['true', '1', 'yes']
    
    tickets_qs = Ticket.objects.all().order_by('-created_at')
    units = Unit.objects.filter(is_active=True)
    departments = Department.objects.filter(is_active=True)
    category = request.GET.get('category', 'all')
    status = request.GET.get('status')
    priority = request.GET.get('priority')
    unit_id = request.GET.get('unit')
    dept_id = request.GET.get('department')
    assigned_person = request.GET.get('assigned_person', '').strip()
    created_by_role = request.GET.get('created_by_role')
    ticket_number = request.GET.get('ticket_number', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    search = request.GET.get('search', '').strip()
    
    if category == 'open': 
        tickets_qs = tickets_qs.filter(status='Open')
    elif category == 'assigned': 
        tickets_qs = tickets_qs.filter(status='Assigned')
    elif category == 'hold': 
        tickets_qs = tickets_qs.filter(status='Hold')
    elif category == 'escalated': 
        tickets_qs = tickets_qs.filter(status='Escalated')
    elif category == 'closed': 
        tickets_qs = tickets_qs.filter(status='Closed')
    elif category == 'critical': 
        tickets_qs = tickets_qs.filter(priority='Critical')
    
    if unit_id: 
        tickets_qs = tickets_qs.filter(unit_id=unit_id)
    if dept_id: 
        tickets_qs = tickets_qs.filter(department_id=dept_id)
    if status: 
        tickets_qs = tickets_qs.filter(status=status)
    if priority: 
        tickets_qs = tickets_qs.filter(priority=priority)
    if assigned_person: 
        tickets_qs = tickets_qs.filter(assigned_person__icontains=assigned_person)
    if created_by_role: 
        tickets_qs = tickets_qs.filter(created_by_role=created_by_role)
    if ticket_number: 
        tickets_qs = tickets_qs.filter(ticket_number__icontains=ticket_number)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            from_datetime = timezone.make_aware(
                datetime.combine(date_from_obj, datetime.min.time())
            )
            tickets_qs = tickets_qs.filter(created_at__gte=from_datetime)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            to_datetime = timezone.make_aware(
                datetime.combine(date_to_obj, datetime.max.time())
            )
            tickets_qs = tickets_qs.filter(created_at__lte=to_datetime)
        except ValueError:
            pass
    
    if search: 
        tickets_qs = tickets_qs.filter(
            Q(ticket_number__icontains=search) | 
            Q(subject__icontains=search) | 
            Q(unit__code__icontains=search) | 
            Q(department__name__icontains=search)
        )
    
    if is_ajax:
        filter_value = 'All'
        filter_type = 'status'
        
        if status:
            filter_value = status
            filter_type = 'status'
        elif priority:
            filter_value = priority
            filter_type = 'priority'
        elif unit_id:
            unit = Unit.objects.filter(pk=unit_id).first()
            filter_value = unit.code if unit else 'Unit'
            filter_type = 'unit'
        elif category and category != 'all':
            filter_value = category.capitalize()
            filter_type = 'status'

        try:
            try:
                html = render_to_string('admin_panel/_ticket_list_modal.html', {
                    'tickets': tickets_qs[:50],
                    'status_label': filter_value,
                    'filter_type': filter_type,
                    'filter_value': filter_value,
                }, request=request)
                return JsonResponse({'html': html, 'success': True, 'count': tickets_qs.count()})
            except TemplateDoesNotExist:
                html = generate_admin_ticket_list_html(tickets_qs[:50], filter_value)
                return JsonResponse({'html': html, 'success': True, 'count': tickets_qs.count()})
            except Exception as e:
                logger.error(f"AJAX error in all_tickets (template): {str(e)}")
                html = generate_admin_ticket_list_html(tickets_qs[:50], filter_value)
                return JsonResponse({'html': html, 'success': True, 'count': tickets_qs.count()})
        except Exception as e:
            logger.error(f"AJAX error in all_tickets: {str(e)}")
            return JsonResponse({
                'success': False, 
                'error': str(e),
                'message': 'Error loading tickets. Please try again.'
            }, status=500)
    
    paginator = Paginator(tickets_qs, 10)
    page_number = request.GET.get('page')
    try: 
        tickets_page = paginator.page(page_number)
    except PageNotAnInteger: 
        tickets_page = paginator.page(1)
    except EmptyPage: 
        tickets_page = paginator.page(paginator.num_pages)
    
    context = {
        'tickets': tickets_page, 
        'units': units, 
        'departments': departments,
        'status_choices': Ticket.STATUS_CHOICES, 
        'priority_choices': Ticket.PRIORITY_CHOICES,
        'created_by_choices': Ticket.CREATED_BY_CHOICES, 
        'selected_status': status,
        'selected_priority': priority, 
        'selected_unit': unit_id, 
        'selected_department': dept_id,
        'selected_assigned_person': assigned_person, 
        'selected_created_by_role': created_by_role,
        'selected_ticket_number': ticket_number, 
        'date_from': date_from, 
        'date_to': date_to,
        'search_query': search, 
        'category': category,
    }
    return render(request, 'admin_panel/all_tickets.html', context)


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def ticket_detail_admin(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    history = ticket.history.all().order_by('timestamp')
    can_reopen = False
    reopen_time_left = None
    reopen_deadline = None
    if ticket.status == 'Closed' and ticket.closed_at:
        reopen_deadline = ticket.closed_at + timedelta(hours=48)
        if timezone.now() < reopen_deadline: 
            can_reopen = True
            reopen_time_left = reopen_deadline - timezone.now()
    time_to_close_str = ""
    if ticket.status == 'Closed' and ticket.created_at and ticket.closed_at:
        time_to_close = ticket.closed_at - ticket.created_at
        time_to_close_str = format_timedelta_display(time_to_close)
    
    if request.method == 'POST':
        action_type = request.POST.get('action_type')
        remarks = request.POST.get('remarks', '')
        
        with transaction.atomic():
            if action_type == 'Assign':
                assigned_person = request.POST.get('assigned_person', '').strip()
                if not assigned_person: 
                    messages.error(request, "Assigned Person Name is mandatory.")
                    return redirect('admin_ticket_detail', pk=pk)
                ticket.status = 'Assigned'
                ticket.assigned_person = assigned_person
                ticket.save()
                TicketHistory.objects.create(
                    ticket=ticket, 
                    action=f"Assigned to {assigned_person}", 
                    remarks=remarks, 
                    performed_by=f"Admin {request.user.username}"
                )
                messages.success(request, f'Ticket assigned to {assigned_person}.')
                
            elif action_type == 'Hold':
                hold_reason = request.POST.get('hold_reason', '').strip()
                if not hold_reason: 
                    messages.error(request, "Hold Reason is mandatory.")
                    return redirect('admin_ticket_detail', pk=pk)
                ticket.status = 'Hold'
                ticket.hold_reason = hold_reason
                ticket.save()
                TicketHistory.objects.create(
                    ticket=ticket, 
                    action="Status changed to Hold", 
                    remarks=f"Reason: {hold_reason}", 
                    performed_by=f"Admin {request.user.username}"
                )
                messages.success(request, 'Ticket placed on Hold.')
                
            elif action_type == 'Escalate':
                vendor_ticket = request.POST.get('vendor_ticket_number', '').strip()
                ticket.status = 'Escalated'
                if vendor_ticket: 
                    ticket.vendor_ticket_number = vendor_ticket
                ticket.escalated_at = timezone.now()
                ticket.save()
                remark_str = f"Vendor Ticket: {vendor_ticket}" if vendor_ticket else "Escalated without vendor ticket number"
                TicketHistory.objects.create(
                    ticket=ticket, 
                    action="Escalated to ERP Vendor", 
                    remarks=remark_str, 
                    performed_by=f"Admin {request.user.username}"
                )
                messages.success(request, 'Ticket escalated to ERP vendor.')
                
            elif action_type == 'Close':
                closing_remarks = request.POST.get('closing_remarks', '').strip()
                error_type = request.POST.get('error_type', '').strip()
                if not closing_remarks: 
                    messages.error(request, "Closing Remarks are mandatory.")
                    return redirect('admin_ticket_detail', pk=pk)
                if not error_type: 
                    messages.error(request, "Error Classification is mandatory.")
                    return redirect('admin_ticket_detail', pk=pk)
                ticket.status = 'Closed'
                ticket.closing_remarks = closing_remarks
                ticket.error_type = error_type
                ticket.closed_by = request.user.username
                ticket.closed_at = timezone.now()
                ticket.save()
                TicketHistory.objects.create(
                    ticket=ticket, 
                    action=f"Closed by {request.user.username}", 
                    remarks=f"Error Type: {error_type} | {closing_remarks}", 
                    performed_by=f"Admin {request.user.username}"
                )
                send_ticket_email(ticket, 'Closed', remarks=closing_remarks)
                messages.success(request, 'Ticket closed successfully.')
                
            elif action_type == 'Reopen':
                if not can_reopen: 
                    messages.error(request, "Cannot reopen - 48 hours elapsed.")
                    return redirect('admin_ticket_detail', pk=pk)
                remarks = request.POST.get('remarks', '').strip()
                if not remarks: 
                    messages.error(request, "Reason for reopening is mandatory.")
                    return redirect('admin_ticket_detail', pk=pk)
                reopen_ticket_logic(ticket, f"Admin {request.user.username}", remarks)
                messages.success(request, 'Ticket reopened successfully.')
                
        return redirect('admin_ticket_detail', pk=pk)
    
    context = {
        'ticket': ticket,
        'history': history,
        'can_reopen': can_reopen,
        'time_to_close': time_to_close_str,
        'reopen_time_left': reopen_time_left,
        'reopen_deadline_iso': reopen_deadline.isoformat() if reopen_deadline else None,
    }
    return render(request, 'admin_panel/ticket_detail.html', context)


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def reports(request):
    tickets_qs = Ticket.objects.all().order_by('-created_at')
    units = Unit.objects.all()
    departments = Department.objects.all()
    
    unit_id = request.GET.get('unit', '').strip()
    dept_id = request.GET.get('department', '').strip()
    priority = request.GET.get('priority', '').strip()
    status = request.GET.get('status', '').strip()
    assigned_person = request.GET.get('assigned_person', '').strip()
    created_by_role = request.GET.get('created_by_role', '').strip()
    error_type = request.GET.get('error_type', '').strip()
    vendor_ticket = request.GET.get('vendor_ticket_number', '').strip()
    
    created_start = request.GET.get('created_start', '').strip()
    created_end = request.GET.get('created_end', '').strip()
    closed_start = request.GET.get('closed_start', '').strip()
    closed_end = request.GET.get('closed_end', '').strip()
    escalated_start = request.GET.get('escalated_start', '').strip()
    escalated_end = request.GET.get('escalated_end', '').strip()
    
    category = request.GET.get('category', 'all').strip()
    is_reopened = request.GET.get('is_reopened', '').strip()
    
    if category == 'open': 
        tickets_qs = tickets_qs.filter(status='Open')
    elif category == 'assigned': 
        tickets_qs = tickets_qs.filter(status='Assigned')
    elif category == 'hold': 
        tickets_qs = tickets_qs.filter(status='Hold')
    elif category == 'escalated': 
        tickets_qs = tickets_qs.filter(status='Escalated')
    elif category == 'closed': 
        tickets_qs = tickets_qs.filter(status='Closed')
    elif category == 'escalated_closed': 
        tickets_qs = tickets_qs.filter(status='Closed', escalated_at__isnull=False)
    elif category == 'reopened': 
        tickets_qs = tickets_qs.filter(history__action='Ticket Reopened').distinct()
    
    if unit_id and unit_id != '': 
        tickets_qs = tickets_qs.filter(unit_id=unit_id)
    if dept_id and dept_id != '': 
        tickets_qs = tickets_qs.filter(department_id=dept_id)
    if priority and priority != '': 
        tickets_qs = tickets_qs.filter(priority=priority)
    if status and status != '': 
        tickets_qs = tickets_qs.filter(status=status)
    if assigned_person and assigned_person != '': 
        tickets_qs = tickets_qs.filter(assigned_person__icontains=assigned_person)
    if created_by_role and created_by_role != '': 
        tickets_qs = tickets_qs.filter(created_by_role=created_by_role)
    if error_type and error_type != '': 
        tickets_qs = tickets_qs.filter(error_type=error_type)
    if vendor_ticket and vendor_ticket != '': 
        tickets_qs = tickets_qs.filter(vendor_ticket_number__icontains=vendor_ticket)
    
    if created_start and created_start != '':
        try:
            created_start_date = datetime.strptime(created_start, '%Y-%m-%d').date()
            from_datetime = timezone.make_aware(
                datetime.combine(created_start_date, datetime.min.time())
            )
            tickets_qs = tickets_qs.filter(created_at__gte=from_datetime)
        except ValueError:
            pass
    
    if created_end and created_end != '':
        try:
            created_end_date = datetime.strptime(created_end, '%Y-%m-%d').date()
            to_datetime = timezone.make_aware(
                datetime.combine(created_end_date, datetime.max.time())
            )
            tickets_qs = tickets_qs.filter(created_at__lte=to_datetime)
        except ValueError:
            pass
    
    if closed_start and closed_start != '':
        try:
            closed_start_date = datetime.strptime(closed_start, '%Y-%m-%d').date()
            from_datetime = timezone.make_aware(
                datetime.combine(closed_start_date, datetime.min.time())
            )
            tickets_qs = tickets_qs.filter(closed_at__gte=from_datetime)
        except ValueError:
            pass
    
    if closed_end and closed_end != '':
        try:
            closed_end_date = datetime.strptime(closed_end, '%Y-%m-%d').date()
            to_datetime = timezone.make_aware(
                datetime.combine(closed_end_date, datetime.max.time())
            )
            tickets_qs = tickets_qs.filter(closed_at__lte=to_datetime)
        except ValueError:
            pass
    
    if escalated_start and escalated_start != '':
        try:
            escalated_start_date = datetime.strptime(escalated_start, '%Y-%m-%d').date()
            from_datetime = timezone.make_aware(
                datetime.combine(escalated_start_date, datetime.min.time())
            )
            tickets_qs = tickets_qs.filter(escalated_at__gte=from_datetime)
        except ValueError:
            pass
    
    if escalated_end and escalated_end != '':
        try:
            escalated_end_date = datetime.strptime(escalated_end, '%Y-%m-%d').date()
            to_datetime = timezone.make_aware(
                datetime.combine(escalated_end_date, datetime.max.time())
            )
            tickets_qs = tickets_qs.filter(escalated_at__lte=to_datetime)
        except ValueError:
            pass
    
    if is_reopened == 'yes': 
        tickets_qs = tickets_qs.filter(history__action='Ticket Reopened').distinct()
    elif is_reopened == 'no': 
        tickets_qs = tickets_qs.exclude(history__action='Ticket Reopened')
    
    total_count = tickets_qs.count()
    open_count = tickets_qs.filter(status='Open').count()
    assigned_count = tickets_qs.filter(status='Assigned').count()
    hold_count = tickets_qs.filter(status='Hold').count()
    escalated_count = tickets_qs.filter(status='Escalated').count()
    closed_count = tickets_qs.filter(status='Closed').count()
    
    if 'export' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=GPLAST_Report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tickets Report"
        
        title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        data_font = Font(name='Calibri', size=11)
        title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        
        ws.merge_cells('A1:Y1')
        ws['A1'] = "GPLAST Ticket Report"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40
        
        headers = ["Ticket Number","Status","Unit Code","Unit Full Name","Department","Employee ID","Employee Name","Mobile","Email","Screen Number","Subject","Description","Priority","Error Type","Created By Role","Time to Close","Admin Reason","Assigned Person","Hold Reason","Closing Remarks","Closed By","Vendor Ticket","Created At","Closed At","Escalated At"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[3].height = 25
        
        row_idx = 4
        for t in tickets_qs:
            c_at = t.created_at.strftime('%d-%b-%Y %I:%M %p') if t.created_at else ""
            cl_at = t.closed_at.strftime('%d-%b-%Y %I:%M %p') if t.closed_at else ""
            esc_at = t.escalated_at.strftime('%d-%b-%Y %I:%M %p') if t.escalated_at else ""
            ttc = format_timedelta_display(t.closed_at - t.created_at) if t.status == 'Closed' and t.created_at and t.closed_at else ""
            
            row_data = [
                t.ticket_number,
                t.status,
                t.unit.code if t.unit else '',
                t.unit.full_name if t.unit else '',
                t.department.name if t.department else '',
                t.employee_id,
                t.employee_name,
                t.mobile,
                t.email,
                t.screen_number,
                t.subject,
                t.description,
                t.priority,
                t.error_type,
                t.created_by_role,
                ttc,
                t.admin_creation_reason or '',
                t.assigned_person or '',
                t.hold_reason or '',
                t.closing_remarks or '',
                t.closed_by or '',
                t.vendor_ticket_number or '',
                c_at,
                cl_at,
                esc_at
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.font = data_font
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            row_idx += 1
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col if cell.row > 1)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 3, 12)
            
        wb.save(response)
        return response
    
    paginator = Paginator(tickets_qs, 15)
    page_number = request.GET.get('page')
    try: 
        tickets_page = paginator.page(page_number)
    except PageNotAnInteger: 
        tickets_page = paginator.page(1)
    except EmptyPage: 
        tickets_page = paginator.page(paginator.num_pages)
    
    error_type_choices = Ticket.objects.filter(
        error_type__isnull=False
    ).exclude(
        error_type__exact=''
    ).values_list('error_type', flat=True).distinct().order_by('error_type')
    
    error_type_choices_list = [(et, et) for et in error_type_choices]
    
    context = {
        'tickets': tickets_page,
        'total_count': total_count,
        'open_count': open_count,
        'assigned_count': assigned_count,
        'hold_count': hold_count,
        'escalated_count': escalated_count,
        'closed_count': closed_count,
        'units': units,
        'departments': departments,
        'category': category,
        'is_reopened': is_reopened,
        'status_choices': Ticket.STATUS_CHOICES,
        'priority_choices': Ticket.PRIORITY_CHOICES,
        'created_by_choices': Ticket.CREATED_BY_CHOICES,
        'error_type_choices': error_type_choices_list,
        'selected_unit': unit_id,
        'selected_department': dept_id,
        'selected_status': status,
        'selected_priority': priority,
        'selected_assigned_person': assigned_person,
        'selected_created_by_role': created_by_role,
        'selected_error_type': error_type,
        'selected_vendor_ticket': vendor_ticket,
        'created_start': created_start,
        'created_end': created_end,
        'closed_start': closed_start,
        'closed_end': closed_end,
        'escalated_start': escalated_start,
        'escalated_end': escalated_end,
    }
    return render(request, 'admin_panel/reports.html', context)


def _get_contact_data():
    contact_obj, _ = AdminContact.objects.get_or_create(
        id=1,
        defaults={'admin_name': "IT ADMIN", 'admin_phone': "9999999999", 'admin_email': "admin@gplast.com"}
    )
    return contact_obj


def _get_employee_directory_data(request):
    emp_search = request.GET.get('emp_search', '').strip()
    employees_qs = EmployeeMaster.objects.all().order_by('employee_id')
    if emp_search:
        employees_qs = employees_qs.filter(
            Q(employee_id__icontains=emp_search) |
            Q(employee_name__icontains=emp_search) |
            Q(email__icontains=emp_search)
        )
    return employees_qs, emp_search


def _get_credentials_data():
    all_credentials = DepartmentCredential.objects.select_related('unit', 'department').order_by('unit__code', 'department__name')
    credentials_by_unit = []
    
    for unit in Unit.objects.filter(is_active=True).order_by('code'):
        unit_creds = all_credentials.filter(unit=unit)
        if unit_creds.exists():
            credentials_by_unit.append({
                'unit': unit, 
                'credentials': unit_creds
            })
    
    for unit in Unit.objects.filter(is_active=False).order_by('code'):
        unit_creds = all_credentials.filter(unit=unit)
        if unit_creds.exists():
            credentials_by_unit.append({
                'unit': unit, 
                'credentials': unit_creds
            })
    
    return all_credentials, credentials_by_unit


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def settings_page(request):
    contact_obj = _get_contact_data()
    employees_qs, emp_search = _get_employee_directory_data(request)
    all_credentials, credentials_by_unit = _get_credentials_data()
    
    context = {
        'contact': contact_obj,
        'contact_form': AdminContactForm(instance=contact_obj),
        'unit_form': UnitForm(),
        'dept_form': DepartmentForm(),
        'email_form': AdminNotificationEmailForm(),
        'units': Unit.objects.all().order_by('code'),
        'departments': Department.objects.all().order_by('unit__code', 'name'),
        'emails': AdminNotificationEmail.objects.all().order_by('-created_at'),
        'employees': employees_qs,
        'emp_search': emp_search,
        'all_units': Unit.objects.filter(is_active=True).order_by('code'),
        'credentials': all_credentials,
        'credentials_by_unit': credentials_by_unit,
        'employee_users': User.objects.filter(is_staff=False, is_active=True).order_by('username'),
    }
    return render(request, 'admin_panel/settings.html', context)


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def settings_contact(request):
    if request.method == 'POST':
        contact_obj = AdminContact.objects.first()
        if not contact_obj:
            contact_obj = AdminContact.objects.create(
                admin_name="IT ADMIN",
                admin_phone="9999999999",
                admin_email="admin@gplast.com"
            )
        
        admin_name = request.POST.get('admin_name', '').strip()
        admin_email = request.POST.get('admin_email', '').strip()
        
        if admin_name:
            contact_obj.admin_name = admin_name
        if admin_email:
            contact_obj.admin_email = admin_email
        
        contact_obj.save()
        messages.success(request, "IT Support Contact updated successfully.")
    return redirect('settings_page')


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def settings_units(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            form = UnitForm(request.POST)
            if form.is_valid(): 
                unit = form.save(commit=False)
                unit.created_by = request.user.username
                unit.save()
                messages.success(request, f"Unit '{unit.code}' added.")
        elif action == 'edit':
            unit = get_object_or_404(Unit, pk=request.POST.get('unit_id'))
            form = UnitForm(request.POST, instance=unit)
            if form.is_valid(): 
                form.save()
                messages.success(request, f"Unit '{unit.code}' updated.")
        elif action == 'toggle':
            unit = get_object_or_404(Unit, pk=request.POST.get('unit_id'))
            unit.is_active = not unit.is_active
            unit.save()
            messages.success(request, f"Unit '{unit.code}' {'activated' if unit.is_active else 'deactivated'}.")
    return redirect('settings_page')


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def settings_departments(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            form = DepartmentForm(request.POST)
            if form.is_valid(): 
                dept = form.save()
                messages.success(request, f"Department '{dept.name}' added.")
        elif action == 'edit':
            dept = get_object_or_404(Department, pk=request.POST.get('dept_id'))
            form = DepartmentForm(request.POST, instance=dept)
            if form.is_valid(): 
                form.save()
                messages.success(request, f"Department '{dept.name}' updated.")
        elif action == 'toggle':
            dept = get_object_or_404(Department, pk=request.POST.get('dept_id'))
            dept.is_active = not dept.is_active
            dept.save()
            messages.success(request, f"Department '{dept.name}' {'activated' if dept.is_active else 'deactivated'}.")
    return redirect('settings_page')


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def settings_emails(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            form = AdminNotificationEmailForm(request.POST)
            if form.is_valid(): 
                email = form.save()
                messages.success(request, f"Email '{email.email}' added.")
        elif action == 'delete':
            email_obj = get_object_or_404(AdminNotificationEmail, pk=request.POST.get('email_id'))
            email_str = email_obj.email
            email_obj.delete()
            messages.success(request, f"Email '{email_str}' deleted.")
    return redirect('settings_page')


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def settings_passwords(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'change_my_password':
            form = AdminPasswordChangeForm(user=request.user, data=request.POST)
            if form.is_valid(): 
                user = form.save()
                update_session_auth_hash(request, user)
                messages.success(request, 'Password updated!')
            else:
                for e in form.errors.values():
                    for err in e: 
                        messages.error(request, f"Error: {err}")
        elif action == 'set_user_password':
            user_id = request.POST.get('user')
            if not user_id:
                messages.error(request, "Please select an employee.")
                return redirect('settings_page')
            selected_user = get_object_or_404(User, pk=user_id, is_staff=False)
            form = AdminSetUserPasswordForm(user=selected_user, data=request.POST)
            if form.is_valid(): 
                form.save()
                messages.success(request, f"Password reset for '{selected_user.username}'.")
            else:
                for e in form.errors.values():
                    for err in e: 
                        messages.error(request, f"Error: {err}")
    return redirect('settings_page')


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def settings_employees(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_employee':
            eid = request.POST.get('employee_id','').strip().upper()
            ename = request.POST.get('employee_name','').strip().upper()
            mob = request.POST.get('mobile','').strip()
            email = request.POST.get('email','').strip()
            uid = request.POST.get('unit')
            did = request.POST.get('department')
            
            if not all([eid,ename,mob,email]): 
                messages.error(request, "All mandatory fields are required.")
                return redirect('settings_page')
            
            try:
                EmployeeMaster.objects.create(
                    employee_id=eid,
                    employee_name=ename,
                    mobile=mob,
                    email=email,
                    unit_id=uid or None,
                    department_id=did or None
                )
                messages.success(request, f'Employee "{eid}" added.')
            except IntegrityError: 
                messages.error(request, f'Employee ID "{eid}" already exists.')
        
        elif action == 'bulk_upload':
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': 'Please select an Excel file.'
                    })
                messages.error(request, "Please select an Excel file.")
                return redirect('settings_page')
            
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': 'Invalid file format. Only .xlsx and .xls files are supported.'
                    })
                messages.error(request, "Invalid file format. Only .xlsx and .xls files are supported.")
                return redirect('settings_page')
            
            try:
                df = pd.read_excel(excel_file, dtype=str)
                
                if df.empty:
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': 'The uploaded file is empty.'
                        })
                    messages.error(request, "The uploaded file is empty.")
                    return redirect('settings_page')
                
                required_columns = ['Employee ID', 'Employee Name', 'Mobile', 'Email']
                missing_columns = []
                for col in required_columns:
                    if col not in df.columns:
                        found = False
                        for existing_col in df.columns:
                            if existing_col.lower() == col.lower():
                                found = True
                                break
                        if not found:
                            missing_columns.append(col)
                
                if missing_columns:
                    error_msg = f"Missing required columns: {', '.join(missing_columns)}"
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': error_msg
                        })
                    messages.error(request, error_msg)
                    return redirect('settings_page')
                
                validation_errors = []
                successful_rows = 0
                
                all_units = {unit.code.upper(): unit for unit in Unit.objects.filter(is_active=True)}
                all_departments = {dept.name.upper(): dept for dept in Department.objects.filter(is_active=True)}
                
                for idx, row in df.iterrows():
                    row_num = idx + 2
                    
                    eid = None
                    ename = None
                    mob = None
                    email = None
                    uc = None
                    dn = None
                    
                    for col in df.columns:
                        col_lower = col.lower()
                        if col_lower in ['employee id', 'employee_id', 'employeeid']:
                            eid = str(row.get(col, '')).strip().upper()
                        elif col_lower in ['employee name', 'employee_name', 'employeename', 'name']:
                            ename = str(row.get(col, '')).strip().upper()
                        elif col_lower in ['mobile', 'phone', 'contact']:
                            mob = str(row.get(col, '')).strip()
                        elif col_lower in ['email', 'email id', 'email_id', 'emailid']:
                            email = str(row.get(col, '')).strip().lower()
                        elif col_lower in ['unit code', 'unit_code', 'unitcode', 'unit']:
                            uc = str(row.get(col, '')).strip().upper()
                        elif col_lower in ['department', 'dept', 'department name', 'department_name']:
                            dn = str(row.get(col, '')).strip().upper()
                    
                    if not eid:
                        validation_errors.append({'row': row_num, 'message': 'Employee ID is required'})
                        continue
                    if not ename:
                        validation_errors.append({'row': row_num, 'message': 'Employee Name is required'})
                        continue
                    if not mob:
                        validation_errors.append({'row': row_num, 'message': 'Mobile number is required'})
                        continue
                    if not email:
                        validation_errors.append({'row': row_num, 'message': 'Email is required'})
                        continue
                    if '@' not in email or '.' not in email:
                        validation_errors.append({'row': row_num, 'message': f'Invalid email format: {email}'})
                        continue
                    if not mob.isdigit() or len(mob) != 10:
                        validation_errors.append({'row': row_num, 'message': f'Mobile number must be 10 digits: {mob}'})
                        continue
                    if uc and uc not in all_units:
                        valid_units = ', '.join(list(all_units.keys())[:5])
                        if len(all_units) > 5:
                            valid_units += f' and {len(all_units) - 5} more'
                        validation_errors.append({'row': row_num, 'message': f'Invalid Unit Code "{uc}". Valid units: {valid_units}'})
                        continue
                    if dn and dn not in all_departments:
                        valid_depts = ', '.join(list(all_departments.keys())[:5])
                        if len(all_departments) > 5:
                            valid_depts += f' and {len(all_departments) - 5} more'
                        validation_errors.append({'row': row_num, 'message': f'Invalid Department "{dn}". Valid departments: {valid_depts}'})
                        continue
                    
                    successful_rows += 1
                
                if validation_errors:
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': f'Validation failed. Found {len(validation_errors)} error(s).',
                            'errors': validation_errors
                        })
                    messages.error(request, f'Validation failed. Found {len(validation_errors)} error(s).')
                    for err in validation_errors[:5]:
                        messages.error(request, f'Row {err["row"]}: {err["message"]}')
                    if len(validation_errors) > 5:
                        messages.error(request, f'... and {len(validation_errors) - 5} more errors.')
                    return redirect('settings_page')
                
                success_count = 0
                error_count = 0
                
                for idx, row in df.iterrows():
                    try:
                        eid = None
                        ename = None
                        mob = None
                        email = None
                        uc = None
                        dn = None
                        
                        for col in df.columns:
                            col_lower = col.lower()
                            if col_lower in ['employee id', 'employee_id', 'employeeid']:
                                eid = str(row.get(col, '')).strip().upper()
                            elif col_lower in ['employee name', 'employee_name', 'employeename', 'name']:
                                ename = str(row.get(col, '')).strip().upper()
                            elif col_lower in ['mobile', 'phone', 'contact']:
                                mob = str(row.get(col, '')).strip()
                            elif col_lower in ['email', 'email id', 'email_id', 'emailid']:
                                email = str(row.get(col, '')).strip().lower()
                            elif col_lower in ['unit code', 'unit_code', 'unitcode', 'unit']:
                                uc = str(row.get(col, '')).strip().upper()
                            elif col_lower in ['department', 'dept', 'department name', 'department_name']:
                                dn = str(row.get(col, '')).strip().upper()
                        
                        if not eid or not ename or not mob or not email:
                            error_count += 1
                            continue
                        
                        unit_obj = all_units.get(uc) if uc else None
                        dept_obj = all_departments.get(dn) if dn else None
                        
                        EmployeeMaster.objects.update_or_create(
                            employee_id=eid,
                            defaults={
                                'employee_name': ename,
                                'mobile': mob,
                                'email': email,
                                'unit': unit_obj,
                                'department': dept_obj,
                                'is_active': True
                            }
                        )
                        success_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        logger.error(f"Error processing row {idx+2}: {str(e)}")
                
                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'message': f'Successfully processed {success_count} employees. {error_count} skipped.'
                    })
                
                if success_count > 0:
                    messages.success(request, f'Successfully uploaded {success_count} employees.')
                if error_count > 0:
                    messages.warning(request, f'{error_count} rows were skipped due to errors.')
                
            except Exception as e:
                logger.error(f"Bulk upload error: {str(e)}")
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': f'Error processing file: {str(e)}'
                    })
                messages.error(request, f'Error processing file: {str(e)}')
        
        elif action == 'edit_employee':
            emp = get_object_or_404(EmployeeMaster, pk=request.POST.get('emp_id'))
            emp.employee_id = request.POST.get('employee_id','').strip().upper()
            emp.employee_name = request.POST.get('employee_name','').strip().upper()
            emp.mobile = request.POST.get('mobile','').strip()
            emp.email = request.POST.get('email','').strip()
            emp.unit_id = request.POST.get('unit') or None
            emp.department_id = request.POST.get('department') or None
            try: 
                emp.save()
                messages.success(request, 'Employee updated successfully.')
            except IntegrityError: 
                messages.error(request, 'Employee ID already exists.')
        
        elif action == 'toggle_employee':
            emp = get_object_or_404(EmployeeMaster, pk=request.POST.get('emp_id'))
            emp.is_active = not emp.is_active
            emp.save()
            messages.success(request, f'Employee {"activated" if emp.is_active else "deactivated"}.')
        
        elif action == 'delete_employee':
            emp = get_object_or_404(EmployeeMaster, pk=request.POST.get('emp_id'))
            eid = emp.employee_id
            emp.delete()
            messages.success(request, f'Employee "{eid}" deleted.')
    
    return redirect('settings_page')


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def download_employee_list(request):
    emps = EmployeeMaster.objects.all().order_by('employee_id')
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=Employee_List_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    tf = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    hf = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    df = Font(name='Calibri', size=11)
    tfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hfill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    ws.merge_cells('A1:H1')
    ws['A1'] = "Employee Directory"
    ws['A1'].font = tf
    ws['A1'].fill = tfill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    for ci, h in enumerate(['Employee ID', 'Name', 'Mobile', 'Email', 'Unit Code', 'Unit Name', 'Department', 'Status'], 1):
        c = ws.cell(row=3, column=ci)
        c.value = h
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 25
    for ri, emp in enumerate(emps, 4):
        rd = [
            emp.employee_id,
            emp.employee_name,
            emp.mobile,
            emp.email,
            emp.unit.code if emp.unit else '',
            emp.unit.full_name if emp.unit else '',
            emp.department.name if emp.department else '',
            'Active' if emp.is_active else 'Inactive'
        ]
        for ci, v in enumerate(rd, 1):
            c = ws.cell(row=ri, column=ci)
            c.value = v
            c.font = df
    for col in ws.columns:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(
            max(len(str(c.value or '')) for c in col if c.row > 1) + 3, 12
        )
    wb.save(resp)
    return resp


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def download_employee_template(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Employee_Upload_Template.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Template"
    
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    headers = ['Employee ID', 'Employee Name', 'Mobile', 'Email', 'Unit Code', 'Department']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    sample_data = [
        ['EMP001', 'JOHN DOE', '9876543210', 'john.doe@company.com', 'GPL', 'Production'],
        ['EMP002', 'JANE SMITH', '9876543211', 'jane.smith@company.com', 'GPLAST', 'QA'],
        ['EMP003', 'MIKE JOHNSON', '9876543212', 'mike.johnson@company.com', 'IMD', 'Purchase'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = Font(name='Calibri', size=11)
    
    note_row = len(sample_data) + 3
    note_cell = ws.cell(row=note_row, column=1)
    note_cell.value = "Mandatory fields: Employee ID, Employee Name, Mobile, Email"
    note_cell.font = Font(name='Calibri', size=10, italic=True, color='FF0000')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
    
    validation_row = note_row + 1
    validation_cell = ws.cell(row=validation_row, column=1)
    validation_cell.value = "Note: Unit Code must match existing active units in the system. Department must match existing active departments."
    validation_cell.font = Font(name='Calibri', size=9, italic=True, color='666666')
    ws.merge_cells(start_row=validation_row, start_column=1, end_row=validation_row, end_column=6)
    
    column_widths = [15, 20, 15, 25, 12, 15]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def settings_credentials(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_credential':
            uid = request.POST.get('unit')
            did = request.POST.get('department')
            uname = request.POST.get('username','').strip()
            pwd = request.POST.get('password','').strip()
            if not all([uid, did, uname, pwd]): 
                messages.error(request, "All fields are required.")
                return redirect('settings_page')
            if DepartmentCredential.objects.filter(unit_id=uid, department_id=did).exists():
                messages.error(request, "Credential already exists for this department.")
                return redirect('settings_page')
            try:
                cred = DepartmentCredential.objects.create(
                    unit_id=uid,
                    department_id=did,
                    username=uname,
                    password=pwd
                )
                if not User.objects.filter(username=uname).exists():
                    User.objects.create_user(username=uname, password=pwd, is_staff=False)
                u = Unit.objects.get(pk=uid)
                d = Department.objects.get(pk=did)
                messages.success(request, f'Credential for {u.code}-{d.name} added successfully!')
            except Exception as ex: 
                messages.error(request, f'Error: {ex}')
        elif action == 'edit_credential':
            cred = get_object_or_404(DepartmentCredential, pk=request.POST.get('cred_id'))
            ou = cred.username
            nu = request.POST.get('username','').strip()
            np = request.POST.get('password','').strip()
            cred.username = nu
            if np: 
                cred.password = np
            try:
                cred.save()
                user = User.objects.filter(username=ou).first()
                if user:
                    if ou != nu:
                        user.username = nu
                    if np:
                        user.set_password(np)
                    user.save()
                elif not User.objects.filter(username=nu).exists():
                    User.objects.create_user(username=nu, password=np or cred.password, is_staff=False)
                messages.success(request, 'Credential updated successfully!')
            except Exception as ex: 
                messages.error(request, f'Error: {ex}')
        elif action == 'toggle_credential':
            cred = get_object_or_404(DepartmentCredential, pk=request.POST.get('cred_id'))
            cred.is_active = not cred.is_active
            cred.save()
            user = User.objects.filter(username=cred.username).first()
            if user:
                user.is_active = cred.is_active
                user.save()
            messages.success(request, f'Credential {"activated" if cred.is_active else "deactivated"}.')
        elif action == 'delete_credential':
            cred = get_object_or_404(DepartmentCredential, pk=request.POST.get('cred_id'))
            info = f'{cred.unit.code}-{cred.department.name}'
            uname = cred.username
            user = User.objects.filter(username=uname).first()
            if user:
                user.is_active = False
                user.save()
            cred.delete()
            messages.success(request, f'Credential for {info} deleted.')
    return redirect('settings_page')


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def download_credentials(request):
    creds = DepartmentCredential.objects.all().select_related('unit', 'department').order_by('unit__code', 'department__name')
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=Credentials_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credentials"
    tf = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    hf = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    df = Font(name='Calibri', size=11)
    tfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hfill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    ws.merge_cells('A1:F1')
    ws['A1'] = "Department Credentials"
    ws['A1'].font = tf
    ws['A1'].fill = tfill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    for ci, h in enumerate(['Unit Code', 'Unit Name', 'Department', 'Username', 'Password', 'Status'], 1):
        c = ws.cell(row=3, column=ci)
        c.value = h
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 25
    for ri, cred in enumerate(creds, 4):
        rd = [
            cred.unit.code,
            cred.unit.full_name,
            cred.department.name,
            cred.username,
            cred.password,
            'Active' if cred.is_active else 'Inactive'
        ]
        for ci, v in enumerate(rd, 1):
            c = ws.cell(row=ri, column=ci)
            c.value = v
            c.font = df
    for col in ws.columns:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(
            max(len(str(c.value or '')) for c in col if c.row > 1) + 3, 12
        )
    wb.save(resp)
    return resp


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def test_notifications(request): 
    return render(request, 'admin_panel/test_notifications.html')


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def test_success_message(request): 
    messages.success(request, 'Test success message.')
    return redirect('test_notifications')


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def test_error_message(request): 
    messages.error(request, 'Test error message.')
    return redirect('test_notifications')


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def test_warning_message(request): 
    messages.warning(request, 'Test warning message.')
    return redirect('test_notifications')


@login_required
@user_passes_test(is_admin, login_url='admin_dashboard')
def test_info_message(request): 
    messages.info(request, 'Test info message.')
    return redirect('test_notifications')


def get_units(request):
    try:
        units = Unit.objects.filter(is_active=True).order_by('code')
        units_list = []
        for unit in units:
            units_list.append({
                'id': unit.id,
                'code': unit.code,
                'name': unit.full_name or unit.code,
                'full_name': unit.full_name or unit.code
            })
        return JsonResponse({
            'units': units_list, 
            'success': True,
            'count': len(units_list)
        })
    except Exception as e:
        logger.error(f"Error in get_units: {str(e)}")
        return JsonResponse({
            'success': False, 
            'error': str(e),
            'units': []
        })


def get_departments_by_unit(request):
    unit_id = request.GET.get('unit_id')
    
    try:
        if unit_id:
            departments = Department.objects.filter(unit_id=unit_id, is_active=True).order_by('name')
        else:
            departments = Department.objects.filter(is_active=True).order_by('name')
        
        departments_list = []
        for dept in departments:
            departments_list.append({
                'id': dept.id,
                'name': dept.name,
                'unit_id': dept.unit_id
            })
        
        return JsonResponse({
            'success': True,
            'departments': departments_list,
            'count': len(departments_list)
        })
    except Exception as e:
        logger.error(f"Error in get_departments_by_unit: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e),
            'departments': []
        })


def get_employee_details(request):
    eid = request.GET.get('employee_id', '').strip().upper()
    u_uid = request.GET.get('unit_id', '')
    u_did = request.GET.get('department_id', '')
    
    if not eid: 
        return JsonResponse({
            'found': False, 
            'message': 'Please enter an Employee ID.'
        })
    
    try:
        emp = EmployeeMaster.objects.get(employee_id=eid, is_active=True)
        
        mismatches = []
        if u_uid and emp.unit_id and str(emp.unit_id) != str(u_uid):
            mismatches.append(f'Unit: {emp.unit.code if emp.unit else "Unknown"} (expected: {u_uid})')
        if u_did and emp.department_id and str(emp.department_id) != str(u_did):
            mismatches.append(f'Department: {emp.department.name if emp.department else "Unknown"} (expected: {u_did})')
        
        if mismatches:
            return JsonResponse({
                'found': False, 
                'message': f'Employee belongs to different department/unit: {", ".join(mismatches)}',
                'mismatch': True,
                'employee': {
                    'employee_id': emp.employee_id,
                    'employee_name': emp.employee_name,
                    'mobile': emp.mobile,
                    'email': emp.email,
                    'unit_id': emp.unit_id or None,
                    'unit_code': emp.unit.code if emp.unit else None,
                    'department_id': emp.department_id or None,
                    'department_name': emp.department.name if emp.department else None
                }
            })
        
        return JsonResponse({
            'found': True,
            'employee': {
                'employee_id': emp.employee_id,
                'employee_name': emp.employee_name,
                'mobile': emp.mobile,
                'email': emp.email,
                'unit_id': emp.unit_id or None,
                'unit_code': emp.unit.code if emp.unit else None,
                'department_id': emp.department_id or None,
                'department_name': emp.department.name if emp.department else None
            }
        })
    except EmployeeMaster.DoesNotExist:
        return JsonResponse({
            'found': False, 
            'message': f'Employee "{eid}" not found.'
        })
    except Exception as e:
        logger.error(f"Error in get_employee_details: {str(e)}")
        return JsonResponse({
            'found': False, 
            'message': 'Error fetching employee details.'
        })


def get_employees_by_department(request):
    dept_id = request.GET.get('department_id')
    
    if not dept_id:
        return JsonResponse({
            'success': False, 
            'message': 'Department ID is required'
        })
    
    try:
        department = Department.objects.get(id=dept_id, is_active=True)
        employees = EmployeeMaster.objects.filter(
            department=department
        ).order_by('employee_id')
        
        employee_list = []
        for emp in employees:
            employee_list.append({
                'id': emp.id,
                'employee_id': emp.employee_id or '-',
                'employee_name': emp.employee_name or '-',
                'mobile': emp.mobile or '-',
                'email': emp.email or '-',
                'is_active': emp.is_active
            })
        
        return JsonResponse({
            'success': True,
            'department_name': department.name,
            'department_id': department.id,
            'employees': employee_list,
            'count': len(employee_list)
        })
        
    except Department.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'message': 'Department not found'
        })
    except Exception as e:
        logger.error(f"Error in get_employees_by_department: {str(e)}")
        return JsonResponse({
            'success': False, 
            'message': str(e)
        })