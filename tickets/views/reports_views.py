# tickets/views/reports_views.py

"""
Reports Views - Reports, Export, Download Excel
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Count, Q, DateField, Value, CharField, OuterRef, Subquery
from django.db.models.functions import TruncMonth, Cast, Coalesce
from django.db import transaction
from django.contrib import messages
from django.template.loader import render_to_string
from django.template import TemplateDoesNotExist
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging

from tickets.models import (
    Unit, Department, Ticket, TicketHistory, EmployeeMaster,
    AdminContact, AdminNotificationEmail, SettingsAuditLog,
    ERPHolderMapping
)
from tickets.forms import AdminTicketForm, CloseTicketForm
from tickets.utils import send_ticket_email

from .utils import (
    is_admin,
    format_timedelta_display,
    reopen_ticket_logic,
    generate_admin_ticket_list_html,
)

logger = logging.getLogger(__name__)


def _aging_category(days):
    if days <= 7:
        return '0-7 Days'
    if days <= 15:
        return '8-15 Days'
    if days <= 30:
        return '16-30 Days'
    if days <= 60:
        return '31-60 Days'
    return '>60 Days'


@login_required
@user_passes_test(is_admin, login_url='login')  # ✅ FIXED: Removed 'tickets:'
def escalated_aging_report(request):
    """Show and export currently escalated tickets grouped by age."""
    now = timezone.now()
    unit_id = request.GET.get('unit', '').strip()
    department_id = request.GET.get('department', '').strip()
    priority = request.GET.get('priority', '').strip()
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    selected_category = request.GET.get('aging_category', '').strip()

    tickets_qs = Ticket.objects.filter(
        status='Escalated', escalated_at__isnull=False
    ).select_related('department').order_by('escalated_at')

    if department_id:
        tickets_qs = tickets_qs.filter(department_id=department_id)
    if unit_id:
        tickets_qs = tickets_qs.filter(department__unit_id=unit_id)
    if priority:
        tickets_qs = tickets_qs.filter(priority=priority)

    current_tz = timezone.get_current_timezone()
    for value, lookup, boundary in (
        (start_date, 'escalated_at__gte', datetime.min.time()),
        (end_date, 'escalated_at__lte', datetime.max.time()),
    ):
        if value:
            try:
                date_value = datetime.strptime(value, '%Y-%m-%d').date()
                date_time = timezone.make_aware(
                    datetime.combine(date_value, boundary), current_tz
                )
                tickets_qs = tickets_qs.filter(**{lookup: date_time})
            except ValueError:
                pass

    rows = []
    counts = {label: 0 for label in (
        '0-7 Days', '8-15 Days', '16-30 Days', '31-60 Days', '>60 Days'
    )}
    for ticket in tickets_qs:
        aging_days = max(0, (now - ticket.escalated_at).days)
        category = _aging_category(aging_days)
        counts[category] += 1
        rows.append({
            'ticket': ticket,
            'aging_days': aging_days,
            'category': category,
        })

    if selected_category in counts:
        rows = [row for row in rows if row['category'] == selected_category]

    if 'export' in request.GET:
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            'attachment; filename=Escalated_Aging_Report_'
            f'{now.strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Escalated Aging'
        headers = [
            'Ticket Number', 'Subject', 'Department', 'Priority',
            'Escalated Date', 'Aging Days', 'Aging Category'
        ]
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1F4E79')
        for row in rows:
            ticket = row['ticket']
            worksheet.append([
                ticket.ticket_number,
                ticket.subject,
                ticket.department.name if ticket.department else '',
                ticket.priority,
                ticket.escalated_at.astimezone(current_tz).strftime('%d-%b-%Y %I:%M %p'),
                row['aging_days'],
                row['category'],
            ])
        for column, width in {'A': 18, 'B': 36, 'C': 24, 'D': 14, 'E': 24, 'F': 14, 'G': 16}.items():
            worksheet.column_dimensions[column].width = width
        worksheet.freeze_panes = 'A2'
        workbook.save(response)
        return response

    context = {
        'aging_rows': rows,
        'aging_counts': counts,
        'aging_categories': [
            {'label': label, 'count': counts[label]}
            for label in counts
        ],
        'total_escalated': len(rows),
        'units': Unit.objects.filter(is_active=True).order_by('code'),
        'departments': Department.objects.all().order_by('name'),
        'priority_choices': Ticket.PRIORITY_CHOICES,
        'selected_unit': unit_id,
        'selected_department': department_id,
        'selected_priority': priority,
        'start_date': start_date,
        'end_date': end_date,
        'selected_category': selected_category,
        'chart_labels': list(counts.keys()),
        'chart_values': list(counts.values()),
    }
    filter_query = request.GET.copy()
    filter_query.pop('aging_category', None)
    filter_query.pop('export', None)
    context['filter_query'] = filter_query.urlencode()
    return render(request, 'admin_panel/escalated_aging_report.html', context)


# ============================================================
# REPORTS VIEW - COMPLETELY REWRITTEN
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')  # ✅ FIXED: Removed 'tickets:'
def reports(request):
    """
    Reports page with advanced filtering and export
    """
    # ✅ Start with base queryset
    tickets_qs = Ticket.objects.all()
    
    # ✅ Create a copy for unfiltered stats before applying filters
    all_tickets_for_stats = tickets_qs

    # ✅ Get all filter parameters from GET
    unit_id = request.GET.get('unit', '').strip()
    dept_id = request.GET.get('department', '').strip()
    priority = request.GET.get('priority', '').strip()
    status = request.GET.get('status', '').strip()
    assigned_person = request.GET.get('assigned_person', '').strip()
    created_by_role = request.GET.get('created_by_role', '').strip()
    error_type = request.GET.get('error_type', '').strip()
    vendor_ticket = request.GET.get('vendor_ticket_number', '').strip()
    main_error_type = request.GET.get('main_error_type', '').strip()
    sub_error_type = request.GET.get('sub_error_type', '').strip()
    erp_id_filter = request.GET.get('erp_id', '').strip()
    created_start = request.GET.get('created_start', '').strip()
    created_end = request.GET.get('created_end', '').strip()
    closed_start = request.GET.get('closed_start', '').strip()
    closed_end = request.GET.get('closed_end', '').strip()
    escalated_start = request.GET.get('escalated_start', '').strip()
    escalated_end = request.GET.get('escalated_end', '').strip()
    category = request.GET.get('category', 'all').strip()
    is_reopened = request.GET.get('is_reopened', '').strip()
    
    # ✅ DEBUG - Log all filters
    logger.info(f"Filters received - category: {category}, unit: {unit_id}, status: {status}, priority: {priority}")
    
    # ============================================================
    # APPLY CATEGORY FILTERS
    # ============================================================
    if category == 'open':
        tickets_qs = tickets_qs.filter(status='Open')
    elif category == 'assigned':
        tickets_qs = tickets_qs.filter(status='Assigned')
    elif category == 'hold':
        tickets_qs = tickets_qs.filter(status='Hold')
    elif category == 'escalated':
        tickets_qs = tickets_qs.filter(status='Escalated')
    elif category == 'closed':
        tickets_qs = tickets_qs.filter(status='Closed')
    elif category == 'escalated_closed':
        tickets_qs = tickets_qs.filter(status='Closed', escalated_at__isnull=False)
    elif category == 'reopened':
        tickets_qs = tickets_qs.filter(history__action='Ticket Reopened').distinct()
    
    # ============================================================
    # APPLY STANDARD FILTERS
    # ============================================================
    if unit_id:
        tickets_qs = tickets_qs.filter(unit_id=unit_id)
    if dept_id:
        tickets_qs = tickets_qs.filter(department_id=dept_id)
    if priority:
        tickets_qs = tickets_qs.filter(priority=priority)
    if status:
        tickets_qs = tickets_qs.filter(status=status)
    if assigned_person:
        tickets_qs = tickets_qs.filter(assigned_person__icontains=assigned_person)
    if created_by_role:
        tickets_qs = tickets_qs.filter(created_by_role=created_by_role)
    if error_type:
        tickets_qs = tickets_qs.filter(error_type=error_type)
    if vendor_ticket:
        tickets_qs = tickets_qs.filter(vendor_ticket_number__icontains=vendor_ticket)
    
    # ============================================================
    # APPLY ERROR TYPE FILTERS
    # ============================================================
    if main_error_type:
        tickets_qs = tickets_qs.filter(main_error_type=main_error_type)
    if sub_error_type and sub_error_type != 'All':
        tickets_qs = tickets_qs.filter(sub_error_type=sub_error_type)
    
    # ============================================================
    # APPLY ERP ID FILTER
    # ============================================================
    if erp_id_filter:
        # Get employee IDs with this ERP ID
        employee_ids = ERPHolderMapping.objects.filter(
            erp_user_id=erp_id_filter
        ).values_list('employee__employee_id', flat=True)
        if employee_ids.exists():
            tickets_qs = tickets_qs.filter(employee_id__in=employee_ids)
        else:
            tickets_qs = tickets_qs.none()
    
    # ============================================================
    # APPLY DATE RANGE FILTERS
    # ============================================================
    current_tz = timezone.get_current_timezone()
    
    if created_start:
        try:
            created_start_date = datetime.strptime(created_start, '%Y-%m-%d').date()
            start_dt = datetime.combine(created_start_date, datetime.min.time())
            from_datetime = timezone.make_aware(start_dt, current_tz)
            tickets_qs = tickets_qs.filter(created_at__gte=from_datetime)
        except ValueError:
            pass
    
    if created_end:
        try:
            created_end_date = datetime.strptime(created_end, '%Y-%m-%d').date()
            end_dt = datetime.combine(created_end_date, datetime.max.time())
            to_datetime = timezone.make_aware(end_dt, current_tz)
            tickets_qs = tickets_qs.filter(created_at__lte=to_datetime)
        except ValueError:
            pass
    
    if closed_start:
        try:
            closed_start_date = datetime.strptime(closed_start, '%Y-%m-%d').date()
            start_dt = datetime.combine(closed_start_date, datetime.min.time())
            from_datetime = timezone.make_aware(start_dt, current_tz)
            tickets_qs = tickets_qs.filter(closed_at__gte=from_datetime)
        except ValueError:
            pass
    
    if closed_end:
        try:
            closed_end_date = datetime.strptime(closed_end, '%Y-%m-%d').date()
            end_dt = datetime.combine(closed_end_date, datetime.max.time())
            to_datetime = timezone.make_aware(end_dt, current_tz)
            tickets_qs = tickets_qs.filter(closed_at__lte=to_datetime)
        except ValueError:
            pass
    
    if escalated_start:
        try:
            escalated_start_date = datetime.strptime(escalated_start, '%Y-%m-%d').date()
            start_dt = datetime.combine(escalated_start_date, datetime.min.time())
            from_datetime = timezone.make_aware(start_dt, current_tz)
            tickets_qs = tickets_qs.filter(escalated_at__gte=from_datetime)
        except ValueError:
            pass
    
    if escalated_end:
        try:
            escalated_end_date = datetime.strptime(escalated_end, '%Y-%m-%d').date()
            end_dt = datetime.combine(escalated_end_date, datetime.max.time())
            to_datetime = timezone.make_aware(end_dt, current_tz)
            tickets_qs = tickets_qs.filter(escalated_at__lte=to_datetime)
        except ValueError:
            pass
    
    # ============================================================
    # APPLY REOPENED FILTER
    # ============================================================
    if is_reopened == 'yes':
        tickets_qs = tickets_qs.filter(history__action='Ticket Reopened').distinct()
    elif is_reopened == 'no':
        tickets_qs = tickets_qs.exclude(history__action='Ticket Reopened')
    
    # ============================================================
    # ORDER AND ANNOTATE
    # ============================================================
    tickets_qs = tickets_qs.order_by('-created_at')
    
    # ✅ Annotate with ERP ID
    erp_subquery = ERPHolderMapping.objects.filter(
        employee__employee_id=OuterRef('employee_id')
    ).values('erp_user_id')[:1]
    
    tickets_qs = tickets_qs.annotate(
        erp_id=Coalesce(Subquery(erp_subquery, output_field=CharField()), Value('Not Mapped'))
    )
    
    # ✅ DEBUG - Log final count
    total_count = tickets_qs.count()
    logger.info(f"Total tickets after all filters: {total_count}")
    
    # ============================================================
    # CALCULATE STATS
    # ============================================================
    open_count = all_tickets_for_stats.filter(status='Open').count()
    assigned_count = all_tickets_for_stats.filter(status='Assigned').count()
    hold_count = all_tickets_for_stats.filter(status='Hold').count()
    escalated_count = all_tickets_for_stats.filter(status='Escalated').count()
    closed_count = all_tickets_for_stats.filter(status='Closed').count()
    
    # ============================================================
    # GET DATA FOR DROPDOWNS
    # ============================================================
    units = Unit.objects.all()
    departments = Department.objects.all()
    employees = EmployeeMaster.objects.all().order_by('employee_name')
    erp_ids = ERPHolderMapping.objects.filter(
        erp_user_id__isnull=False
    ).exclude(
        erp_user_id__exact=''
    ).values_list('erp_user_id', flat=True).distinct().order_by('erp_user_id')
    
    # ============================================================
    # EXPORT TO EXCEL - ✅ FIXED
    # ============================================================
    if 'export' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=GPLAST_Report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tickets Report"
        
        title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        data_font = Font(name='Calibri', size=11)
        title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        now_local = timezone.now().astimezone(current_tz)
        report_time = now_local.strftime('%d-%b-%Y %I:%M:%S %p')
        
        ws.merge_cells('A1:AB1')
        ws['A1'] = f"GPLAST TICKET REPORT - Generated: {report_time}  |  Total Entries: {tickets_qs.count()}"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 45
        
        headers = [
            "Ticket Number","Status","Unit Code","Unit Full Name","Department",
            "Employee ID","ERP ID","Employee Name","Mobile","Email","Screen Number",
            "Subject","Description","Priority","Error Type","Created By Role",
            "Time to Close","Admin Reason","Assigned Person","Hold Reason",
            "Main Error Type","Sub Error Type","Closing Remarks","Closed By",
            "Vendor Ticket","Created At","Closed At","Escalated At"
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        ws.row_dimensions[3].height = 25
        
        row_idx = 4
        for t in tickets_qs:
            c_at = t.created_at.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p') if t.created_at else ""
            cl_at = t.closed_at.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p') if t.closed_at else ""
            esc_at = t.escalated_at.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p') if t.escalated_at else ""
            ttc = format_timedelta_display(t.closed_at - t.created_at) if t.status == 'Closed' and t.created_at and t.closed_at else ""
            
            row_data = [
                t.ticket_number, t.status, t.unit.code if t.unit else '', t.unit.full_name if t.unit else '',
                t.department.name if t.department else '', t.employee_id, t.erp_id, t.employee_name,
                t.mobile, t.email, 
                t.screen_number,  # ✅ FIXED: Use screen_number instead of get_screen_display
                t.subject, t.description, t.priority,
                t.error_type, t.created_by_role, ttc, t.admin_creation_reason or '',
                t.assigned_person or '', t.hold_reason or '', t.main_error_type or 'N/A',
                t.sub_error_type or 'N/A', t.closing_remarks or '', t.closed_by or '',
                t.vendor_ticket_number or '', c_at, cl_at, esc_at
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.font = data_font
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
            row_idx += 1
        
        column_widths = {
            'A': 18, 'B': 14, 'C': 14, 'D': 25, 'E': 20, 'F': 14, 'G': 16,
            'H': 22, 'I': 16, 'J': 25, 'K': 16, 'L': 30, 'M': 40, 'N': 14,
            'O': 20, 'P': 18, 'Q': 16, 'R': 20, 'S': 20, 'T': 20, 'U': 22,
            'V': 22, 'W': 30, 'X': 18, 'Y': 18, 'Z': 22, 'AA': 22, 'AB': 22
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        wb.save(response)
        return response
    
    # ============================================================
    # PAGINATION
    # ============================================================
    paginator = Paginator(tickets_qs, 15)
    page_number = request.GET.get('page')
    try:
        tickets_page = paginator.page(page_number)
    except PageNotAnInteger:
        tickets_page = paginator.page(1)
    except EmptyPage:
        tickets_page = paginator.page(paginator.num_pages)
    
    # ============================================================
    # ERROR TYPE CHOICES
    # ============================================================
    error_type_choices = Ticket.objects.filter(
        error_type__isnull=False
    ).exclude(
        error_type__exact=''
    ).values_list('error_type', flat=True).distinct().order_by('error_type')
    error_type_choices_list = [(et, et) for et in error_type_choices]

    filter_query = request.GET.copy()
    filter_query.pop('page', None)
    
    # ============================================================
    # CONTEXT
    # ============================================================
    context = {
        'tickets': tickets_page,
        'total_count': total_count,
        'open_count': open_count,
        'assigned_count': assigned_count,
        'hold_count': hold_count,
        'escalated_count': escalated_count,
        'closed_count': closed_count,
        'units': units,
        'departments': departments,
        'employees': employees,
        'erp_ids': erp_ids,
        'category': category,
        'is_reopened': is_reopened,
        'status_choices': Ticket.STATUS_CHOICES,
        'priority_choices': Ticket.PRIORITY_CHOICES,
        'created_by_choices': Ticket.CREATED_BY_CHOICES,
        'error_type_choices': error_type_choices_list,
        'selected_unit': unit_id,
        'selected_department': dept_id,
        'selected_status': status,
        'selected_priority': priority,
        'selected_assigned_person': assigned_person,
        'selected_created_by_role': created_by_role,
        'selected_error_type': error_type,
        'selected_vendor_ticket': vendor_ticket,
        'created_start': created_start,
        'created_end': created_end,
        'closed_start': closed_start,
        'closed_end': closed_end,
        'escalated_start': escalated_start,
        'escalated_end': escalated_end,
        'selected_main_error_type': main_error_type,
        'selected_sub_error_type': sub_error_type,
        'selected_erp_id': erp_id_filter,
        'filter_query': filter_query.urlencode(),
    }
    return render(request, 'admin_panel/reports.html', context)


# ============================================================
# DOWNLOAD TICKET EXCEL - ✅ FIXED
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')  # ✅ FIXED: Removed 'tickets:'
def download_ticket_excel(request, pk):
    """Export single ticket details to Excel"""
    ticket = get_object_or_404(Ticket, pk=pk)
    
    erp_id = 'Not Mapped'
    if ticket.employee_id:
        erp_mapping = ERPHolderMapping.objects.filter(
            employee__employee_id=ticket.employee_id
        ).first()
        if erp_mapping:
            erp_id = erp_mapping.erp_user_id
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Ticket_{ticket.ticket_number}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ticket {ticket.ticket_number}"
    
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    section_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    label_font = Font(name='Calibri', size=11, bold=True, color='1A2A6C')
    data_font = Font(name='Calibri', size=11, color='333333')
    history_header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    history_data_font = Font(name='Calibri', size=10, color='333333')
    
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    section_fill = PatternFill(start_color='FF6B00', end_color='FF6B00', fill_type='solid')
    label_fill = PatternFill(start_color='E8EDF5', end_color='E8EDF5', fill_type='solid')
    history_header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    ws.merge_cells('A1:F1')
    ws['A1'] = f"GPLAST TICKET DETAILS - {ticket.ticket_number}"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    row = 3
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "BASIC INFORMATION"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 30
    row += 1
    
    basic_info = [
        ('Ticket Number', ticket.ticket_number),
        ('Subject', ticket.subject),
        ('Description', ticket.description or ''),
        ('Priority', ticket.priority),
        ('Status', ticket.status),
        ('Error Type', ticket.error_type or 'Not Set'),
        ('Created Date', timezone.localtime(ticket.created_at).strftime('%d-%b-%Y %I:%M %p') if ticket.created_at else ''),
        ('Updated Date', timezone.localtime(ticket.updated_at).strftime('%d-%b-%Y %I:%M %p') if ticket.updated_at else ''),
    ]
    
    for label, value in basic_info:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=1).fill = label_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.cell(row=row, column=2, value=value).font = data_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
    
    row += 1
    
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "EMPLOYEE DETAILS"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 30
    row += 1
    
    emp_info = [
        ('Employee Name', ticket.employee_name),
        ('Employee ID', ticket.employee_id),
        ('ERP ID', erp_id),
        ('Mobile', ticket.mobile),
        ('Email', ticket.email),
        ('Unit', ticket.unit.full_name if ticket.unit else ''),
        ('Department', ticket.department.name if ticket.department else ''),
        # ✅ FIXED: Use screen_number instead of get_screen_display
        ('Screen/Module', ticket.screen_number),
    ]
    
    for label, value in emp_info:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=1).fill = label_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.cell(row=row, column=2, value=value).font = data_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
    
    row += 1
    
    row += 1
    
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "ASSIGNMENT & STATUS"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 30
    row += 1
    
    assign_info = [
        ('Created By Role', ticket.created_by_role),
        ('Assigned To', ticket.assigned_person or 'Not Assigned'),
        ('Hold Reason', ticket.hold_reason or ''),
        ('Vendor Ticket', ticket.vendor_ticket_number or ''),
        ('Admin Creation Reason', ticket.admin_creation_reason or ''),
    ]
    
    for label, value in assign_info:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=1).fill = label_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.cell(row=row, column=2, value=value).font = data_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
    
    row += 1
    
    if ticket.status == 'Closed':
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "CLOSING DETAILS"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 30
        row += 1
        
        closing_info = [
            ('Closed By', ticket.closed_by or ''),
            ('Closed Date', timezone.localtime(ticket.closed_at).strftime('%d-%b-%Y %I:%M %p') if ticket.closed_at else ''),
            ('Main Error Type', ticket.main_error_type or 'N/A'),
            ('Sub Error Type', ticket.sub_error_type or 'N/A'),
            ('Closing Remarks', ticket.closing_remarks or ''),
            ('Time to Close', format_timedelta_display(ticket.closed_at - ticket.created_at) if ticket.closed_at else ''),
        ]
        
        for label, value in closing_info:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=1).fill = label_fill
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.cell(row=row, column=2, value=value).font = data_font
            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            row += 1
        
        row += 1
    
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "AUDIT HISTORY"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 30
    row += 1
    
    history_headers = ['Timestamp', 'Action', 'Remarks', 'Performed By']
    for col_idx, header in enumerate(history_headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = history_header_font
        cell.fill = history_header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    for history in ticket.history.all().order_by('timestamp'):
        ws.cell(row=row, column=1, value=timezone.localtime(history.timestamp).strftime('%d-%b-%Y %I:%M %p')).font = history_data_font
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=row, column=2, value=history.action).font = history_data_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=row, column=3, value=history.remarks or '').font = history_data_font
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=row, column=4, value=history.get_performed_by_display()).font = history_data_font
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        row += 1
    
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"Report generated on {timezone.now().strftime('%d-%b-%Y %I:%M %p')} | GPLAST Support System"
    ws[f'A{row}'].font = Font(name='Calibri', size=9, italic=True, color='666666')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    ws.freeze_panes = 'A1'
    
    wb.save(response)
    return response


# ============================================================
# EXPORT CLOSED TICKETS (LAST 30 DAYS) TO EXCEL - ✅ FIXED
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')  # ✅ FIXED: Removed 'tickets:'
def export_closed_tickets_30_days(request):
    """Export closed tickets from the last 30 days to Excel"""
    thirty_days_ago = timezone.now() - timedelta(days=30)
    
    tickets_qs = Ticket.objects.filter(
        status='Closed',
        closed_at__gte=thirty_days_ago
    ).order_by('-closed_at')
    
    current_tz = timezone.get_current_timezone()
    now_local = timezone.now().astimezone(current_tz)
    report_time = now_local.strftime('%d-%b-%Y %I:%M:%S %p')
    
    # Build mapping of employee_id to ERP ID
    employee_ids = tickets_qs.values_list('employee_id', flat=True).distinct()
    erp_mappings = {}
    if employee_ids:
        erp_mappings_qs = ERPHolderMapping.objects.filter(
            employee__employee_id__in=employee_ids
        ).select_related('employee')
        for mapping in erp_mappings_qs:
            erp_mappings[mapping.employee.employee_id] = mapping.erp_user_id
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Closed_Tickets_30_Days_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Closed Tickets (30 Days)"
    
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10)
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    ws.merge_cells('A1:AB1')
    ws['A1'] = f"CLOSED TICKETS - LAST 30 DAYS"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45
    
    ws.merge_cells('A2:AB2')
    ws['A2'] = f"Generated: {report_time}  |  Total Closed Tickets: {tickets_qs.count()}  |  Period: {thirty_days_ago.strftime('%d-%b-%Y')} to {timezone.now().strftime('%d-%b-%Y')}"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    
    headers = [
        'Ticket Number', 'Status', 'Unit Code', 'Unit Name', 'Department',
        'Employee ID', 'ERP ID', 'Employee Name', 'Mobile', 'Email', 
        'Screen/Module',  # ✅ This column will use screen_number
        'Subject', 'Description', 'Priority', 'Error Type', 'Created By Role',
        'Admin Creation Reason', 'Assigned Person', 'Hold Reason',
        'Main Error Type', 'Sub Error Type', 'Closing Remarks', 'Closed By', 
        'Vendor Ticket Number', 'Created At', 'Closed At', 'Time to Close', 
        'Escalated At'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[4].height = 30
    
    row_idx = 5
    for ticket in tickets_qs:
        created_at_local = ticket.created_at.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p') if ticket.created_at else ''
        closed_at_local = ticket.closed_at.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p') if ticket.closed_at else ''
        escalated_at_local = ticket.escalated_at.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p') if ticket.escalated_at else ''
        
        time_to_close = ''
        if ticket.created_at and ticket.closed_at:
            duration = ticket.closed_at - ticket.created_at
            days = duration.days
            hours = duration.seconds // 3600
            minutes = (duration.seconds % 3600) // 60
            time_to_close = f"{days}d {hours}h {minutes}m" if days > 0 else f"{hours}h {minutes}m"
        
        erp_id = erp_mappings.get(ticket.employee_id, 'Not Mapped')
        
        row_data = [
            ticket.ticket_number, ticket.status, ticket.unit.code if ticket.unit else '',
            ticket.unit.full_name if ticket.unit else '', ticket.department.name if ticket.department else '',
            ticket.employee_id, erp_id, ticket.employee_name, ticket.mobile, ticket.email,
            ticket.screen_number,  # ✅ FIXED: Use screen_number instead of get_screen_display
            ticket.subject, ticket.description or '', ticket.priority,
            ticket.error_type or '', ticket.created_by_role, ticket.admin_creation_reason or '',
            ticket.assigned_person or '', ticket.hold_reason or '', ticket.main_error_type or 'N/A',
            ticket.sub_error_type or 'N/A', ticket.closing_remarks or '', ticket.closed_by or '',
            ticket.vendor_ticket_number or '', created_at_local, closed_at_local, time_to_close,
            escalated_at_local,
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        row_idx += 1
    
    column_widths = {
        'A': 18, 'B': 14, 'C': 12, 'D': 25, 'E': 20, 'F': 14, 'G': 16,
        'H': 22, 'I': 16, 'J': 25, 'K': 16, 'L': 30, 'M': 40, 'N': 14,
        'O': 20, 'P': 18, 'Q': 25, 'R': 20, 'S': 20, 'T': 22, 'U': 22,
        'V': 30, 'W': 18, 'X': 18, 'Y': 22, 'Z': 22, 'AA': 16, 'AB': 22
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    wb.save(response)
    return response