# tickets/views/employee_views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from django.utils import timezone
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.core.exceptions import ValidationError
from django.template.loader import render_to_string
from datetime import timedelta, datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from tickets.models import (
    Ticket, 
    Unit, 
    Department, 
    AdminContact,           
    AdminNotificationEmail, 
    EmployeeMaster,         
    TicketHistory,
    ReopenAttachment,
    SettingsAuditLog,
    DepartmentCredential    
)
from tickets.forms import TicketForm
from tickets.utils import send_ticket_email, validate_attachment
import logging

logger = logging.getLogger(__name__)


# ============================================================
# ROLE REDIRECT
# ============================================================
def role_redirect(request):
    """
    Redirect users based on their role
    - Admin users go to admin_dashboard
    - Regular users go to employee_dashboard
    """
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('admin_dashboard')
        else:
            return redirect('employee_dashboard')
    else:
        from django.contrib.auth.views import LoginView
        return redirect('login')


# ============================================================
# EMPLOYEE DASHBOARD
# ============================================================
@login_required
def employee_dashboard(request):
    """Employee dashboard view"""
    contact = AdminContact.objects.first()
    
    if request.user.is_staff:
        tickets = Ticket.objects.all()
    else:
        tickets = Ticket.objects.filter(
            Q(created_by_user=request.user) | 
            Q(assigned_person__icontains=request.user.username)
        ).distinct()
    
    kpis = {
        'total': tickets.count(),
        'open': tickets.filter(status='Open').count(),
        'assigned': tickets.filter(status='Assigned').count(),
        'hold': tickets.filter(status='Hold').count(),
        'escalated': tickets.filter(status='Escalated').count(),
        'closed': tickets.filter(status='Closed').count(),
        'critical': tickets.filter(priority='Critical').count(),
    }
    
    status_counts = list(tickets.values('status').annotate(count=Count('id')))
    charts_data = {
        'dept_status': {item['status']: item['count'] for item in status_counts},
        'dept_priority': dict(tickets.values('priority').annotate(count=Count('id')).values_list('priority', 'count')),
    }
    
    latest_tickets = tickets.order_by('-created_at')[:10]
    
    show_department_tickets = False
    department_name = None
    unit_name = None
    
    try:
        employee = EmployeeMaster.objects.filter(email=request.user.email).first()
        if employee and employee.department:
            show_department_tickets = True
            department_name = employee.department.name
            unit_name = employee.department.unit.full_name if employee.department.unit else None
    except:
        pass
    
    context = {
        'kpis': kpis,
        'charts_data': charts_data,
        'latest_tickets': latest_tickets,
        'contact': contact,
        'show_department_tickets': show_department_tickets,
        'department_name': department_name,
        'unit_name': unit_name,
    }
    return render(request, 'employee/dashboard.html', context)


# ============================================================
# CREATE TICKET
# ============================================================
@login_required
def create_ticket(request):
    """Create a new ticket - Employee must be from same department as logged-in user"""
    
    user_credential = None
    try:
        user_credential = DepartmentCredential.objects.filter(
            username=request.user.username,
            is_active=True
        ).select_related('unit', 'department').first()
    except Exception as e:
        logger.error(f"Error fetching user credential: {e}")
    
    if user_credential:
        user_unit_id = user_credential.unit_id
        user_dept_id = user_credential.department_id
    else:
        user_unit_id = None
        user_dept_id = None
    
    if request.method == 'POST':
        form = TicketForm(request.POST, request.FILES)
        
        employee_id = request.POST.get('employee_id', '').strip().upper()
        employee = EmployeeMaster.objects.filter(employee_id=employee_id, is_active=True).first()
        
        if employee:
            if user_credential:
                if employee.unit_id != user_credential.unit_id:
                    messages.error(request, f'❌ Employee "{employee_id}" belongs to a different Unit. You can only create tickets for employees in your unit ({user_credential.unit.code}).')
                    form = TicketForm(initial={
                        'unit': user_credential.unit_id,
                        'department': user_credential.department_id,
                    })
                    context = {
                        'form': form,
                        'units': Unit.objects.filter(is_active=True),
                        'departments': Department.objects.filter(is_active=True),
                        'user_credential': user_credential,
                    }
                    return render(request, 'employee/create_ticket.html', context)
                
                if employee.department_id != user_credential.department_id:
                    messages.error(request, f'❌ Employee "{employee_id}" belongs to a different Department. You can only create tickets for employees in your department ({user_credential.department.name}).')
                    form = TicketForm(initial={
                        'unit': user_credential.unit_id,
                        'department': user_credential.department_id,
                    })
                    context = {
                        'form': form,
                        'units': Unit.objects.filter(is_active=True),
                        'departments': Department.objects.filter(is_active=True),
                        'user_credential': user_credential,
                    }
                    return render(request, 'employee/create_ticket.html', context)
        else:
            if employee_id:
                messages.error(request, f'❌ Employee ID "{employee_id}" not found. Please enter a valid Employee ID.')
                form = TicketForm(initial={
                    'unit': user_credential.unit_id if user_credential else None,
                    'department': user_credential.department_id if user_credential else None,
                })
                context = {
                    'form': form,
                    'units': Unit.objects.filter(is_active=True),
                    'departments': Department.objects.filter(is_active=True),
                    'user_credential': user_credential,
                }
                return render(request, 'employee/create_ticket.html', context)
        
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.created_by_user = request.user
            ticket.created_by_role = 'Employee'
            
            if user_credential:
                ticket.unit = user_credential.unit
                ticket.department = user_credential.department
            
            ticket.save()
            
            TicketHistory.objects.create(
                ticket=ticket,
                action='Created Ticket',
                remarks='Ticket created by employee',
                performed_by=request.user.username
            )
            
            try:
                send_ticket_email(ticket, 'Created')
                logger.info(f"Email sent for ticket {ticket.ticket_number}")
            except Exception as e:
                logger.error(f"Failed to send email for ticket {ticket.ticket_number}: {e}")
            
            messages.success(request, f'Ticket #{ticket.ticket_number} created successfully!')
            return redirect('ticket_detail', ticket_id=ticket.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        initial_data = {}
        if user_credential:
            initial_data['unit'] = user_credential.unit_id
            initial_data['department'] = user_credential.department_id
        
        form = TicketForm(initial=initial_data)
    
    departments = Department.objects.filter(is_active=True)
    if user_credential:
        departments = departments.filter(unit_id=user_credential.unit_id)
    
    context = {
        'form': form,
        'units': Unit.objects.filter(is_active=True),
        'departments': departments,
        'user_credential': user_credential,
    }
    return render(request, 'employee/create_ticket.html', context)


# ============================================================
# ALL TICKETS - WITH 30 DAY CLOSED FILTER AND EXCEL EXPORT
# ============================================================
@login_required
def all_tickets(request):
    """
    View all tickets with 30-day filter for closed tickets
    - Shows closed tickets from last 30 days by default
    - All non-closed tickets are always shown
    - Search works across ALL history
    - Pagination: 20 per page
    - Excel export with current filters
    """
    tickets = Ticket.objects.all().order_by('-created_at')
    
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    search_query = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # ✅ NEW: Main Error Type and Sub Error Type filters
    main_error_type = request.GET.get('main_error_type', '').strip()
    sub_error_type = request.GET.get('sub_error_type', '').strip()
    
    # 30-DAY CLOSED TICKETS FILTER (DEFAULT VIEW)
    if not status_filter:
        thirty_days_ago = timezone.now() - timedelta(days=30)
        tickets = tickets.filter(
            Q(status='Closed', closed_at__gte=thirty_days_ago) |
            ~Q(status='Closed')
        )
    
    if status_filter:
        tickets = tickets.filter(status=status_filter)
    
    if priority_filter:
        tickets = tickets.filter(priority=priority_filter)
    
    # ✅ NEW: Apply Main Error Type filter
    if main_error_type and main_error_type != '':
        tickets = tickets.filter(main_error_type=main_error_type)
    
    # ✅ NEW: Apply Sub Error Type filter
    if sub_error_type and sub_error_type != '' and sub_error_type != 'All':
        tickets = tickets.filter(sub_error_type=sub_error_type)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            from_datetime = timezone.make_aware(
                datetime.combine(date_from_obj, datetime.min.time())
            )
            tickets = tickets.filter(created_at__gte=from_datetime)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            to_datetime = timezone.make_aware(
                datetime.combine(date_to_obj, datetime.max.time())
            )
            tickets = tickets.filter(created_at__lte=to_datetime)
        except ValueError:
            pass
    
    if search_query:
        tickets = tickets.filter(
            Q(ticket_number__icontains=search_query) |
            Q(subject__icontains=search_query) |
            Q(employee_name__icontains=search_query) |
            Q(employee_id__icontains=search_query) |
            Q(unit__code__icontains=search_query) |
            Q(department__name__icontains=search_query)
        )
    
    thirty_days_ago = timezone.now() - timedelta(days=30)
    closed_count_30_days = Ticket.objects.filter(
        status='Closed',
        closed_at__gte=thirty_days_ago
    ).count()
    
    # ============================================================
    # EXCEL EXPORT - Export filtered tickets
    # ============================================================
    if request.GET.get('export') == 'excel':
        return export_filtered_tickets_excel(request, tickets)
    
    is_ajax = request.GET.get('ajax', False)
    if is_ajax:
        tickets = tickets[:50]
        html = render_to_string('employee/_ticket_list_modal.html', {
            'tickets': tickets,
        }, request=request)
        return JsonResponse({
            'html': html,
            'success': True,
            'count': tickets.count()
        })
    
    paginator = Paginator(tickets, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'closed_count_30_days': closed_count_30_days,
        'status_choices': Ticket.STATUS_CHOICES,
        'priority_choices': Ticket.PRIORITY_CHOICES,
        'units': Unit.objects.filter(is_active=True),
        'departments': Department.objects.filter(is_active=True),
        # ✅ NEW: Pass error type filters to template
        'selected_main_error_type': main_error_type,
        'selected_sub_error_type': sub_error_type,
    }
    return render(request, 'employee/all_tickets.html', context)


# ============================================================
# MY TICKETS - WITH EXCEL EXPORT
# ============================================================
@login_required
def my_tickets(request):
    """View tickets created by the logged-in user with full filtering and Excel export"""
    tickets = Ticket.objects.filter(created_by_user=request.user).order_by('-created_at')
    
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    assigned_person_filter = request.GET.get('assigned_person', '')
    search_query = request.GET.get('search', '')
    ticket_number_filter = request.GET.get('ticket_number', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # ✅ NEW: Main Error Type and Sub Error Type filters
    main_error_type = request.GET.get('main_error_type', '').strip()
    sub_error_type = request.GET.get('sub_error_type', '').strip()
    
    selected_status = status_filter
    selected_priority = priority_filter
    selected_assigned_person = assigned_person_filter
    selected_ticket_number = ticket_number_filter
    
    if status_filter:
        tickets = tickets.filter(status=status_filter)
    
    if priority_filter:
        tickets = tickets.filter(priority=priority_filter)
    
    if assigned_person_filter:
        tickets = tickets.filter(assigned_person=assigned_person_filter)
    
    if ticket_number_filter:
        tickets = tickets.filter(ticket_number__icontains=ticket_number_filter)
    
    # ✅ NEW: Apply Main Error Type filter
    if main_error_type and main_error_type != '':
        tickets = tickets.filter(main_error_type=main_error_type)
    
    # ✅ NEW: Apply Sub Error Type filter
    if sub_error_type and sub_error_type != '' and sub_error_type != 'All':
        tickets = tickets.filter(sub_error_type=sub_error_type)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            from_datetime = timezone.make_aware(
                datetime.combine(date_from_obj, datetime.min.time())
            )
            tickets = tickets.filter(created_at__gte=from_datetime)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            to_datetime = timezone.make_aware(
                datetime.combine(date_to_obj, datetime.max.time())
            )
            tickets = tickets.filter(created_at__lte=to_datetime)
        except ValueError:
            pass
    
    if search_query:
        tickets = tickets.filter(
            Q(ticket_number__icontains=search_query) |
            Q(subject__icontains=search_query) |
            Q(employee_name__icontains=search_query) |
            Q(employee_id__icontains=search_query) |
            Q(unit__code__icontains=search_query) |
            Q(department__name__icontains=search_query)
        )
    
    total = tickets.count()
    open_tickets = tickets.filter(status='Open').count()
    assigned_tickets = tickets.filter(status='Assigned').count()
    hold_tickets = tickets.filter(status='Hold').count()
    escalated_tickets = tickets.filter(status='Escalated').count()
    closed_tickets = tickets.filter(status='Closed').count()
    
    thirty_days_ago = timezone.now() - timedelta(days=30)
    closed_count_30_days = Ticket.objects.filter(
        created_by_user=request.user,
        status='Closed',
        closed_at__gte=thirty_days_ago
    ).count()
    
    employees = EmployeeMaster.objects.filter(is_active=True).order_by('employee_name')
    
    # ============================================================
    # EXCEL EXPORT - Export filtered my tickets
    # ============================================================
    if request.GET.get('export') == 'excel':
        return export_filtered_my_tickets_excel(request, tickets)
    
    is_ajax = request.GET.get('ajax', False)
    if is_ajax:
        tickets = tickets[:50]
        html = render_to_string('employee/_ticket_list_modal.html', {
            'tickets': tickets,
        }, request=request)
        return JsonResponse({
            'html': html,
            'success': True,
            'count': tickets.count()
        })
    
    paginator = Paginator(tickets, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total': total,
        'open_tickets': open_tickets,
        'assigned_tickets': assigned_tickets,
        'hold_tickets': hold_tickets,
        'escalated_tickets': escalated_tickets,
        'closed_tickets': closed_tickets,
        'closed_count_30_days': closed_count_30_days,
        'search_query': search_query,
        'selected_status': selected_status,
        'selected_priority': selected_priority,
        'selected_assigned_person': selected_assigned_person,
        'selected_ticket_number': selected_ticket_number,
        'date_from': date_from,
        'date_to': date_to,
        'status_choices': Ticket.STATUS_CHOICES,
        'priority_choices': Ticket.PRIORITY_CHOICES,
        'employees': employees,
        # ✅ NEW: Pass error type filters to template
        'selected_main_error_type': main_error_type,
        'selected_sub_error_type': sub_error_type,
    }
    return render(request, 'employee/my_tickets.html', context)


# ============================================================
# EXPORT FILTERED TICKETS TO EXCEL (ALL TICKETS)
# ============================================================
def export_filtered_tickets_excel(request, tickets_qs):
    """Export filtered tickets from all tickets view to Excel"""
    
    current_tz = timezone.get_current_timezone()
    now_utc = timezone.now()
    if timezone.is_naive(now_utc):
        now_utc = timezone.make_aware(now_utc, timezone.utc)
    now_local = now_utc.astimezone(current_tz)
    report_time = now_local.strftime('%d-%b-%Y %I:%M:%S %p')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=All_Tickets_Export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Tickets"
    
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10)
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    ws.merge_cells('A1:AB1')
    ws['A1'] = f"ALL TICKETS - GPLAST SUPPORT SYSTEM"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45
    
    ws.merge_cells('A2:AB2')
    ws['A2'] = f"Generated: {report_time}  |  Total Tickets: {tickets_qs.count()}"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    
    # ✅ UPDATED: Added Main Error Type and Sub Error Type columns
    headers = [
        'Ticket Number', 'Status', 'Unit Code', 'Unit Name', 'Department',
        'Employee ID', 'Employee Name', 'Mobile', 'Email', 'Screen/Module',
        'Subject', 'Description', 'Priority', 'Error Type', 'Created By Role',
        'Assigned Person', 'Hold Reason', 'Closing Remarks', 'Closed By',
        'Vendor Ticket', 'Main Error Type', 'Sub Error Type',
        'Created At', 'Closed At', 'Time to Close', 'Escalated At'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[4].height = 30
    
    row_idx = 5
    for ticket in tickets_qs:
        if ticket.created_at:
            if timezone.is_naive(ticket.created_at):
                utc_time = timezone.make_aware(ticket.created_at, timezone.utc)
            else:
                utc_time = ticket.created_at
            created_at_local = utc_time.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p')
        else:
            created_at_local = ''
        
        if ticket.closed_at:
            if timezone.is_naive(ticket.closed_at):
                utc_time = timezone.make_aware(ticket.closed_at, timezone.utc)
            else:
                utc_time = ticket.closed_at
            closed_at_local = utc_time.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p')
        else:
            closed_at_local = ''
        
        if ticket.escalated_at:
            if timezone.is_naive(ticket.escalated_at):
                utc_time = timezone.make_aware(ticket.escalated_at, timezone.utc)
            else:
                utc_time = ticket.escalated_at
            escalated_at_local = utc_time.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p')
        else:
            escalated_at_local = ''
        
        time_to_close = ''
        if ticket.created_at and ticket.closed_at:
            duration = ticket.closed_at - ticket.created_at
            days = duration.days
            hours = duration.seconds // 3600
            minutes = (duration.seconds % 3600) // 60
            if days > 0:
                time_to_close = f"{days}d {hours}h {minutes}m"
            else:
                time_to_close = f"{hours}h {minutes}m"
        
        row_data = [
            ticket.ticket_number,
            ticket.status,
            ticket.unit.code if ticket.unit else '',
            ticket.unit.full_name if ticket.unit else '',
            ticket.department.name if ticket.department else '',
            ticket.employee_id,
            ticket.employee_name,
            ticket.mobile,
            ticket.email,
            ticket.screen_number,
            ticket.subject,
            ticket.description or '',
            ticket.priority,
            ticket.error_type or '',
            ticket.created_by_role,
            ticket.assigned_person or '',
            ticket.hold_reason or '',
            ticket.closing_remarks or '',
            ticket.closed_by or '',
            ticket.vendor_ticket_number or '',
            ticket.main_error_type or 'N/A',  # ✅ NEW
            ticket.sub_error_type or 'N/A',   # ✅ NEW
            created_at_local,
            closed_at_local,
            time_to_close,
            escalated_at_local,
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        row_idx += 1
    
    column_widths = {
        'A': 18, 'B': 14, 'C': 12, 'D': 25, 'E': 20,
        'F': 14, 'G': 22, 'H': 16, 'I': 25, 'J': 16,
        'K': 30, 'L': 40, 'M': 14, 'N': 20, 'O': 18,
        'P': 20, 'Q': 20, 'R': 30, 'S': 18, 'T': 18,
        'U': 22, 'V': 22, 'W': 16, 'X': 22, 'Y': 16, 'Z': 22
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    wb.save(response)
    return response


# ============================================================
# EXPORT FILTERED MY TICKETS TO EXCEL
# ============================================================
def export_filtered_my_tickets_excel(request, tickets_qs):
    """Export filtered tickets from my tickets view to Excel"""
    
    current_tz = timezone.get_current_timezone()
    now_utc = timezone.now()
    if timezone.is_naive(now_utc):
        now_utc = timezone.make_aware(now_utc, timezone.utc)
    now_local = now_utc.astimezone(current_tz)
    report_time = now_local.strftime('%d-%b-%Y %I:%M:%S %p')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=My_Tickets_Export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Tickets"
    
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10)
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    ws.merge_cells('A1:AB1')
    ws['A1'] = f"MY TICKETS - GPLAST SUPPORT SYSTEM"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45
    
    ws.merge_cells('A2:AB2')
    ws['A2'] = f"Generated: {report_time}  |  Total Tickets: {tickets_qs.count()}"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    
    # ✅ UPDATED: Added Main Error Type and Sub Error Type columns
    headers = [
        'Ticket Number', 'Status', 'Unit Code', 'Unit Name', 'Department',
        'Employee ID', 'Employee Name', 'Mobile', 'Email', 'Screen/Module',
        'Subject', 'Description', 'Priority', 'Error Type', 'Created By Role',
        'Assigned Person', 'Hold Reason', 'Closing Remarks', 'Closed By',
        'Vendor Ticket', 'Main Error Type', 'Sub Error Type',
        'Created At', 'Closed At', 'Time to Close', 'Escalated At'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[4].height = 30
    
    row_idx = 5
    for ticket in tickets_qs:
        if ticket.created_at:
            if timezone.is_naive(ticket.created_at):
                utc_time = timezone.make_aware(ticket.created_at, timezone.utc)
            else:
                utc_time = ticket.created_at
            created_at_local = utc_time.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p')
        else:
            created_at_local = ''
        
        if ticket.closed_at:
            if timezone.is_naive(ticket.closed_at):
                utc_time = timezone.make_aware(ticket.closed_at, timezone.utc)
            else:
                utc_time = ticket.closed_at
            closed_at_local = utc_time.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p')
        else:
            closed_at_local = ''
        
        if ticket.escalated_at:
            if timezone.is_naive(ticket.escalated_at):
                utc_time = timezone.make_aware(ticket.escalated_at, timezone.utc)
            else:
                utc_time = ticket.escalated_at
            escalated_at_local = utc_time.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p')
        else:
            escalated_at_local = ''
        
        time_to_close = ''
        if ticket.created_at and ticket.closed_at:
            duration = ticket.closed_at - ticket.created_at
            days = duration.days
            hours = duration.seconds // 3600
            minutes = (duration.seconds % 3600) // 60
            if days > 0:
                time_to_close = f"{days}d {hours}h {minutes}m"
            else:
                time_to_close = f"{hours}h {minutes}m"
        
        row_data = [
            ticket.ticket_number,
            ticket.status,
            ticket.unit.code if ticket.unit else '',
            ticket.unit.full_name if ticket.unit else '',
            ticket.department.name if ticket.department else '',
            ticket.employee_id,
            ticket.employee_name,
            ticket.mobile,
            ticket.email,
            ticket.screen_number,
            ticket.subject,
            ticket.description or '',
            ticket.priority,
            ticket.error_type or '',
            ticket.created_by_role,
            ticket.assigned_person or '',
            ticket.hold_reason or '',
            ticket.closing_remarks or '',
            ticket.closed_by or '',
            ticket.vendor_ticket_number or '',
            ticket.main_error_type or 'N/A',  # ✅ NEW
            ticket.sub_error_type or 'N/A',   # ✅ NEW
            created_at_local,
            closed_at_local,
            time_to_close,
            escalated_at_local,
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        row_idx += 1
    
    column_widths = {
        'A': 18, 'B': 14, 'C': 12, 'D': 25, 'E': 20,
        'F': 14, 'G': 22, 'H': 16, 'I': 25, 'J': 16,
        'K': 30, 'L': 40, 'M': 14, 'N': 20, 'O': 18,
        'P': 20, 'Q': 20, 'R': 30, 'S': 18, 'T': 18,
        'U': 22, 'V': 22, 'W': 16, 'X': 22, 'Y': 16, 'Z': 22
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    wb.save(response)
    return response


# ============================================================
# TICKET DETAIL VIEW
# ============================================================
@login_required
def employee_ticket_detail(request, ticket_id):
    """View ticket details with individual ticket Excel download"""
    ticket = get_object_or_404(Ticket, id=ticket_id)
    history = TicketHistory.objects.filter(ticket=ticket).order_by('timestamp')
    
    # Build attachments list for display
    attachments = []
    if ticket.attachment_1:
        attachments.append({'file': ticket.attachment_1, 'name': 'Attachment 1'})
    if ticket.attachment_2:
        attachments.append({'file': ticket.attachment_2, 'name': 'Attachment 2'})
    if ticket.attachment_3:
        attachments.append({'file': ticket.attachment_3, 'name': 'Attachment 3'})
    
    can_reopen = False
    reopen_deadline_iso = ''
    if ticket.status == 'Closed' and ticket.closed_at:
        time_diff = timezone.now() - ticket.closed_at
        if time_diff.total_seconds() <= 48 * 3600:
            can_reopen = True
            reopen_deadline = ticket.closed_at + timedelta(hours=48)
            reopen_deadline_iso = reopen_deadline.isoformat()
    
    time_to_close = 'N/A'
    if ticket.closed_at and ticket.created_at:
        time_diff = ticket.closed_at - ticket.created_at
        hours = time_diff.total_seconds() / 3600
        if hours < 1:
            minutes = int(time_diff.total_seconds() / 60)
            time_to_close = f'{minutes} minutes'
        else:
            time_to_close = f'{hours:.1f} hours'
    
    employees = EmployeeMaster.objects.filter(is_active=True)
    
    context = {
        'ticket': ticket,
        'history': history,
        'attachments': attachments,
        'can_reopen': can_reopen,
        'reopen_deadline_iso': reopen_deadline_iso,
        'time_to_close': time_to_close,
        'employees': employees,
    }
    return render(request, 'employee/ticket_detail.html', context)


@login_required
def ticket_detail(request, ticket_id):
    """Alias for employee_ticket_detail"""
    return employee_ticket_detail(request, ticket_id)


# ============================================================
# DOWNLOAD INDIVIDUAL TICKET EXCEL
# ============================================================
@login_required
def download_individual_ticket_excel(request, ticket_id):
    """Download a single ticket details as Excel file"""
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Ticket_{ticket.ticket_number}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ticket {ticket.ticket_number}"
    
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    section_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    label_font = Font(name='Calibri', size=11, bold=True, color='1A2A6C')
    data_font = Font(name='Calibri', size=11, color='333333')
    
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    section_fill = PatternFill(start_color='FF6B00', end_color='FF6B00', fill_type='solid')
    label_fill = PatternFill(start_color='E8EDF5', end_color='E8EDF5', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"GPLAST TICKET DETAILS - {ticket.ticket_number}"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    row = 3
    
    # Basic Information
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "BASIC INFORMATION"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 30
    row += 1
    
    basic_info = [
        ('Ticket Number', ticket.ticket_number),
        ('Subject', ticket.subject),
        ('Description', ticket.description or ''),
        ('Priority', ticket.priority),
        ('Status', ticket.status),
        ('Error Type', ticket.error_type or 'Not Set'),
        ('Created Date', ticket.created_at.strftime('%d-%b-%Y %I:%M %p') if ticket.created_at else ''),
        ('Updated Date', ticket.updated_at.strftime('%d-%b-%Y %I:%M %p') if ticket.updated_at else ''),
    ]
    
    for label, value in basic_info:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=1).fill = label_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
        
        ws.cell(row=row, column=2, value=value).font = data_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
    
    row += 1
    
    # Employee Details
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "EMPLOYEE DETAILS"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 30
    row += 1
    
    emp_info = [
        ('Employee Name', ticket.employee_name),
        ('Employee ID', ticket.employee_id),
        ('Mobile', ticket.mobile),
        ('Email', ticket.email),
        ('Unit', ticket.unit.full_name if ticket.unit else ''),
        ('Department', ticket.department.name if ticket.department else ''),
        ('Screen/Module', ticket.screen_number),
    ]
    
    for label, value in emp_info:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=1).fill = label_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
        
        ws.cell(row=row, column=2, value=value).font = data_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
    
    row += 1
    
    row += 1
    
    # Closing Details (if closed) - ✅ UPDATED with new error types
    if ticket.status == 'Closed':
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "CLOSING DETAILS"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 30
        row += 1
        
        time_to_close_str = ''
        if ticket.created_at and ticket.closed_at:
            duration = ticket.closed_at - ticket.created_at
            days = duration.days
            hours = duration.seconds // 3600
            minutes = (duration.seconds % 3600) // 60
            if days > 0:
                time_to_close_str = f"{days}d {hours}h {minutes}m"
            else:
                time_to_close_str = f"{hours}h {minutes}m"
        
        closing_info = [
            ('Closed By', ticket.closed_by or ''),
            ('Closed Date', ticket.closed_at.strftime('%d-%b-%Y %I:%M %p') if ticket.closed_at else ''),
            ('Main Error Type', ticket.main_error_type or 'N/A'),  # ✅ NEW
            ('Sub Error Type', ticket.sub_error_type or 'N/A'),   # ✅ NEW
            ('Closing Remarks', ticket.closing_remarks or ''),
            ('Time to Close', time_to_close_str),
        ]
        
        for label, value in closing_info:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=1).fill = label_fill
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
            
            ws.cell(row=row, column=2, value=value).font = data_font
            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            row += 1
        
        row += 1
    
    # Footer
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"Report generated on {timezone.now().strftime('%d-%b-%Y %I:%M %p')} | GPLAST Support System"
    ws[f'A{row}'].font = Font(name='Calibri', size=9, italic=True, color='666666')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    wb.save(response)
    return response


# ============================================================
# UPDATE TICKET STATUS
# ============================================================
@login_required
def update_ticket_status(request, ticket_id):
    """Update ticket status (Assign, Hold, Escalate, Close, Reopen)"""
    if request.method != 'POST':
        return redirect('ticket_detail', ticket_id=ticket_id)
    
    ticket = get_object_or_404(Ticket, id=ticket_id)
    action_type = request.POST.get('action_type')
    
    if action_type == 'Assign':
        assigned_person = request.POST.get('assigned_person')
        remarks = request.POST.get('remarks', '')
        
        if assigned_person:
            ticket.assigned_person = assigned_person
            ticket.status = 'Assigned'
            ticket.save()
            
            TicketHistory.objects.create(
                ticket=ticket,
                action=f'Assigned to {assigned_person}',
                remarks=remarks,
                performed_by=request.user.username
            )
            send_ticket_email(ticket, 'Assigned', remarks=remarks, request=request)
            messages.success(request, f'Ticket assigned to {assigned_person}')
        else:
            messages.error(request, 'Please select an employee to assign')
    
    elif action_type == 'Hold':
        hold_reason = request.POST.get('hold_reason')
        if hold_reason:
            ticket.hold_reason = hold_reason
            ticket.status = 'Hold'
            ticket.save()
            
            TicketHistory.objects.create(
                ticket=ticket,
                action='Put on Hold',
                remarks=hold_reason,
                performed_by=request.user.username
            )
            send_ticket_email(ticket, 'Hold', remarks=hold_reason)
            messages.success(request, 'Ticket put on hold')
        else:
            messages.error(request, 'Please provide a reason for holding')
    
    elif action_type == 'Escalate':
        vendor_ticket = request.POST.get('vendor_ticket_number', '')
        remarks = request.POST.get('remarks', '')
        
        if vendor_ticket:
            ticket.vendor_ticket_number = vendor_ticket
        ticket.status = 'Escalated'
        ticket.escalated_at = timezone.now()
        ticket.save()
        
        TicketHistory.objects.create(
            ticket=ticket,
            action='Escalated to Vendor',
            remarks=f'Vendor Ticket: {vendor_ticket or "N/A"} - {remarks}',
            performed_by=request.user.username
        )
        send_ticket_email(ticket, 'Escalated', remarks=remarks)
        messages.success(request, 'Ticket escalated successfully')
    
    elif action_type == 'Close':
        # ✅ UPDATED: Get new error type fields
        main_error_type = request.POST.get('main_error_type', '').strip()
        sub_error_type = request.POST.get('sub_error_type', '').strip()
        closing_remarks = request.POST.get('closing_remarks', '').strip()
        
        if main_error_type and sub_error_type and closing_remarks:
            ticket.main_error_type = main_error_type
            ticket.sub_error_type = sub_error_type
            ticket.closing_remarks = closing_remarks
            ticket.status = 'Closed'
            ticket.closed_by = request.user.username
            ticket.closed_at = timezone.now()
            ticket.save()
            
            TicketHistory.objects.create(
                ticket=ticket,
                action='Closed Ticket',
                remarks=f'Main Error: {main_error_type}\nSub Error: {sub_error_type}\nClosing Remarks: {closing_remarks}',
                performed_by=request.user.username
            )
            
            try:
                send_ticket_email(ticket, 'Closed', remarks=closing_remarks)
            except Exception as e:
                logger.error(f"Failed to send closing email: {e}")
            
            messages.success(request, 'Ticket closed successfully')
        else:
            messages.error(request, 'Please provide main error type, sub error type, and closing remarks')
    
    elif action_type == 'Reopen':
        remarks = request.POST.get('remarks', '')
        uploaded_files = request.FILES.getlist('reopen_attachments')
        
        if remarks:
            if ticket.closed_at and (timezone.now() - ticket.closed_at).total_seconds() <= 48 * 3600:
                try:
                    for uploaded_file in uploaded_files:
                        validate_attachment(uploaded_file)
                except ValidationError as error:
                    messages.error(request, str(error))
                    return redirect('ticket_detail', ticket_id=ticket_id)
                ticket.status = 'Open'
                ticket.closed_at = None
                ticket.closed_by = None
                ticket.closing_remarks = ''
                ticket.save()
                
                TicketHistory.objects.create(
                    ticket=ticket,
                    action='Reopened Ticket',
                    remarks=remarks,
                    performed_by=request.user.username
                )
                reopen_attachments = [
                    ReopenAttachment.objects.create(
                        ticket=ticket,
                        file=uploaded_file,
                        uploaded_by=request.user.username,
                    ).file
                    for uploaded_file in uploaded_files
                ]
                
                try:
                    send_ticket_email(ticket, 'Reopened', remarks=remarks, attachments=reopen_attachments)
                except Exception as e:
                    logger.error(f"Failed to send reopen email: {e}")
                
                messages.success(request, 'Ticket reopened successfully')
            else:
                messages.error(request, 'Cannot reopen ticket after 48 hours')
        else:
            messages.error(request, 'Please provide a reason for reopening')
    
    return redirect('ticket_detail', ticket_id=ticket_id)


# ============================================================
# EXPORT CLOSED TICKETS (LAST 30 DAYS) TO EXCEL
# ============================================================
@login_required
def export_closed_tickets_30_days(request):
    """
    Export closed tickets from the last 30 days to Excel with new error type fields
    """
    thirty_days_ago = timezone.now() - timedelta(days=30)
    
    tickets_qs = Ticket.objects.filter(
        status='Closed',
        closed_at__gte=thirty_days_ago
    ).order_by('-closed_at')
    
    current_tz = timezone.get_current_timezone()
    now_utc = timezone.now()
    if timezone.is_naive(now_utc):
        now_utc = timezone.make_aware(now_utc, timezone.utc)
    now_local = now_utc.astimezone(current_tz)
    report_time = now_local.strftime('%d-%b-%Y %I:%M:%S %p')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Closed_Tickets_30_Days_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Closed Tickets (30 Days)"
    
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10)
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    ws.merge_cells('A1:AA1')
    ws['A1'] = f"CLOSED TICKETS - LAST 30 DAYS"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45
    
    ws.merge_cells('A2:AA2')
    ws['A2'] = f"Generated: {report_time}  |  Total Closed Tickets: {tickets_qs.count()}  |  Period: {thirty_days_ago.strftime('%d-%b-%Y')} to {timezone.now().strftime('%d-%b-%Y')}"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    
    # ✅ UPDATED: Added Main Error Type and Sub Error Type columns
    headers = [
        'Ticket Number', 'Status', 'Unit Code', 'Unit Name', 'Department',
        'Employee ID', 'Employee Name', 'Mobile', 'Email', 'Screen/Module',
        'Subject', 'Description', 'Priority', 'Error Type', 'Created By Role',
        'Admin Creation Reason', 'Assigned Person', 'Hold Reason',
        'Main Error Type', 'Sub Error Type', 'Closing Remarks', 'Closed By',
        'Vendor Ticket Number', 'Created At', 'Closed At', 'Time to Close', 'Escalated At'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[4].height = 30
    
    row_idx = 5
    for ticket in tickets_qs:
        if ticket.created_at:
            if timezone.is_naive(ticket.created_at):
                utc_time = timezone.make_aware(ticket.created_at, timezone.utc)
            else:
                utc_time = ticket.created_at
            created_at_local = utc_time.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p')
        else:
            created_at_local = ''
        
        if ticket.closed_at:
            if timezone.is_naive(ticket.closed_at):
                utc_time = timezone.make_aware(ticket.closed_at, timezone.utc)
            else:
                utc_time = ticket.closed_at
            closed_at_local = utc_time.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p')
        else:
            closed_at_local = ''
        
        if ticket.escalated_at:
            if timezone.is_naive(ticket.escalated_at):
                utc_time = timezone.make_aware(ticket.escalated_at, timezone.utc)
            else:
                utc_time = ticket.escalated_at
            escalated_at_local = utc_time.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p')
        else:
            escalated_at_local = ''
        
        time_to_close = ''
        if ticket.created_at and ticket.closed_at:
            duration = ticket.closed_at - ticket.created_at
            days = duration.days
            hours = duration.seconds // 3600
            minutes = (duration.seconds % 3600) // 60
            if days > 0:
                time_to_close = f"{days}d {hours}h {minutes}m"
            else:
                time_to_close = f"{hours}h {minutes}m"
        
        row_data = [
            ticket.ticket_number,
            ticket.status,
            ticket.unit.code if ticket.unit else '',
            ticket.unit.full_name if ticket.unit else '',
            ticket.department.name if ticket.department else '',
            ticket.employee_id,
            ticket.employee_name,
            ticket.mobile,
            ticket.email,
            ticket.screen_number,
            ticket.subject,
            ticket.description or '',
            ticket.priority,
            ticket.error_type or '',
            ticket.created_by_role,
            ticket.admin_creation_reason or '',
            ticket.assigned_person or '',
            ticket.hold_reason or '',
            ticket.main_error_type or 'N/A',  # ✅ NEW
            ticket.sub_error_type or 'N/A',   # ✅ NEW
            ticket.closing_remarks or '',
            ticket.closed_by or '',
            ticket.vendor_ticket_number or '',
            created_at_local,
            closed_at_local,
            time_to_close,
            escalated_at_local,
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        row_idx += 1
    
    column_widths = {
        'A': 18, 'B': 14, 'C': 12, 'D': 25, 'E': 20,
        'F': 14, 'G': 22, 'H': 16, 'I': 25, 'J': 16,
        'K': 30, 'L': 40, 'M': 14, 'N': 20, 'O': 18,
        'P': 25, 'Q': 20, 'R': 20, 'S': 22, 'T': 22,
        'U': 30, 'V': 18, 'W': 18, 'X': 22, 'Y': 22,
        'Z': 16, 'AA': 22
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    wb.save(response)
    return response


# ============================================================
# GET EMPLOYEE DETAILS (AJAX)
# ============================================================
@login_required
def get_employee_details(request):
    """AJAX endpoint to get employee details by employee_id"""
    employee_id = request.GET.get('employee_id', '').strip().upper()
    
    if not employee_id:
        return JsonResponse({'error': 'Employee ID required'}, status=400)
    
    try:
        employee = EmployeeMaster.objects.filter(
            employee_id=employee_id,
            is_active=True
        ).select_related('unit', 'department').first()
        
        if not employee:
            return JsonResponse({'error': 'Employee not found'}, status=404)
        
        data = {
            'employee_name': employee.employee_name,
            'email': employee.email or '',
            'mobile': employee.mobile or '',
            'unit_id': employee.unit_id,
            'unit_name': employee.unit.full_name if employee.unit else '',
            'department_id': employee.department_id,
            'department_name': employee.department.name if employee.department else '',
        }
        return JsonResponse(data)
    
    except Exception as e:
        logger.error(f"Error fetching employee details: {e}")
        return JsonResponse({'error': str(e)}, status=500)