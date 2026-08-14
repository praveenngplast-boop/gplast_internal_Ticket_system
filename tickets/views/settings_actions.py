# tickets/views/settings_actions.py

"""
Settings Actions (POST Handlers) - All settings POST actions
- Contact
- Units
- Departments
- Emails
- Passwords
- Employees (Add, Edit, Toggle, Delete, Bulk Upload)
- Credentials (Add, Edit, Toggle, Delete)
- Downloads (Employee List, Employee Template, Credentials)
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.models import User
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.db import IntegrityError
from django.contrib import messages
from datetime import datetime
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
import logging

from tickets.models import (
    Unit, Department, AdminContact, AdminNotificationEmail, 
    EmployeeMaster, DepartmentCredential, SettingsAuditLog
)
from tickets.forms import (
    UnitForm, DepartmentForm, AdminNotificationEmailForm, 
    AdminPasswordChangeForm, AdminSetUserPasswordForm,
    AdminContactForm
)

from .utils import (
    is_admin,
    _get_contact_data,
    _get_employee_directory_data,
    _get_credentials_data,
    format_timedelta_display,
    get_client_ip,
)

logger = logging.getLogger(__name__)


# ============================================================
# AUDIT LOG HELPER FUNCTION
# ============================================================

def log_settings_change(request, action_type, setting_type, setting_name, 
                        old_value=None, new_value=None, change_summary=None, remarks=None):
    """
    Helper function to log settings changes
    """
    try:
        SettingsAuditLog.objects.create(
            performed_by=request.user if request.user.is_authenticated else None,
            performed_by_name=request.user.username if request.user.is_authenticated else 'System',
            action_type=action_type,
            setting_type=setting_type,
            setting_name=setting_name,
            old_value=str(old_value) if old_value else None,
            new_value=str(new_value) if new_value else None,
            change_summary=change_summary or '',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            remarks=remarks or '',
        )
    except Exception as e:
        logger.error(f"Failed to log settings change: {str(e)}")


# ============================================================
# CONTACT SETTINGS
# ============================================================

@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_contact(request):
    """
    Update Helpdesk Contact information
    POST: admin_name, admin_email
    Redirects to: settings_communication
    """
    if request.method == 'POST':
        contact_obj = AdminContact.objects.first()
        if not contact_obj:
            contact_obj = AdminContact.objects.create(
                admin_name="IT ADMIN",
                admin_phone="9999999999",
                admin_email="admin@gplast.com"
            )
        
        old_name = contact_obj.admin_name
        old_email = contact_obj.admin_email
        admin_name = request.POST.get('admin_name', '').strip()
        admin_email = request.POST.get('admin_email', '').strip()
        
        changed = False
        change_details = []
        
        if admin_name and admin_name != contact_obj.admin_name:
            contact_obj.admin_name = admin_name
            changed = True
            change_details.append(f"Name: {old_name} → {admin_name}")
        
        if admin_email and admin_email != contact_obj.admin_email:
            contact_obj.admin_email = admin_email
            changed = True
            change_details.append(f"Email: {old_email} → {admin_email}")
        
        if changed:
            contact_obj.save()
            messages.success(request, "IT Support Contact updated successfully.")
            
            log_settings_change(
                request,
                action_type='UPDATE',
                setting_type='CONTACT',
                setting_name='Helpdesk Contact',
                old_value=f"Name: {old_name}, Email: {old_email}",
                new_value=f"Name: {admin_name}, Email: {admin_email}",
                change_summary='; '.join(change_details),
                remarks="Contact information updated"
            )
        else:
            messages.info(request, "No changes made to contact information.")
    
    return redirect('settings_communication')


# ============================================================
# UNIT SETTINGS
# ============================================================

@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_units(request):
    """
    Manage Units: Add, Edit, Toggle Active/Inactive
    POST: action (add/edit/toggle), unit_id, code, full_name
    Redirects to: settings_units_departments
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            form = UnitForm(request.POST)
            if form.is_valid(): 
                unit = form.save(commit=False)
                unit.created_by = request.user.username
                unit.save()
                messages.success(request, f"Unit '{unit.code}' added.")
                
                log_settings_change(
                    request,
                    action_type='CREATE',
                    setting_type='UNIT',
                    setting_name=f"Unit: {unit.code}",
                    new_value=f"Code: {unit.code}, Name: {unit.full_name}",
                    change_summary=f"Added unit '{unit.code}' - {unit.full_name}",
                    remarks=f"Unit added by {request.user.username}"
                )
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        
        elif action == 'edit':
            unit = get_object_or_404(Unit, pk=request.POST.get('unit_id'))
            old_code = unit.code
            old_name = unit.full_name
            
            form = UnitForm(request.POST, instance=unit)
            if form.is_valid(): 
                form.save()
                messages.success(request, f"Unit '{unit.code}' updated.")
                
                change_details = []
                if old_code != unit.code:
                    change_details.append(f"Code: {old_code} → {unit.code}")
                if old_name != unit.full_name:
                    change_details.append(f"Name: {old_name} → {unit.full_name}")
                
                log_settings_change(
                    request,
                    action_type='UPDATE',
                    setting_type='UNIT',
                    setting_name=f"Unit: {unit.code}",
                    old_value=f"Code: {old_code}, Name: {old_name}",
                    new_value=f"Code: {unit.code}, Name: {unit.full_name}",
                    change_summary='; '.join(change_details) if change_details else 'Unit updated',
                    remarks=f"Unit updated by {request.user.username}"
                )
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        
        elif action == 'toggle':
            unit = get_object_or_404(Unit, pk=request.POST.get('unit_id'))
            old_status = 'Active' if unit.is_active else 'Inactive'
            unit.is_active = not unit.is_active
            unit.save()
            new_status = 'Active' if unit.is_active else 'Inactive'
            
            messages.success(request, f"Unit '{unit.code}' {'activated' if unit.is_active else 'deactivated'}.")
            
            log_settings_change(
                request,
                action_type='TOGGLE',
                setting_type='UNIT',
                setting_name=f"Unit: {unit.code}",
                old_value=f"Status: {old_status}",
                new_value=f"Status: {new_status}",
                change_summary=f"Status changed from {old_status} to {new_status}",
                remarks=f"Unit toggled by {request.user.username}"
            )
    
    return redirect('settings_units_departments')


# ============================================================
# DEPARTMENT SETTINGS
# ============================================================

@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_departments(request):
    """
    Manage Departments: Add, Edit, Toggle Active/Inactive
    POST: action (add/edit/toggle), dept_id, name, unit
    Redirects to: settings_units_departments
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            form = DepartmentForm(request.POST)
            if form.is_valid(): 
                dept = form.save()
                messages.success(request, f"Department '{dept.name}' added.")
                
                log_settings_change(
                    request,
                    action_type='CREATE',
                    setting_type='DEPARTMENT',
                    setting_name=f"Department: {dept.name}",
                    new_value=f"Name: {dept.name}, Unit: {dept.unit.code}",
                    change_summary=f"Added department '{dept.name}' under unit '{dept.unit.code}'",
                    remarks=f"Department added by {request.user.username}"
                )
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        
        elif action == 'edit':
            dept = get_object_or_404(Department, pk=request.POST.get('dept_id'))
            old_name = dept.name
            old_unit = dept.unit.code
            
            form = DepartmentForm(request.POST, instance=dept)
            if form.is_valid(): 
                form.save()
                messages.success(request, f"Department '{dept.name}' updated.")
                
                change_details = []
                if old_name != dept.name:
                    change_details.append(f"Name: {old_name} → {dept.name}")
                if old_unit != dept.unit.code:
                    change_details.append(f"Unit: {old_unit} → {dept.unit.code}")
                
                log_settings_change(
                    request,
                    action_type='UPDATE',
                    setting_type='DEPARTMENT',
                    setting_name=f"Department: {dept.name}",
                    old_value=f"Name: {old_name}, Unit: {old_unit}",
                    new_value=f"Name: {dept.name}, Unit: {dept.unit.code}",
                    change_summary='; '.join(change_details) if change_details else 'Department updated',
                    remarks=f"Department updated by {request.user.username}"
                )
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        
        elif action == 'toggle':
            dept = get_object_or_404(Department, pk=request.POST.get('dept_id'))
            old_status = 'Active' if dept.is_active else 'Inactive'
            dept.is_active = not dept.is_active
            dept.save()
            new_status = 'Active' if dept.is_active else 'Inactive'
            
            messages.success(request, f"Department '{dept.name}' {'activated' if dept.is_active else 'deactivated'}.")
            
            log_settings_change(
                request,
                action_type='TOGGLE',
                setting_type='DEPARTMENT',
                setting_name=f"Department: {dept.name}",
                old_value=f"Status: {old_status}",
                new_value=f"Status: {new_status}",
                change_summary=f"Status changed from {old_status} to {new_status}",
                remarks=f"Department toggled by {request.user.username}"
            )
    
    return redirect('settings_units_departments')


# ============================================================
# EMAIL SETTINGS
# ============================================================

@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_emails(request):
    """
    Manage Notification Emails: Add, Delete
    POST: action (add/delete), email, email_id
    Redirects to: settings_communication
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            form = AdminNotificationEmailForm(request.POST)
            if form.is_valid(): 
                email = form.save()
                messages.success(request, f"Email '{email.email}' added.")
                
                log_settings_change(
                    request,
                    action_type='CREATE',
                    setting_type='EMAIL',
                    setting_name=f"Email: {email.email}",
                    new_value=email.email,
                    change_summary=f"Added notification email: {email.email}",
                    remarks=f"Email added by {request.user.username}"
                )
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        
        elif action == 'delete':
            email_obj = get_object_or_404(AdminNotificationEmail, pk=request.POST.get('email_id'))
            email_str = email_obj.email
            email_obj.delete()
            messages.success(request, f"Email '{email_str}' deleted.")
            
            log_settings_change(
                request,
                action_type='DELETE',
                setting_type='EMAIL',
                setting_name=f"Email: {email_str}",
                old_value=email_str,
                change_summary=f"Deleted notification email: {email_str}",
                remarks=f"Email deleted by {request.user.username}"
            )
    
    return redirect('settings_communication')


# ============================================================
# PASSWORD SETTINGS
# ============================================================

@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_passwords(request):
    """
    Change Admin Password and Reset Employee Password
    POST: action (change_my_password / set_user_password)
    Redirects to: settings_credentials_page
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'change_my_password':
            form = AdminPasswordChangeForm(user=request.user, data=request.POST)
            if form.is_valid(): 
                user = form.save()
                update_session_auth_hash(request, user)
                messages.success(request, 'Password updated!')
                
                log_settings_change(
                    request,
                    action_type='UPDATE',
                    setting_type='PASSWORD',
                    setting_name=f"Admin: {request.user.username}",
                    change_summary="Admin password changed",
                    remarks=f"Admin password changed by {request.user.username}"
                )
            else:
                for e in form.errors.values():
                    for err in e: 
                        messages.error(request, f"Error: {err}")
        
        elif action == 'set_user_password':
            user_id = request.POST.get('user')
            if not user_id:
                messages.error(request, "Please select an employee.")
                return redirect('settings_credentials_page')
            
            selected_user = get_object_or_404(User, pk=user_id, is_staff=False)
            form = AdminSetUserPasswordForm(user=selected_user, data=request.POST)
            if form.is_valid(): 
                form.save()
                messages.success(request, f"Password reset for '{selected_user.username}'.")
                
                log_settings_change(
                    request,
                    action_type='UPDATE',
                    setting_type='PASSWORD',
                    setting_name=f"Employee: {selected_user.username}",
                    change_summary=f"Password reset for {selected_user.username}",
                    remarks=f"Employee password reset by {request.user.username}"
                )
            else:
                for e in form.errors.values():
                    for err in e: 
                        messages.error(request, f"Error: {err}")
    
    return redirect('settings_credentials_page')


# ============================================================
# EMPLOYEE MANAGEMENT (ALL ACTIONS)
# ============================================================

@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_employees(request):
    """
    Complete Employee Management Actions:
    - Add Employee
    - Edit Employee
    - Toggle Active/Inactive
    - Toggle Can Assign
    - Delete Employee (Must be deactivated first)
    - Bulk Upload (AJAX)
    POST: action, employee_id, employee_name, mobile, email, unit, department, can_assign_ticket
    Redirects to: settings_employees_page
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # ========== ADD EMPLOYEE ==========
        if action == 'add_employee':
            eid = request.POST.get('employee_id','').strip().upper()
            ename = request.POST.get('employee_name','').strip().upper()
            mob = request.POST.get('mobile','').strip()
            email = request.POST.get('email','').strip()
            uid = request.POST.get('unit')
            did = request.POST.get('department')
            
            if not all([eid,ename,mob,email]): 
                messages.error(request, "All mandatory fields are required.")
                return redirect('settings_employees_page')
            
            try:
                emp = EmployeeMaster.objects.create(
                    employee_id=eid,
                    employee_name=ename,
                    mobile=mob,
                    email=email,
                    unit_id=uid or None,
                    department_id=did or None,
                    is_active=True
                )
                messages.success(request, f'Employee "{eid}" added.')
                
                log_settings_change(
                    request,
                    action_type='CREATE',
                    setting_type='EMPLOYEE',
                    setting_name=f"Employee: {eid} - {ename}",
                    new_value=f"ID: {eid}, Name: {ename}, Mobile: {mob}, Email: {email}",
                    change_summary=f"Added employee {eid} - {ename}",
                    remarks=f"Employee added by {request.user.username}"
                )
            except IntegrityError: 
                messages.error(request, f'Employee ID "{eid}" already exists.')
        
        # ========== BULK UPLOAD ==========
        elif action == 'bulk_upload':
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': 'Please select an Excel file.'
                    })
                messages.error(request, "Please select an Excel file.")
                return redirect('settings_employees_page')
            
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': 'Invalid file format. Only .xlsx and .xls files are supported.'
                    })
                messages.error(request, "Invalid file format. Only .xlsx and .xls files are supported.")
                return redirect('settings_employees_page')
            
            try:
                df = pd.read_excel(excel_file, dtype=str)
                
                if df.empty:
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': 'The uploaded file is empty.'
                        })
                    messages.error(request, "The uploaded file is empty.")
                    return redirect('settings_employees_page')
                
                required_columns = ['Employee ID', 'Employee Name', 'Mobile', 'Email']
                missing_columns = []
                for col in required_columns:
                    if col not in df.columns:
                        found = False
                        for existing_col in df.columns:
                            if existing_col.lower() == col.lower():
                                found = True
                                break
                        if not found:
                            missing_columns.append(col)
                
                if missing_columns:
                    error_msg = f"Missing required columns: {', '.join(missing_columns)}"
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': error_msg
                        })
                    messages.error(request, error_msg)
                    return redirect('settings_employees_page')
                
                validation_errors = []
                successful_rows = 0
                
                all_units = {unit.code.upper(): unit for unit in Unit.objects.filter(is_active=True)}
                all_departments = {dept.name.upper(): dept for dept in Department.objects.filter(is_active=True)}
                
                for idx, row in df.iterrows():
                    row_num = idx + 2
                    
                    eid = None
                    ename = None
                    mob = None
                    email = None
                    uc = None
                    dn = None
                    
                    for col in df.columns:
                        col_lower = col.lower()
                        if col_lower in ['employee id', 'employee_id', 'employeeid']:
                            eid = str(row.get(col, '')).strip().upper()
                        elif col_lower in ['employee name', 'employee_name', 'employeename', 'name']:
                            ename = str(row.get(col, '')).strip().upper()
                        elif col_lower in ['mobile', 'phone', 'contact']:
                            mob = str(row.get(col, '')).strip()
                        elif col_lower in ['email', 'email id', 'email_id', 'emailid']:
                            email = str(row.get(col, '')).strip().lower()
                        elif col_lower in ['unit code', 'unit_code', 'unitcode', 'unit']:
                            uc = str(row.get(col, '')).strip().upper()
                        elif col_lower in ['department', 'dept', 'department name', 'department_name']:
                            dn = str(row.get(col, '')).strip().upper()
                    
                    if not eid:
                        validation_errors.append({'row': row_num, 'message': 'Employee ID is required'})
                        continue
                    if not ename:
                        validation_errors.append({'row': row_num, 'message': 'Employee Name is required'})
                        continue
                    if not mob:
                        validation_errors.append({'row': row_num, 'message': 'Mobile number is required'})
                        continue
                    if not email:
                        validation_errors.append({'row': row_num, 'message': 'Email is required'})
                        continue
                    if '@' not in email or '.' not in email:
                        validation_errors.append({'row': row_num, 'message': f'Invalid email format: {email}'})
                        continue
                    if not mob.isdigit() or len(mob) != 10:
                        validation_errors.append({'row': row_num, 'message': f'Mobile number must be 10 digits: {mob}'})
                        continue
                    if uc and uc not in all_units:
                        valid_units = ', '.join(list(all_units.keys())[:5])
                        if len(all_units) > 5:
                            valid_units += f' and {len(all_units) - 5} more'
                        validation_errors.append({'row': row_num, 'message': f'Invalid Unit Code "{uc}". Valid units: {valid_units}'})
                        continue
                    if dn and dn not in all_departments:
                        valid_depts = ', '.join(list(all_departments.keys())[:5])
                        if len(all_departments) > 5:
                            valid_depts += f' and {len(all_departments) - 5} more'
                        validation_errors.append({'row': row_num, 'message': f'Invalid Department "{dn}". Valid departments: {valid_depts}'})
                        continue
                    
                    successful_rows += 1
                
                if validation_errors:
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': f'Validation failed. Found {len(validation_errors)} error(s).',
                            'errors': validation_errors
                        })
                    messages.error(request, f'Validation failed. Found {len(validation_errors)} error(s).')
                    for err in validation_errors[:5]:
                        messages.error(request, f'Row {err["row"]}: {err["message"]}')
                    if len(validation_errors) > 5:
                        messages.error(request, f'... and {len(validation_errors) - 5} more errors.')
                    return redirect('settings_employees_page')
                
                success_count = 0
                error_count = 0
                
                for idx, row in df.iterrows():
                    try:
                        eid = None
                        ename = None
                        mob = None
                        email = None
                        uc = None
                        dn = None
                        
                        for col in df.columns:
                            col_lower = col.lower()
                            if col_lower in ['employee id', 'employee_id', 'employeeid']:
                                eid = str(row.get(col, '')).strip().upper()
                            elif col_lower in ['employee name', 'employee_name', 'employeename', 'name']:
                                ename = str(row.get(col, '')).strip().upper()
                            elif col_lower in ['mobile', 'phone', 'contact']:
                                mob = str(row.get(col, '')).strip()
                            elif col_lower in ['email', 'email id', 'email_id', 'emailid']:
                                email = str(row.get(col, '')).strip().lower()
                            elif col_lower in ['unit code', 'unit_code', 'unitcode', 'unit']:
                                uc = str(row.get(col, '')).strip().upper()
                            elif col_lower in ['department', 'dept', 'department name', 'department_name']:
                                dn = str(row.get(col, '')).strip().upper()
                        
                        if not eid or not ename or not mob or not email:
                            error_count += 1
                            continue
                        
                        unit_obj = all_units.get(uc) if uc else None
                        dept_obj = all_departments.get(dn) if dn else None
                        
                        EmployeeMaster.objects.update_or_create(
                            employee_id=eid,
                            defaults={
                                'employee_name': ename,
                                'mobile': mob,
                                'email': email,
                                'unit': unit_obj,
                                'department': dept_obj,
                                'is_active': True
                            }
                        )
                        success_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        logger.error(f"Error processing row {idx+2}: {str(e)}")
                
                # Log bulk upload summary
                if success_count > 0:
                    log_settings_change(
                        request,
                        action_type='CREATE',
                        setting_type='EMPLOYEE',
                        setting_name=f"Bulk Upload: {success_count} employees",
                        new_value=f"Successfully uploaded {success_count} employees",
                        change_summary=f"Bulk uploaded {success_count} employees",
                        remarks=f"Bulk upload by {request.user.username}"
                    )
                
                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'message': f'Successfully processed {success_count} employees. {error_count} skipped.'
                    })
                
                if success_count > 0:
                    messages.success(request, f'Successfully uploaded {success_count} employees.')
                if error_count > 0:
                    messages.warning(request, f'{error_count} rows were skipped due to errors.')
                
            except Exception as e:
                logger.error(f"Bulk upload error: {str(e)}")
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': f'Error processing file: {str(e)}'
                    })
                messages.error(request, f'Error processing file: {str(e)}')
        
        # ========== EDIT EMPLOYEE ==========
        elif action == 'edit_employee':
            emp = get_object_or_404(EmployeeMaster, pk=request.POST.get('emp_id'))
            
            old_id = emp.employee_id
            old_name = emp.employee_name
            old_mobile = emp.mobile
            old_email = emp.email
            old_unit = emp.unit.code if emp.unit else 'None'
            old_dept = emp.department.name if emp.department else 'None'
            old_can_assign = 'Yes' if emp.can_assign_ticket else 'No'
            
            emp.employee_id = request.POST.get('employee_id','').strip().upper()
            emp.employee_name = request.POST.get('employee_name','').strip().upper()
            emp.mobile = request.POST.get('mobile','').strip()
            emp.email = request.POST.get('email','').strip()
            emp.unit_id = request.POST.get('unit') or None
            emp.department_id = request.POST.get('department') or None
            emp.can_assign_ticket = request.POST.get('can_assign_ticket') == 'on'
            
            try: 
                emp.save()
                messages.success(request, 'Employee updated successfully.')
                
                change_details = []
                if old_id != emp.employee_id:
                    change_details.append(f"ID: {old_id} → {emp.employee_id}")
                if old_name != emp.employee_name:
                    change_details.append(f"Name: {old_name} → {emp.employee_name}")
                if old_mobile != emp.mobile:
                    change_details.append(f"Mobile: {old_mobile} → {emp.mobile}")
                if old_email != emp.email:
                    change_details.append(f"Email: {old_email} → {emp.email}")
                if old_unit != (emp.unit.code if emp.unit else 'None'):
                    change_details.append(f"Unit: {old_unit} → {emp.unit.code if emp.unit else 'None'}")
                if old_dept != (emp.department.name if emp.department else 'None'):
                    change_details.append(f"Department: {old_dept} → {emp.department.name if emp.department else 'None'}")
                if old_can_assign != ('Yes' if emp.can_assign_ticket else 'No'):
                    change_details.append(f"Can Assign: {old_can_assign} → {'Yes' if emp.can_assign_ticket else 'No'}")
                
                log_settings_change(
                    request,
                    action_type='UPDATE',
                    setting_type='EMPLOYEE',
                    setting_name=f"Employee: {emp.employee_id} - {emp.employee_name}",
                    old_value=f"ID: {old_id}, Name: {old_name}, Mobile: {old_mobile}, Email: {old_email}",
                    new_value=f"ID: {emp.employee_id}, Name: {emp.employee_name}, Mobile: {emp.mobile}, Email: {emp.email}",
                    change_summary='; '.join(change_details) if change_details else 'Employee updated',
                    remarks=f"Employee updated by {request.user.username}"
                )
            except IntegrityError: 
                messages.error(request, 'Employee ID already exists.')
        
        # ========== TOGGLE EMPLOYEE ACTIVE/INACTIVE ==========
        elif action == 'toggle_employee':
            emp = get_object_or_404(EmployeeMaster, pk=request.POST.get('emp_id'))
            old_status = 'Active' if emp.is_active else 'Inactive'
            emp.is_active = not emp.is_active
            emp.save()
            new_status = 'Active' if emp.is_active else 'Inactive'
            
            messages.success(request, f'Employee {"activated" if emp.is_active else "deactivated"}.')
            
            log_settings_change(
                request,
                action_type='TOGGLE',
                setting_type='EMPLOYEE',
                setting_name=f"Employee: {emp.employee_id} - {emp.employee_name}",
                old_value=f"Status: {old_status}",
                new_value=f"Status: {new_status}",
                change_summary=f"Status changed from {old_status} to {new_status}",
                remarks=f"Employee toggled by {request.user.username}"
            )
        
        # ========== TOGGLE CAN ASSIGN ==========
        elif action == 'toggle_can_assign':
            emp = get_object_or_404(EmployeeMaster, pk=request.POST.get('emp_id'))
            old_value = 'Yes' if emp.can_assign_ticket else 'No'
            emp.can_assign_ticket = not emp.can_assign_ticket
            emp.save()
            new_value = 'Yes' if emp.can_assign_ticket else 'No'
            
            status = "enabled" if emp.can_assign_ticket else "disabled"
            messages.success(request, f'Employee "{emp.employee_id}" assignment {status}.')
            
            log_settings_change(
                request,
                action_type='UPDATE',
                setting_type='EMPLOYEE',
                setting_name=f"Employee: {emp.employee_id} - {emp.employee_name}",
                old_value=f"Can Assign: {old_value}",
                new_value=f"Can Assign: {new_value}",
                change_summary=f"Can Assign changed from {old_value} to {new_value}",
                remarks=f"Assignment toggled by {request.user.username}"
            )
        
        # ========== DELETE EMPLOYEE - MUST BE DEACTIVATED FIRST ==========
        elif action == 'delete_employee':
            emp = get_object_or_404(EmployeeMaster, pk=request.POST.get('emp_id'))
            eid = emp.employee_id
            ename = emp.employee_name
            
            # ✅ Check if employee is active
            if emp.is_active:
                messages.error(request, f'❌ Cannot delete "{eid}" because they are still ACTIVE. Please deactivate the employee first, then try deleting again.')
                return redirect('settings_employees_page')
            
            # ✅ Employee is inactive, proceed with deletion
            emp_mobile = emp.mobile
            emp_email = emp.email
            
            # ✅ Log before deleting
            log_settings_change(
                request,
                action_type='DELETE',
                setting_type='EMPLOYEE',
                setting_name=f"Employee: {eid} - {ename}",
                old_value=f"ID: {eid}, Name: {ename}, Mobile: {emp_mobile}, Email: {emp_email}, Status: Inactive",
                change_summary=f"Permanently deleted employee {eid} - {ename}",
                remarks=f"Employee permanently deleted by {request.user.username} (was already deactivated)"
            )
            
            # ✅ Hard delete from database
            emp.delete()
            
            messages.success(request, f'✅ Employee "{eid}" has been permanently deleted from the database.')
    
    return redirect('settings_employees_page')


# ============================================================
# DOWNLOAD EMPLOYEE LIST
# ============================================================

@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def download_employee_list(request):
    """
    Download complete employee list as Excel with all fields
    Includes: Employee ID, Name, Mobile, Email, Unit, Department, Status, Can Assign
    """
    emps = EmployeeMaster.objects.all().order_by('employee_id')
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=Employee_List_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    tf = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    hf = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    df = Font(name='Calibri', size=11)
    tfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hfill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    
    ws.merge_cells('A1:I1')
    ws['A1'] = "Employee Directory"
    ws['A1'].font = tf
    ws['A1'].fill = tfill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    for ci, h in enumerate(['Employee ID', 'Name', 'Mobile', 'Email', 'Unit Code', 'Unit Name', 'Department', 'Status', 'Can Assign'], 1):
        c = ws.cell(row=3, column=ci)
        c.value = h
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 25
    
    for ri, emp in enumerate(emps, 4):
        rd = [
            emp.employee_id,
            emp.employee_name,
            emp.mobile,
            emp.email,
            emp.unit.code if emp.unit else '',
            emp.unit.full_name if emp.unit else '',
            emp.department.name if emp.department else '',
            'Active' if emp.is_active else 'Inactive',
            'Yes' if emp.can_assign_ticket else 'No'
        ]
        for ci, v in enumerate(rd, 1):
            c = ws.cell(row=ri, column=ci)
            c.value = v
            c.font = df
    
    for col in ws.columns:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(
            max(len(str(c.value or '')) for c in col if c.row > 1) + 3, 12
        )
    wb.save(resp)
    return resp


# ============================================================
# DOWNLOAD EMPLOYEE TEMPLATE
# ============================================================

@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def download_employee_template(request):
    """
    Download Excel template for bulk employee upload
    Columns: Employee ID, Employee Name, Mobile, Email, Unit Code, Department
    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Employee_Upload_Template.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Template"
    
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    headers = ['Employee ID', 'Employee Name', 'Mobile', 'Email', 'Unit Code', 'Department']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    sample_data = [
        ['EMP001', 'JOHN DOE', '9876543210', 'john.doe@company.com', 'GPL', 'Production'],
        ['EMP002', 'JANE SMITH', '9876543211', 'jane.smith@company.com', 'GPLAST', 'QA'],
        ['EMP003', 'MIKE JOHNSON', '9876543212', 'mike.johnson@company.com', 'IMD', 'Purchase'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = Font(name='Calibri', size=11)
    
    note_row = len(sample_data) + 3
    note_cell = ws.cell(row=note_row, column=1)
    note_cell.value = "Mandatory fields: Employee ID, Employee Name, Mobile, Email"
    note_cell.font = Font(name='Calibri', size=10, italic=True, color='FF0000')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
    
    validation_row = note_row + 1
    validation_cell = ws.cell(row=validation_row, column=1)
    validation_cell.value = "Note: Unit Code must match existing active units in the system. Department must match existing active departments."
    validation_cell.font = Font(name='Calibri', size=9, italic=True, color='666666')
    ws.merge_cells(start_row=validation_row, start_column=1, end_row=validation_row, end_column=6)
    
    column_widths = [15, 20, 15, 25, 12, 15]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    wb.save(response)
    return response


# ============================================================
# CREDENTIALS MANAGEMENT (ALL ACTIONS)
# ============================================================

@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_credentials(request):
    """
    Manage Department Credentials:
    - Add Credential
    - Edit Credential
    - Toggle Active/Inactive
    - Delete Credential
    POST: action (add_credential, edit_credential, toggle_credential, delete_credential)
    Redirects to: settings_credentials_page
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # ========== ADD CREDENTIAL ==========
        if action == 'add_credential':
            uid = request.POST.get('unit')
            did = request.POST.get('department')
            uname = request.POST.get('username','').strip()
            pwd = request.POST.get('password','').strip()
            
            if not all([uid, did, uname, pwd]): 
                messages.error(request, "All fields are required.")
                return redirect('settings_credentials_page')
            
            if DepartmentCredential.objects.filter(unit_id=uid, department_id=did).exists():
                messages.error(request, "Credential already exists for this department.")
                return redirect('settings_credentials_page')
            
            try:
                cred = DepartmentCredential.objects.create(
                    unit_id=uid,
                    department_id=did,
                    username=uname,
                    password=pwd
                )
                if not User.objects.filter(username=uname).exists():
                    User.objects.create_user(username=uname, password=pwd, is_staff=False)
                u = Unit.objects.get(pk=uid)
                d = Department.objects.get(pk=did)
                messages.success(request, f'Credential for {u.code}-{d.name} added successfully!')
                
                log_settings_change(
                    request,
                    action_type='CREATE',
                    setting_type='CREDENTIAL',
                    setting_name=f"Credential: {u.code} - {d.name}",
                    new_value=f"Username: {uname}, Unit: {u.code}, Department: {d.name}",
                    change_summary=f"Added credential for {u.code} - {d.name}",
                    remarks=f"Credential added by {request.user.username}"
                )
            except Exception as ex: 
                messages.error(request, f'Error: {ex}')
        
        # ========== EDIT CREDENTIAL ==========
        elif action == 'edit_credential':
            cred = get_object_or_404(DepartmentCredential, pk=request.POST.get('cred_id'))
            ou = cred.username
            nu = request.POST.get('username','').strip()
            np = request.POST.get('password','').strip()
            
            old_username = cred.username
            old_password = cred.password
            
            cred.username = nu
            if np: 
                cred.password = np
            try:
                cred.save()
                user = User.objects.filter(username=ou).first()
                if user:
                    if ou != nu:
                        user.username = nu
                    if np:
                        user.set_password(np)
                    user.save()
                elif not User.objects.filter(username=nu).exists():
                    User.objects.create_user(username=nu, password=np or cred.password, is_staff=False)
                messages.success(request, 'Credential updated successfully!')
                
                change_details = []
                if old_username != nu:
                    change_details.append(f"Username: {old_username} → {nu}")
                if np:
                    change_details.append("Password changed")
                
                log_settings_change(
                    request,
                    action_type='UPDATE',
                    setting_type='CREDENTIAL',
                    setting_name=f"Credential: {cred.unit.code} - {cred.department.name}",
                    old_value=f"Username: {old_username}",
                    new_value=f"Username: {nu}",
                    change_summary='; '.join(change_details) if change_details else 'Credential updated',
                    remarks=f"Credential updated by {request.user.username}"
                )
            except Exception as ex: 
                messages.error(request, f'Error: {ex}')
        
        # ========== TOGGLE CREDENTIAL ==========
        elif action == 'toggle_credential':
            cred = get_object_or_404(DepartmentCredential, pk=request.POST.get('cred_id'))
            old_status = 'Active' if cred.is_active else 'Inactive'
            cred.is_active = not cred.is_active
            cred.save()
            new_status = 'Active' if cred.is_active else 'Inactive'
            
            user = User.objects.filter(username=cred.username).first()
            if user:
                user.is_active = cred.is_active
                user.save()
            messages.success(request, f'Credential {"activated" if cred.is_active else "deactivated"}.')
            
            log_settings_change(
                request,
                action_type='TOGGLE',
                setting_type='CREDENTIAL',
                setting_name=f"Credential: {cred.unit.code} - {cred.department.name}",
                old_value=f"Status: {old_status}",
                new_value=f"Status: {new_status}",
                change_summary=f"Status changed from {old_status} to {new_status}",
                remarks=f"Credential toggled by {request.user.username}"
            )
        
        # ========== DELETE CREDENTIAL ==========
        elif action == 'delete_credential':
            cred = get_object_or_404(DepartmentCredential, pk=request.POST.get('cred_id'))
            info = f'{cred.unit.code}-{cred.department.name}'
            uname = cred.username
            
            user = User.objects.filter(username=uname).first()
            if user:
                user.is_active = False
                user.save()
            cred.delete()
            messages.success(request, f'Credential for {info} deleted.')
            
            log_settings_change(
                request,
                action_type='DELETE',
                setting_type='CREDENTIAL',
                setting_name=f"Credential: {info}",
                old_value=f"Username: {uname}",
                change_summary=f"Deleted credential for {info}",
                remarks=f"Credential deleted by {request.user.username}"
            )
    
    return redirect('settings_credentials_page')


# ============================================================
# DOWNLOAD CREDENTIALS
# ============================================================

@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def download_credentials(request):
    """
    Download all department credentials as Excel
    Includes: Unit Code, Unit Name, Department, Username, Password, Status
    """
    creds = DepartmentCredential.objects.all().select_related('unit', 'department').order_by('unit__code', 'department__name')
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=Credentials_{timezone.now().strftime("%Y%m%d")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credentials"
    
    tf = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    hf = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    df = Font(name='Calibri', size=11)
    tfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hfill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    
    ws.merge_cells('A1:F1')
    ws['A1'] = "Department Credentials"
    ws['A1'].font = tf
    ws['A1'].fill = tfill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    for ci, h in enumerate(['Unit Code', 'Unit Name', 'Department', 'Username', 'Password', 'Status'], 1):
        c = ws.cell(row=3, column=ci)
        c.value = h
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 25
    
    for ri, cred in enumerate(creds, 4):
        rd = [
            cred.unit.code,
            cred.unit.full_name,
            cred.department.name,
            cred.username,
            cred.password,
            'Active' if cred.is_active else 'Inactive'
        ]
        for ci, v in enumerate(rd, 1):
            c = ws.cell(row=ri, column=ci)
            c.value = v
            c.font = df
    
    for col in ws.columns:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(
            max(len(str(c.value or '')) for c in col if c.row > 1) + 3, 12
        )
    wb.save(resp)
    return resp