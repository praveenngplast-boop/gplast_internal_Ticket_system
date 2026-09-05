# tickets/views/admin_views.py

"""
Admin Views - Dashboard, Create Ticket, All Tickets, Ticket Detail
(Reports views moved to reports_views.py)
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Count, Q, DateField, Value, CharField, OuterRef, Subquery
from django.db.models.functions import TruncMonth, Cast, Coalesce
from django.db import transaction
from django.core.exceptions import ValidationError
from django.contrib import messages
from django.template.loader import render_to_string
from django.template import TemplateDoesNotExist
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging
import json

from tickets.models import (
    Unit, Department, Ticket, TicketHistory, EmployeeMaster,
    AdminContact, AdminNotificationEmail, SettingsAuditLog, ScreenMaster,
    ERPHolderMapping, ReopenAttachment, UnitHead
)
from tickets.forms import AdminTicketForm, CloseTicketForm
from tickets.utils import send_ticket_email, validate_attachment

from .utils import (
    is_admin,
    format_timedelta_display,
    reopen_ticket_logic,
    generate_admin_ticket_list_html,
)

# ✅ Import reports views from separate file
from .reports_views import reports, download_ticket_excel, export_closed_tickets_30_days

logger = logging.getLogger(__name__)


# ============================================================
# ADMIN DASHBOARD
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def admin_dashboard(request):
    """
    Admin dashboard showing:
    - KPIs for all tickets
    - Charts: Status, Unit, Priority, Error Type, Monthly
    - Notification badge count
    """
    all_tickets = Ticket.objects.all()
    
    unviewed_count = Ticket.objects.filter(is_viewed=False).count()
    unviewed_tickets = Ticket.objects.filter(is_viewed=False).order_by('-created_at')[:10]
    
    # ✅ NEW: Get total Unit Heads count
    total_unit_heads = UnitHead.objects.filter(is_active=True).count()
    
    kpis = {
        'total': all_tickets.count(), 
        'open': all_tickets.filter(status='Open').count(),
        'assigned': all_tickets.filter(status='Assigned').count(), 
        'hold': all_tickets.filter(status='Hold').count(),
        'escalated': all_tickets.filter(status='Escalated').count(), 
        'closed': all_tickets.filter(status='Closed').count(),
        'critical': all_tickets.filter(priority='Critical').count(),
        'unviewed': unviewed_count,
        'unit_heads': total_unit_heads,
    }
    
    status_counts = list(all_tickets.values('status').annotate(count=Count('id')))
    chart_status = {item['status']: item['count'] for item in status_counts}
    
    unit_counts = list(all_tickets.filter(unit__isnull=False).values('unit_id', 'unit__code').annotate(count=Count('id')).order_by('unit__code'))
    chart_units = [{'id': item['unit_id'], 'label': item['unit__code'], 'value': item['count']} for item in unit_counts]
    
    prio_counts = list(all_tickets.values('priority').annotate(count=Count('id')))
    chart_priority = {item['priority']: item['count'] for item in prio_counts}
    
    closed_tickets = Ticket.objects.filter(status='Closed')
    
    # Show main_error_type and sub_error_type stats
    main_error_counts = (
        closed_tickets
        .exclude(main_error_type__isnull=True)
        .exclude(main_error_type__exact='')
        .values('main_error_type')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    chart_main_error = {item['main_error_type']: item['count'] for item in main_error_counts}
    
    sub_error_counts = (
        closed_tickets
        .exclude(sub_error_type__isnull=True)
        .exclude(sub_error_type__exact='')
        .values('sub_error_type')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    chart_sub_error = {item['sub_error_type']: item['count'] for item in sub_error_counts}
    
    twelve_months_ago = timezone.now() - timedelta(days=365)
    monthly_counts_qs = Ticket.objects.filter(created_at__gte=twelve_months_ago).annotate(
        month=TruncMonth(Cast('created_at', output_field=DateField()))
    ).values('month').annotate(count=Count('id')).order_by('month')
    chart_monthly = [{'label': item['month'].strftime('%b %Y'), 'value': item['count']} for item in monthly_counts_qs]
    
    charts_data = {
        'status': chart_status,
        'units': chart_units,
        'priority': chart_priority,
        'mainErrorType': chart_main_error,
        'subErrorType': chart_sub_error,
        'monthly': chart_monthly,
    }
    
    context = {
        'kpis': kpis,
        'charts_data': charts_data,
        'unviewed_count': unviewed_count,
        'unviewed_tickets': unviewed_tickets,
    }
    return render(request, 'admin_panel/dashboard.html', context)


# ============================================================
# CREATE TICKET - ADMIN (FIXED: Employee Details Fetch)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def create_ticket_admin(request):
    """
    Admin ticket creation with:
    - Employee assignment options
    - Admin creation reason
    - Admin or Employee role selection
    - 3 attachment fields
    - ✅ FIXED: Proper employee details fetch with ERP ID
    """
    employees = EmployeeMaster.objects.filter(is_active=True, can_assign_ticket=True).order_by('employee_name')
    
    if request.method == 'POST':
        form = AdminTicketForm(request.POST, request.FILES)
        if form.is_valid():
            with transaction.atomic():
                ticket = form.save(commit=False)
                
                if ticket.created_by_role == 'Admin':
                    hist_remark = f"Ticket created by Admin on behalf of employee (Reason: {ticket.admin_creation_reason})"
                    hist_perf = f"Admin {request.user.username}"
                    ticket.created_by_user = request.user
                else:
                    hist_remark = "Ticket Created by Employee (logged by Admin)"
                    hist_perf = "Employee"
                    employee_user, _ = User.objects.get_or_create(
                        username='GPLERPUSERS', 
                        defaults={'is_staff': False, 'password': 'pbkdf2_sha256$720000$j5xL6pS0LpGvLq3sRjVbWk$V/Hq7aYt2x531enqYm5d9f2uZdtsJ7MLd2y221C+L9s='}
                    )
                    ticket.created_by_user = employee_user
                
                ticket.save()
                
                TicketHistory.objects.create(
                    ticket=ticket, 
                    action="Ticket Created", 
                    remarks=hist_remark, 
                    performed_by=hist_perf
                )
            
            # ============================================================
            # EMAIL SENDING DISABLED - COMMENTED OUT
            # ============================================================
            # send_ticket_email(ticket, 'Created')
            # ============================================================
            
            messages.success(request, f'Ticket {ticket.ticket_number} created successfully by Admin!')
            return redirect('admin_dashboard')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = AdminTicketForm()
        # ✅ Set initial error_type choices for GET request
        form.fields['error_type'].choices = [('', 'Select Error Type'), ('New', 'New'), ('Repeated', 'Repeated')]
    
    all_screens = ScreenMaster.objects.all().order_by('screen_name')
    
    # ✅ Get all units and departments for dropdown
    units = Unit.objects.filter(is_active=True).order_by('code')
    departments = Department.objects.filter(is_active=True).order_by('unit__code', 'name')
    
    return render(request, 'admin_panel/create_ticket.html', {
        'form': form,
        'employees': employees,
        'all_screens': all_screens,
        'units': units,
        'departments': departments,
    })


# ============================================================
# ALL TICKETS - ADMIN (FIXED AJAX RESPONSE)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def all_tickets(request):
    """
    Admin ticket listing with filters, AJAX support, and pagination
    """
    is_ajax = request.GET.get('ajax', False)
    
    if isinstance(is_ajax, str):
        is_ajax = is_ajax.lower() in ['true', '1', 'yes']
    
    # ✅ Annotate tickets with ERP ID from ERPHolderMapping
    erp_subquery = ERPHolderMapping.objects.filter(
        employee__employee_id=OuterRef('employee_id')
    ).values('erp_user_id')[:1]
    
    tickets_qs = Ticket.objects.all().order_by('-created_at').annotate(
        erp_id=Coalesce(Subquery(erp_subquery, output_field=CharField()), Value('Not Mapped'))
    )
    
    units = Unit.objects.filter(is_active=True)
    departments = Department.objects.filter(is_active=True)
    category = request.GET.get('category', 'all')
    status = request.GET.get('status')
    priority = request.GET.get('priority')
    unit_id = request.GET.get('unit')
    dept_id = request.GET.get('department')
    assigned_person = request.GET.get('assigned_person', '').strip()
    created_by_role = request.GET.get('created_by_role')
    ticket_number = request.GET.get('ticket_number', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    search = request.GET.get('search', '').strip()
    
    # ✅ Main Error Type and Sub Error Type filters
    main_error_type = request.GET.get('main_error_type', '').strip()
    sub_error_type = request.GET.get('sub_error_type', '').strip()
    
    # ✅ ERP ID filter
    erp_id = request.GET.get('erp_id', '').strip()
    
    # ✅ Screen filter
    screen_number = request.GET.get('screen_number', '').strip()
    
    # ✅ Filter parameter for drill-down
    filter_param = request.GET.get('filter', '').strip()
    
    is_viewed = request.GET.get('is_viewed', '')
    if is_viewed == 'false':
        tickets_qs = tickets_qs.filter(is_viewed=False)
    elif is_viewed == 'true':
        tickets_qs = tickets_qs.filter(is_viewed=True)
    
    employees = EmployeeMaster.objects.filter(is_active=True, can_assign_ticket=True).order_by('employee_name')
    
    if category == 'open': 
        tickets_qs = tickets_qs.filter(status='Open')
    elif category == 'assigned': 
        tickets_qs = tickets_qs.filter(status='Assigned')
    elif category == 'hold': 
        tickets_qs = tickets_qs.filter(status='Hold')
    elif category == 'escalated': 
        tickets_qs = tickets_qs.filter(status='Escalated')
    elif category == 'closed': 
        tickets_qs = tickets_qs.filter(status='Closed')
    elif category == 'critical': 
        tickets_qs = tickets_qs.filter(priority='Critical')
    
    # ✅ Apply filter parameter for drill-down
    if filter_param and filter_param != 'all':
        if filter_param == 'Open':
            tickets_qs = tickets_qs.filter(status='Open')
        elif filter_param == 'Assigned':
            tickets_qs = tickets_qs.filter(status='Assigned')
        elif filter_param == 'Hold':
            tickets_qs = tickets_qs.filter(status='Hold')
        elif filter_param == 'Escalated':
            tickets_qs = tickets_qs.filter(status='Escalated')
        elif filter_param == 'Closed':
            tickets_qs = tickets_qs.filter(status='Closed')
        elif filter_param == 'Critical':
            tickets_qs = tickets_qs.filter(priority='Critical')
    
    if unit_id: 
        tickets_qs = tickets_qs.filter(unit_id=unit_id)
    if dept_id: 
        tickets_qs = tickets_qs.filter(department_id=dept_id)
    if status: 
        tickets_qs = tickets_qs.filter(status=status)
    if priority: 
        tickets_qs = tickets_qs.filter(priority=priority)
    if assigned_person: 
        tickets_qs = tickets_qs.filter(assigned_person__icontains=assigned_person)
    if created_by_role: 
        tickets_qs = tickets_qs.filter(created_by_role=created_by_role)
    if ticket_number: 
        tickets_qs = tickets_qs.filter(ticket_number__icontains=ticket_number)
    
    # ✅ Apply Main Error Type filter
    if main_error_type and main_error_type != '':
        tickets_qs = tickets_qs.filter(main_error_type=main_error_type)
    
    # ✅ Apply Sub Error Type filter
    if sub_error_type and sub_error_type != '' and sub_error_type != 'All':
        tickets_qs = tickets_qs.filter(sub_error_type=sub_error_type)
        
    # ✅ Apply Screen filter
    if screen_number:
        tickets_qs = tickets_qs.filter(screen_number=screen_number)
    
    # ✅ Apply ERP ID filter
    if erp_id and erp_id != '':
        employee_ids_with_erp = ERPHolderMapping.objects.filter(
            erp_user_id__icontains=erp_id
        ).values_list('employee__employee_id', flat=True).distinct()
        tickets_qs = tickets_qs.filter(employee_id__in=employee_ids_with_erp)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            from_datetime = timezone.make_aware(
                datetime.combine(date_from_obj, datetime.min.time())
            )
            tickets_qs = tickets_qs.filter(created_at__gte=from_datetime)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            to_datetime = timezone.make_aware(
                datetime.combine(date_to_obj, datetime.max.time())
            )
            tickets_qs = tickets_qs.filter(created_at__lte=to_datetime)
        except ValueError:
            pass
    
    if search: 
        tickets_qs = tickets_qs.filter(
            Q(ticket_number__icontains=search) | 
            Q(subject__icontains=search) | 
            Q(unit__code__icontains=search) | 
            Q(department__name__icontains=search)
        )
    
    thirty_days_ago = timezone.now() - timedelta(days=30)
    closed_count_30_days = Ticket.objects.filter(
        status='Closed',
        closed_at__gte=thirty_days_ago
    ).count()
    
    # ============================================================
    # ✅ FIXED: AJAX RESPONSE - Return JSON with ticket data
    # ============================================================
    if is_ajax:
        try:
            tickets = tickets_qs[:50]
            
            # Build ticket data with target date
            tickets_data = []
            for ticket in tickets:
                tickets_data.append({
                    'id': ticket.id,
                    'ticket_number': ticket.ticket_number,
                    'subject': ticket.subject,
                    'employee_name': ticket.employee_name,
                    'status': ticket.status,
                    'priority': ticket.priority,
                    'target_date': ticket.target_date.isoformat() if ticket.target_date else None,
                    'created_at': ticket.created_at.isoformat(),
                    'unit': ticket.unit.code if ticket.unit else '',
                    'department': ticket.department.name if ticket.department else '',
                })
            
            return JsonResponse({
                'success': True,
                'tickets': tickets_data,
                'count': tickets_qs.count()
            })
            
        except Exception as e:
            logger.error(f"AJAX error in all_tickets: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': str(e),
                'message': 'Error loading tickets. Please try again.',
                'tickets': [],
                'count': 0
            }, status=500)
    
    # ============================================================
    # REGULAR PAGE RENDER
    # ============================================================
    paginator = Paginator(tickets_qs, 20)
    page_number = request.GET.get('page')
    try: 
        tickets_page = paginator.page(page_number)
    except PageNotAnInteger: 
        tickets_page = paginator.page(1)
    except EmptyPage: 
        tickets_page = paginator.page(paginator.num_pages)
    
    context = {
        'page_obj': tickets_page,
        'units': units, 
        'departments': departments,
        'employees': employees,
        'status_choices': Ticket.STATUS_CHOICES, 
        'priority_choices': Ticket.PRIORITY_CHOICES,
        'created_by_choices': Ticket.CREATED_BY_CHOICES, 
        'selected_status': status,
        'selected_priority': priority, 
        'selected_unit': unit_id, 
        'selected_department': dept_id,
        'selected_assigned_person': assigned_person, 
        'selected_created_by_role': created_by_role,
        'selected_ticket_number': ticket_number, 
        'date_from': date_from, 
        'date_to': date_to,
        'search_query': search, 
        'category': category,
        'is_viewed': is_viewed,
        'closed_count_30_days': closed_count_30_days,
        'selected_main_error_type': main_error_type,
        'selected_sub_error_type': sub_error_type,
        'selected_erp_id': erp_id,
        'selected_screen_number': screen_number,
        'all_screens': ScreenMaster.objects.all().order_by('screen_name'),
    }
    return render(request, 'admin_panel/all_tickets.html', context)


# ============================================================
# TICKET DETAIL - ADMIN (UPDATED WITH PRIORITY CHANGE & TARGET DATE)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def ticket_detail_admin(request, pk):
    """
    Admin ticket detail view with full ticket management
    Includes: Assign, Hold, Escalate, Close, Reopen, Change Priority, and Target Date
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    
    if not ticket.is_viewed:
        ticket.is_viewed = True
        ticket.viewed_at = timezone.now()
        ticket.save()
        logger.info(f"Ticket {ticket.ticket_number} marked as viewed by {request.user.username}")
    
    history = ticket.history.all().order_by('timestamp')
    can_reopen = False
    reopen_time_left = None
    reopen_deadline = None
    
    if ticket.status == 'Closed' and ticket.closed_at:
        reopen_deadline = ticket.closed_at + timedelta(hours=48)
        if timezone.now() < reopen_deadline: 
            can_reopen = True
            reopen_time_left = reopen_deadline - timezone.now()
    
    time_to_close_str = ""
    if ticket.status == 'Closed' and ticket.created_at and ticket.closed_at:
        time_to_close = ticket.closed_at - ticket.created_at
        time_to_close_str = format_timedelta_display(time_to_close)
    
    employees = EmployeeMaster.objects.filter(is_active=True, can_assign_ticket=True).order_by('employee_name')
    
    attachments = []
    if ticket.attachment_1:
        attachments.append({'file': ticket.attachment_1, 'name': 'Attachment 1'})
    if ticket.attachment_2:
        attachments.append({'file': ticket.attachment_2, 'name': 'Attachment 2'})
    if ticket.attachment_3:
        attachments.append({'file': ticket.attachment_3, 'name': 'Attachment 3'})
    
    # Initialize CloseTicketForm for GET requests
    close_form = CloseTicketForm()
    
    # Get ERP ID for this ticket's employee
    erp_id = 'Not Mapped'
    if ticket.employee_id:
        erp_mapping = ERPHolderMapping.objects.filter(
            employee__employee_id=ticket.employee_id
        ).first()
        if erp_mapping:
            erp_id = erp_mapping.erp_user_id

    # ✅ FIXED: Only filter by screen_code, NOT by pk
    screen_object = ScreenMaster.objects.filter(screen_code=ticket.screen_number).first()
    
    if request.method == 'POST':
        action_type = request.POST.get('action_type')
        remarks = request.POST.get('remarks', '')
        
        with transaction.atomic():
            if action_type == 'Assign':
                assigned_person = request.POST.get('assigned_person', '').strip()
                if not assigned_person: 
                    messages.error(request, "Assigned Person Name is mandatory.")
                    return redirect('admin_ticket_detail', pk=ticket.id)
                
                # ✅ Get target date
                target_date_str = request.POST.get('target_date', '').strip()
                target_date = None
                if target_date_str:
                    try:
                        target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
                    except ValueError:
                        messages.error(request, "Invalid target date format.")
                        return redirect('admin_ticket_detail', pk=ticket.id)
                
                ticket.status = 'Assigned'
                ticket.assigned_person = assigned_person
                if target_date:
                    ticket.target_date = target_date
                ticket.save()
                
                TicketHistory.objects.create(
                    ticket=ticket, 
                    action=f"Assigned to {assigned_person}", 
                    remarks=remarks + (f" | Target Date: {target_date.strftime('%d-%b-%Y')}" if target_date else ""), 
                    performed_by=f"Admin {request.user.username}"
                )
                # ============================================================
                # EMAIL SENDING DISABLED - COMMENTED OUT
                # ============================================================
                # send_ticket_email(ticket, 'Assigned')
                # ============================================================
                messages.success(request, f'Ticket assigned to {assigned_person}. Target date set to {target_date.strftime("%d-%b-%Y") if target_date else "Not set"}.')
                return redirect('admin_ticket_detail', pk=ticket.id)
                
            elif action_type == 'Hold':
                hold_reason = request.POST.get('hold_reason', '').strip()
                if not hold_reason: 
                    messages.error(request, "Hold Reason is mandatory.")
                    return redirect('admin_ticket_detail', pk=ticket.id)
                ticket.status = 'Hold'
                ticket.hold_reason = hold_reason
                ticket.save()
                TicketHistory.objects.create(
                    ticket=ticket, 
                    action="Status changed to Hold", 
                    remarks=f"Reason: {hold_reason}", 
                    performed_by=f"Admin {request.user.username}"
                )
                # ============================================================
                # EMAIL SENDING DISABLED - COMMENTED OUT
                # ============================================================
                # send_ticket_email(ticket, 'Hold')
                # ============================================================
                messages.success(request, 'Ticket placed on Hold.')
                return redirect('admin_ticket_detail', pk=ticket.id)
                
            elif action_type == 'Escalate':
                vendor_ticket = request.POST.get('vendor_ticket_number', '').strip()
                ticket.status = 'Escalated'
                if vendor_ticket: 
                    ticket.vendor_ticket_number = vendor_ticket
                ticket.escalated_at = timezone.now()
                ticket.save()
                remark_str = f"Vendor Ticket: {vendor_ticket}" if vendor_ticket else "Escalated without vendor ticket number"
                TicketHistory.objects.create(
                    ticket=ticket, 
                    action="Escalated to ERP Vendor", 
                    remarks=remark_str, 
                    performed_by=f"Admin {request.user.username}"
                )
                # ============================================================
                # EMAIL SENDING DISABLED - COMMENTED OUT
                # ============================================================
                # send_ticket_email(ticket, 'Escalated')
                # ============================================================
                messages.success(request, 'Ticket escalated to ERP vendor.')
                return redirect('admin_ticket_detail', pk=ticket.id)
                
            elif action_type == 'Close':
                close_form = CloseTicketForm(request.POST)
                
                if close_form.is_valid():
                    main_error_type = close_form.cleaned_data['main_error_type']
                    sub_error_type = close_form.cleaned_data['sub_error_type']
                    closing_remarks = close_form.cleaned_data['closing_remarks']
                    
                    ticket.status = 'Closed'
                    ticket.closing_remarks = closing_remarks
                    ticket.closed_by = request.user.username
                    ticket.closed_at = timezone.now()
                    
                    ticket.main_error_type = main_error_type
                    ticket.sub_error_type = sub_error_type
                    
                    ticket.save()
                    
                    TicketHistory.objects.create(
                        ticket=ticket, 
                        action=f"Closed by {request.user.username}", 
                        remarks=f"Main Error: {main_error_type} | Sub Error: {sub_error_type} | {closing_remarks}", 
                        performed_by=f"Admin {request.user.username}"
                    )
                    # ============================================================
                    # EMAIL SENDING DISABLED - COMMENTED OUT
                    # ============================================================
                    # send_ticket_email(ticket, 'Closed')
                    # ============================================================
                    messages.success(request, 'Ticket closed successfully.')
                    return redirect('admin_ticket_detail', pk=ticket.id)
                else:
                    messages.error(request, 'Please fix the errors below.')
                    context = {
                        'ticket': ticket,
                        'history': history,
                        'employees': employees,
                        'attachments': attachments,
                        'can_reopen': can_reopen,
                        'time_to_close': time_to_close_str,
                        'reopen_time_left': reopen_time_left,
                        'reopen_deadline_iso': reopen_deadline.isoformat() if reopen_deadline else None,
                        'close_form': close_form,
                        'erp_id': erp_id,
                        'screen_object': screen_object,
                    }
                    return render(request, 'admin_panel/ticket_detail.html', context)
                
            elif action_type == 'Reopen':
                if not can_reopen: 
                    messages.error(request, "Cannot reopen - 48 hours elapsed.")
                    return redirect('admin_ticket_detail', pk=ticket.id)
                remarks = request.POST.get('remarks', '').strip()
                if not remarks: 
                    messages.error(request, "Reason for reopening is mandatory.")
                    return redirect('admin_ticket_detail', pk=ticket.id)
                uploaded_files = request.FILES.getlist('reopen_attachments')
                try:
                    for uploaded_file in uploaded_files:
                        validate_attachment(uploaded_file)
                except ValidationError as error:
                    messages.error(request, str(error))
                    return redirect('admin_ticket_detail', pk=ticket.id)
                reopen_ticket_logic(ticket, f"Admin {request.user.username}", remarks, uploaded_files)
                messages.success(request, 'Ticket reopened successfully.')
                return redirect('admin_ticket_detail', pk=ticket.id)
                
            # ============================================================
            # PRIORITY CHANGE ACTION
            # ============================================================
            elif action_type == 'ChangePriority':
                # Check if ticket is closed
                if ticket.status == 'Closed':
                    messages.error(request, "Cannot change priority of a closed ticket.")
                    return redirect('admin_ticket_detail', pk=ticket.id)
                
                new_priority = request.POST.get('new_priority', '').strip()
                priority_reason = request.POST.get('priority_reason', '').strip()
                
                # Validate priority
                valid_priorities = ['Critical', 'High', 'Medium', 'Low']
                if new_priority not in valid_priorities:
                    messages.error(request, "Invalid priority selected.")
                    return redirect('admin_ticket_detail', pk=ticket.id)
                
                # Validate reason
                if not priority_reason:
                    messages.error(request, "Please provide a reason for changing priority.")
                    return redirect('admin_ticket_detail', pk=ticket.id)
                
                # Get old priority
                old_priority = ticket.priority
                
                # Update priority
                ticket.priority = new_priority
                ticket.save()
                
                # Create history entry
                history_remark = f"Priority changed from {old_priority} to {new_priority}. Reason: {priority_reason}"
                TicketHistory.objects.create(
                    ticket=ticket,
                    action="Priority Changed",
                    remarks=history_remark,
                    performed_by=f"Admin {request.user.username}"
                )
                
                messages.success(request, f'Priority changed from {old_priority} to {new_priority}.')
                return redirect('admin_ticket_detail', pk=ticket.id)
            
            # ============================================================
            # ✅ NEW: UPDATE TARGET DATE
            # ============================================================
            elif action_type == 'UpdateTargetDate':
                target_date_str = request.POST.get('target_date', '').strip()
                
                if not target_date_str:
                    messages.error(request, "Target date is required.")
                    return redirect('admin_ticket_detail', pk=ticket.id)
                
                try:
                    target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
                    
                    # Check if date is in the past
                    if target_date < timezone.now().date():
                        messages.error(request, "Target date cannot be in the past.")
                        return redirect('admin_ticket_detail', pk=ticket.id)
                    
                    old_target_date = ticket.target_date
                    ticket.target_date = target_date
                    ticket.save()
                    
                    TicketHistory.objects.create(
                        ticket=ticket,
                        action="Target Date Updated",
                        remarks=f"Target date changed from {old_target_date.strftime('%d-%b-%Y') if old_target_date else 'Not set'} to {target_date.strftime('%d-%b-%Y')}",
                        performed_by=f"Admin {request.user.username}"
                    )
                    
                    messages.success(request, f'Target date updated to {target_date.strftime("%d-%b-%Y")}.')
                    return redirect('admin_ticket_detail', pk=ticket.id)
                    
                except ValueError:
                    messages.error(request, "Invalid date format.")
                    return redirect('admin_ticket_detail', pk=ticket.id)
                
        return redirect('admin_ticket_detail', pk=ticket.id)
    
    context = {
        'ticket': ticket,
        'history': history,
        'employees': employees,
        'attachments': attachments,
        'can_reopen': can_reopen,
        'time_to_close': time_to_close_str,
        'reopen_time_left': reopen_time_left,
        'reopen_deadline_iso': reopen_deadline.isoformat() if reopen_deadline else None,
        'close_form': close_form,
        'erp_id': erp_id,
        'screen_object': screen_object,
    }
    return render(request, 'admin_panel/ticket_detail.html', context)


# ============================================================
# ✅ NEW: UPDATE TARGET DATE (Standalone Endpoint)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def update_target_date(request, ticket_id):
    """
    Update the target date for an assigned ticket via AJAX
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=400)
    
    try:
        ticket = Ticket.objects.get(id=ticket_id)
    except Ticket.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Ticket not found'}, status=404)
    
    target_date_str = request.POST.get('target_date', '').strip()
    
    if not target_date_str:
        return JsonResponse({'success': False, 'message': 'Target date is required'})
    
    try:
        from datetime import date
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        
        # Check if date is in the past
        if target_date < date.today():
            return JsonResponse({'success': False, 'message': 'Target date cannot be in the past'})
        
        # Check if ticket is assigned
        if ticket.status not in ['Assigned', 'Open']:
            return JsonResponse({'success': False, 'message': 'Ticket must be in Assigned or Open status to set target date'})
        
        old_target_date = ticket.target_date
        ticket.target_date = target_date
        ticket.save()
        
        # Log the change
        TicketHistory.objects.create(
            ticket=ticket,
            action="Target Date Updated",
            remarks=f"Target date changed from {old_target_date.strftime('%d-%b-%Y') if old_target_date else 'Not set'} to {target_date.strftime('%d-%b-%Y')}",
            performed_by=f"Admin {request.user.username}"
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Target date updated to {target_date.strftime("%d-%b-%Y")}',
            'target_date': target_date.strftime('%d-%b-%Y')
        })
        
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Invalid date format'}, status=400)


# ============================================================
# NOTIFICATION FUNCTIONS - FIXED FOR AJAX
# ============================================================

@login_required
@user_passes_test(is_admin, login_url='login')
def get_notifications(request):
    """
    Get unviewed tickets for AJAX dropdown refresh
    ✅ FIXED: Always returns JSON for AJAX requests
    """
    # Check if it's an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if is_ajax:
        try:
            unviewed_count = Ticket.objects.filter(is_viewed=False).count()
            unviewed_tickets = Ticket.objects.filter(is_viewed=False).order_by('-created_at')[:10]
            
            html = render_to_string('admin_panel/_notification_items.html', {
                'unviewed_tickets': unviewed_tickets,
                'unviewed_count': unviewed_count,
            }, request=request)
            
            return JsonResponse({
                'success': True,
                'html': html,
                'count': unviewed_count
            })
        except Exception as e:
            logger.error(f"Error in get_notifications: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': str(e),
                'count': 0,
                'html': ''
            }, status=500)
    
    # For non-AJAX requests, return a proper page or redirect
    return JsonResponse({
        'success': False,
        'message': 'Invalid request. This endpoint only accepts AJAX requests.',
        'count': 0,
        'html': ''
    }, status=400)


@login_required
@user_passes_test(is_admin, login_url='login')
def mark_all_notifications_read(request):
    """
    Mark all tickets as viewed
    ✅ FIXED: Always returns JSON for AJAX requests
    """
    if request.method == 'POST':
        try:
            count = Ticket.objects.filter(is_viewed=False).update(
                is_viewed=True,
                viewed_at=timezone.now()
            )
            return JsonResponse({
                'success': True,
                'message': f'Marked {count} tickets as read',
                'count': 0
            })
        except Exception as e:
            logger.error(f"Error in mark_all_notifications_read: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
    
    return JsonResponse({'success': False, 'message': 'Invalid request. Use POST.'}, status=400)


@login_required
@user_passes_test(is_admin, login_url='login')
def mark_notification_read(request, ticket_id):
    """
    Mark a single ticket as viewed
    ✅ FIXED: Always returns JSON for AJAX requests
    """
    if request.method == 'POST':
        try:
            ticket = get_object_or_404(Ticket, pk=ticket_id)
            ticket.is_viewed = True
            ticket.viewed_at = timezone.now()
            ticket.save()
            return JsonResponse({
                'success': True,
                'message': f'Ticket {ticket.ticket_number} marked as read'
            })
        except Exception as e:
            logger.error(f"Error in mark_notification_read: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
    
    return JsonResponse({'success': False, 'message': 'Invalid request. Use POST.'}, status=400)


# ============================================================
# TEST NOTIFICATIONS
# ============================================================

@login_required
@user_passes_test(is_admin, login_url='login')
def test_notifications(request): 
    return render(request, 'admin_panel/test_notifications.html')


@login_required
@user_passes_test(is_admin, login_url='login')
def test_success_message(request): 
    messages.success(request, 'Test success message.')
    return redirect('test_notifications')


@login_required
@user_passes_test(is_admin, login_url='login')
def test_error_message(request): 
    messages.error(request, 'Test error message.')
    return redirect('test_notifications')


@login_required
@user_passes_test(is_admin, login_url='login')
def test_warning_message(request): 
    messages.warning(request, 'Test warning message.')
    return redirect('test_notifications')


@login_required
@user_passes_test(is_admin, login_url='login')
def test_info_message(request): 
    messages.info(request, 'Test info message.')
    return redirect('test_notifications')


# ============================================================
# DOWNLOAD AUDIT LOG EXCEL
# ============================================================

@login_required
@user_passes_test(is_admin, login_url='login')
def download_audit_log_excel(request):
    """
    Export filtered audit logs to Excel
    """
    audit_logs = SettingsAuditLog.objects.all().order_by('-created_at')
    
    action = request.GET.get('action', '')
    setting_type = request.GET.get('setting_type', '')
    performed_by = request.GET.get('performed_by', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search = request.GET.get('search', '')
    
    if action:
        audit_logs = audit_logs.filter(action_type=action)
    if setting_type:
        audit_logs = audit_logs.filter(setting_type=setting_type)
    if performed_by:
        audit_logs = audit_logs.filter(performed_by_name__icontains=performed_by)
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            from_datetime = timezone.make_aware(
                datetime.combine(date_from_obj, datetime.min.time())
            )
            audit_logs = audit_logs.filter(created_at__gte=from_datetime)
        except ValueError:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            to_datetime = timezone.make_aware(
                datetime.combine(date_to_obj, datetime.max.time())
            )
            audit_logs = audit_logs.filter(created_at__lte=to_datetime)
        except ValueError:
            pass
    if search:
        audit_logs = audit_logs.filter(
            Q(setting_name__icontains=search) |
            Q(change_summary__icontains=search) |
            Q(performed_by_name__icontains=search)
        )
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Audit_Log_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Log"
    
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10)
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    current_tz = timezone.get_current_timezone()
    now_utc = timezone.now()
    if timezone.is_naive(now_utc):
        now_utc = timezone.make_aware(now_utc, timezone.utc)
    now_local = now_utc.astimezone(current_tz)
    report_time = now_local.strftime('%d-%b-%Y %I:%M:%S %p')
    
    ws.merge_cells('A1:K1')
    ws['A1'] = f"GPLAST SETTINGS AUDIT LOG - Generated: {report_time}  |  Total Entries: {audit_logs.count()}"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45
    
    headers = ['ID', 'Action', 'Setting Type', 'Setting Name', 'Old Value', 'New Value', 
               'Change Summary', 'Performed By', 'IP Address', 'Remarks', 'Created At']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[3].height = 30
    
    row_idx = 4
    for log in audit_logs:
        ws.cell(row=row_idx, column=1, value=log.id).font = data_font
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_idx, column=1).border = thin_border
        
        ws.cell(row=row_idx, column=2, value=log.get_action_type_display()).font = data_font
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row_idx, column=2).border = thin_border
        
        ws.cell(row=row_idx, column=3, value=log.get_setting_type_display()).font = data_font
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row_idx, column=3).border = thin_border
        
        ws.cell(row=row_idx, column=4, value=log.setting_name).font = data_font
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row_idx, column=4).border = thin_border
        
        ws.cell(row=row_idx, column=5, value=log.old_value or '').font = data_font
        ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=row_idx, column=5).border = thin_border
        
        ws.cell(row=row_idx, column=6, value=log.new_value or '').font = data_font
        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=row_idx, column=6).border = thin_border
        
        ws.cell(row=row_idx, column=7, value=log.change_summary or '').font = data_font
        ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=row_idx, column=7).border = thin_border
        
        ws.cell(row=row_idx, column=8, value=log.performed_by_name).font = data_font
        ws.cell(row=row_idx, column=8).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row_idx, column=8).border = thin_border
        
        ws.cell(row=row_idx, column=9, value=log.ip_address or 'N/A').font = data_font
        ws.cell(row=row_idx, column=9).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row_idx, column=9).border = thin_border
        
        ws.cell(row=row_idx, column=10, value=log.remarks or '').font = data_font
        ws.cell(row=row_idx, column=10).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=row_idx, column=10).border = thin_border
        
        if log.created_at:
            if timezone.is_naive(log.created_at):
                utc_time = timezone.make_aware(log.created_at, timezone.utc)
            else:
                utc_time = log.created_at
            local_time = utc_time.astimezone(current_tz)
            formatted_time = local_time.strftime('%d-%b-%Y %I:%M:%S %p')
        else:
            formatted_time = ''
        
        ws.cell(row=row_idx, column=11, value=formatted_time).font = data_font
        ws.cell(row=row_idx, column=11).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_idx, column=11).border = thin_border
        
        row_idx += 1
    
    column_widths = {
        'A': 10, 'B': 20, 'C': 20, 'D': 30, 'E': 35,
        'F': 35, 'G': 40, 'H': 22, 'I': 18, 'J': 35, 'K': 25
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    wb.save(response)
    return response