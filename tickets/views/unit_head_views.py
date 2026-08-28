# tickets/views/unit_head_views.py

"""
Unit Head Views - Dashboard, All Tickets, My Tickets, Ticket Detail, Reports
Unit Heads can only see tickets from their assigned unit.
VIEW ONLY - No actions (Assign, Hold, Escalate, Close, Reopen)
BUT CAN CHANGE PRIORITY for non-closed tickets
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Count, Q, Value, CharField, OuterRef, Subquery
from django.db.models.functions import Coalesce
from django.db import transaction
from django.core.exceptions import ValidationError
from django.template.loader import render_to_string
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging

from tickets.models import (
    Unit, Department, Ticket, TicketHistory, EmployeeMaster,
    ScreenMaster, ERPHolderMapping, ReopenAttachment, UnitHead
)
from tickets.forms import CloseTicketForm
from tickets.utils import send_ticket_email, validate_attachment
from .utils import format_timedelta_display, reopen_ticket_logic

logger = logging.getLogger(__name__)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def get_unit_head_unit(user):
    """Get the unit for a unit head user"""
    try:
        unit_head = UnitHead.objects.filter(user=user, is_active=True).select_related('unit').first()
        return unit_head.unit if unit_head else None
    except:
        return None


def get_unit_head_object(user):
    """Get the unit head object for a user"""
    try:
        return UnitHead.objects.filter(user=user, is_active=True).first()
    except:
        return None


def is_unit_head(user):
    """Check if user is a unit head"""
    try:
        return UnitHead.objects.filter(user=user, is_active=True).exists()
    except:
        return False


def unit_head_required(view_func):
    """Decorator to check if user is a unit head"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not is_unit_head(request.user):
            messages.error(request, "You do not have permission to access this page.")
            return redirect('employee_dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


# ============================================================
# UNIT HEAD DASHBOARD
# ============================================================
@login_required
@unit_head_required
def unit_head_dashboard(request):
    """
    Unit Head dashboard showing:
    - KPIs for tickets from their unit only
    - Charts: Status, Priority
    - Recent tickets from their unit
    """
    unit = get_unit_head_unit(request.user)
    unit_head = get_unit_head_object(request.user)
    
    if not unit:
        messages.error(request, "Your unit head profile is not properly configured.")
        return redirect('employee_dashboard')
    
    # Get tickets for this unit only
    unit_tickets = Ticket.objects.filter(unit=unit)
    
    # KPIs
    kpis = {
        'total': unit_tickets.count(),
        'open': unit_tickets.filter(status='Open').count(),
        'assigned': unit_tickets.filter(status='Assigned').count(),
        'hold': unit_tickets.filter(status='Hold').count(),
        'escalated': unit_tickets.filter(status='Escalated').count(),
        'closed': unit_tickets.filter(status='Closed').count(),
        'critical': unit_tickets.filter(priority='Critical').count(),
    }
    
    # Status chart
    status_counts = list(unit_tickets.values('status').annotate(count=Count('id')))
    chart_status = {item['status']: item['count'] for item in status_counts}
    
    # Priority chart
    prio_counts = list(unit_tickets.values('priority').annotate(count=Count('id')))
    chart_priority = {item['priority']: item['count'] for item in prio_counts}
    
    # Monthly chart (last 12 months)
    twelve_months_ago = timezone.now() - timedelta(days=365)
    monthly_counts = {}
    for i in range(12):
        month = timezone.now() - timedelta(days=30 * i)
        month_label = month.strftime('%b %Y')
        monthly_counts[month_label] = 0
    
    for ticket in unit_tickets.filter(created_at__gte=twelve_months_ago):
        month_label = ticket.created_at.strftime('%b %Y')
        if month_label in monthly_counts:
            monthly_counts[month_label] += 1
    
    chart_monthly = [{'label': k, 'value': v} for k, v in monthly_counts.items()]
    chart_monthly.reverse()
    
    # Build charts_data with proper JSON serializable data
    charts_data = {
        'status': chart_status,
        'priority': chart_priority,
        'monthly': chart_monthly,
    }
    
    # Recent tickets
    recent_tickets = unit_tickets.order_by('-created_at')[:10]
    
    context = {
        'kpis': kpis,
        'charts_data': charts_data,
        'recent_tickets': recent_tickets,
        'unit_head': unit_head,
        'unit': unit,
        'unit_head_unit_code': unit.code,
        'unit_head_unit_name': unit.full_name,
        'unit_head_name': unit_head.name,
    }
    return render(request, 'unit_head/dashboard.html', context)


# ============================================================
# UNIT HEAD - ALL TICKETS (VIEW ONLY) - WITH AGING SUPPORT
# ============================================================
@login_required
@unit_head_required
def unit_head_all_tickets(request):
    """
    Unit Head ticket listing - only tickets from their unit - VIEW ONLY
    Supports AJAX requests for drill-down modal with aging
    """
    unit = get_unit_head_unit(request.user)
    unit_head = get_unit_head_object(request.user)
    
    if not unit:
        messages.error(request, "Your unit head profile is not properly configured.")
        return redirect('employee_dashboard')
    
    # Annotate tickets with ERP ID from ERPHolderMapping
    erp_subquery = ERPHolderMapping.objects.filter(
        employee__employee_id=OuterRef('employee_id')
    ).values('erp_user_id')[:1]
    
    tickets_qs = Ticket.objects.filter(unit=unit).order_by('-created_at').annotate(
        erp_id=Coalesce(Subquery(erp_subquery, output_field=CharField()), Value('Not Mapped'))
    )
    
    # Get filter parameters
    status = request.GET.get('status', '')
    priority = request.GET.get('priority', '')
    ticket_number = request.GET.get('ticket_number', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    search = request.GET.get('search', '').strip()
    main_error_type = request.GET.get('main_error_type', '').strip()
    sub_error_type = request.GET.get('sub_error_type', '').strip()
    
    # Apply filters
    if status:
        tickets_qs = tickets_qs.filter(status=status)
    if priority:
        tickets_qs = tickets_qs.filter(priority=priority)
    if ticket_number:
        tickets_qs = tickets_qs.filter(ticket_number__icontains=ticket_number)
    if main_error_type and main_error_type != '':
        tickets_qs = tickets_qs.filter(main_error_type=main_error_type)
    if sub_error_type and sub_error_type != '' and sub_error_type != 'All':
        tickets_qs = tickets_qs.filter(sub_error_type=sub_error_type)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            from_datetime = timezone.make_aware(
                datetime.combine(date_from_obj, datetime.min.time())
            )
            tickets_qs = tickets_qs.filter(created_at__gte=from_datetime)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            to_datetime = timezone.make_aware(
                datetime.combine(date_to_obj, datetime.max.time())
            )
            tickets_qs = tickets_qs.filter(created_at__lte=to_datetime)
        except ValueError:
            pass
    
    if search:
        tickets_qs = tickets_qs.filter(
            Q(ticket_number__icontains=search) |
            Q(subject__icontains=search) |
            Q(employee_name__icontains=search) |
            Q(employee_id__icontains=search)
        )
    
    # ============================================================
    # CHECK IF AJAX REQUEST FOR DRILL-DOWN MODAL
    # Using request.headers for proper AJAX detection
    # ============================================================
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if is_ajax:
        tickets = tickets_qs[:50]  # Limit for modal display
        html = render_to_string('unit_head/_ticket_list_modal.html', {
            'tickets': tickets,
        }, request=request)
        return JsonResponse({
            'html': html,
            'success': True,
            'count': tickets_qs.count()
        })
    
    # ============================================================
    # REGULAR PAGE RENDER
    # ============================================================
    paginator = Paginator(tickets_qs, 20)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # Build filter query for pagination
    filter_query = request.GET.copy()
    filter_query.pop('page', None)
    
    context = {
        'page_obj': page_obj,
        'unit_head': unit_head,
        'unit': unit,
        'unit_head_unit_code': unit.code,
        'unit_head_unit_name': unit.full_name,
        'status_choices': Ticket.STATUS_CHOICES,
        'priority_choices': Ticket.PRIORITY_CHOICES,
        'selected_status': status,
        'selected_priority': priority,
        'selected_ticket_number': ticket_number,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search,
        'selected_main_error_type': main_error_type,
        'selected_sub_error_type': sub_error_type,
        'filter_query': filter_query.urlencode(),
    }
    return render(request, 'unit_head/all_tickets.html', context)


# ============================================================
# UNIT HEAD - MY TICKETS (VIEW ONLY)
# ============================================================
@login_required
@unit_head_required
def unit_head_my_tickets(request):
    """
    Tickets created by the unit head - filtered by their unit - VIEW ONLY
    """
    unit = get_unit_head_unit(request.user)
    unit_head = get_unit_head_object(request.user)
    
    if not unit:
        messages.error(request, "Your unit head profile is not properly configured.")
        return redirect('employee_dashboard')
    
    tickets_qs = Ticket.objects.filter(
        unit=unit,
        created_by_user=request.user
    ).order_by('-created_at')
    
    # Get filter parameters
    status = request.GET.get('status', '')
    priority = request.GET.get('priority', '')
    ticket_number = request.GET.get('ticket_number', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    search = request.GET.get('search', '').strip()
    
    # Apply filters
    if status:
        tickets_qs = tickets_qs.filter(status=status)
    if priority:
        tickets_qs = tickets_qs.filter(priority=priority)
    if ticket_number:
        tickets_qs = tickets_qs.filter(ticket_number__icontains=ticket_number)
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            from_datetime = timezone.make_aware(
                datetime.combine(date_from_obj, datetime.min.time())
            )
            tickets_qs = tickets_qs.filter(created_at__gte=from_datetime)
        except ValueError:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            to_datetime = timezone.make_aware(
                datetime.combine(date_to_obj, datetime.max.time())
            )
            tickets_qs = tickets_qs.filter(created_at__lte=to_datetime)
        except ValueError:
            pass
    if search:
        tickets_qs = tickets_qs.filter(
            Q(ticket_number__icontains=search) |
            Q(subject__icontains=search) |
            Q(employee_name__icontains=search)
        )
    
    # KPIs
    total = tickets_qs.count()
    open_tickets = tickets_qs.filter(status='Open').count()
    assigned_tickets = tickets_qs.filter(status='Assigned').count()
    hold_tickets = tickets_qs.filter(status='Hold').count()
    escalated_tickets = tickets_qs.filter(status='Escalated').count()
    closed_tickets = tickets_qs.filter(status='Closed').count()
    
    # Pagination
    paginator = Paginator(tickets_qs, 20)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # Build filter query for pagination
    filter_query = request.GET.copy()
    filter_query.pop('page', None)
    
    context = {
        'page_obj': page_obj,
        'unit_head': unit_head,
        'unit': unit,
        'unit_head_unit_code': unit.code,
        'unit_head_unit_name': unit.full_name,
        'total': total,
        'open_tickets': open_tickets,
        'assigned_tickets': assigned_tickets,
        'hold_tickets': hold_tickets,
        'escalated_tickets': escalated_tickets,
        'closed_tickets': closed_tickets,
        'status_choices': Ticket.STATUS_CHOICES,
        'priority_choices': Ticket.PRIORITY_CHOICES,
        'selected_status': status,
        'selected_priority': priority,
        'selected_ticket_number': ticket_number,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search,
        'filter_query': filter_query.urlencode(),
    }
    return render(request, 'unit_head/my_tickets.html', context)


# ============================================================
# UNIT HEAD - TICKET DETAIL (UPDATED WITH PRIORITY CHANGE)
# ============================================================
@login_required
@unit_head_required
def unit_head_ticket_detail(request, ticket_id):
    """
    Unit Head ticket detail view - VIEW ONLY except Priority Change
    - Only if ticket belongs to their unit
    - No action buttons (Assign, Hold, Escalate, Close, Reopen)
    - CAN CHANGE PRIORITY for non-closed tickets
    - Shows ticket info, employee details, attachments, history
    """
    unit = get_unit_head_unit(request.user)
    unit_head = get_unit_head_object(request.user)
    
    if not unit:
        messages.error(request, "Your unit head profile is not properly configured.")
        return redirect('employee_dashboard')
    
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Security: Ensure ticket belongs to unit head's unit
    if ticket.unit != unit:
        messages.error(request, "You do not have permission to view this ticket.")
        return redirect('unit_head_all_tickets')
    
    history = ticket.history.all().order_by('timestamp')
    
    time_to_close_str = ""
    if ticket.status == 'Closed' and ticket.created_at and ticket.closed_at:
        time_to_close = ticket.closed_at - ticket.created_at
        time_to_close_str = format_timedelta_display(time_to_close)
    
    # Attachments
    attachments = []
    if ticket.attachment_1:
        attachments.append({'file': ticket.attachment_1, 'name': 'Attachment 1'})
    if ticket.attachment_2:
        attachments.append({'file': ticket.attachment_2, 'name': 'Attachment 2'})
    if ticket.attachment_3:
        attachments.append({'file': ticket.attachment_3, 'name': 'Attachment 3'})
    
    # Get ERP ID for this ticket's employee
    erp_id = 'Not Mapped'
    if ticket.employee_id:
        erp_mapping = ERPHolderMapping.objects.filter(
            employee__employee_id=ticket.employee_id
        ).first()
        if erp_mapping:
            erp_id = erp_mapping.erp_user_id
    
    # Get screen object (view only)
    screen_object = ScreenMaster.objects.filter(screen_code=ticket.screen_number).first()
    
    # ============================================================
    # HANDLE PRIORITY CHANGE POST REQUEST
    # ============================================================
    if request.method == 'POST':
        action_type = request.POST.get('action_type')
        
        if action_type == 'ChangePriority':
            # Check if ticket is closed
            if ticket.status == 'Closed':
                messages.error(request, "Cannot change priority of a closed ticket.")
                return redirect('unit_head_ticket_detail', ticket_id=ticket.id)
            
            new_priority = request.POST.get('new_priority', '').strip()
            priority_reason = request.POST.get('priority_reason', '').strip()
            
            # Validate priority
            valid_priorities = ['Critical', 'High', 'Medium', 'Low']
            if new_priority not in valid_priorities:
                messages.error(request, "Invalid priority selected.")
                return redirect('unit_head_ticket_detail', ticket_id=ticket.id)
            
            # Validate reason
            if not priority_reason:
                messages.error(request, "Please provide a reason for changing priority.")
                return redirect('unit_head_ticket_detail', ticket_id=ticket.id)
            
            with transaction.atomic():
                # Get old priority
                old_priority = ticket.priority
                
                # Update priority
                ticket.priority = new_priority
                ticket.save()
                
                # Create history entry
                history_remark = f"Priority changed from {old_priority} to {new_priority}. Reason: {priority_reason}"
                TicketHistory.objects.create(
                    ticket=ticket,
                    action="Priority Changed",
                    remarks=history_remark,
                    performed_by=f"Unit Head {unit_head.name}"
                )
                
                messages.success(request, f'Priority changed from {old_priority} to {new_priority}.')
                return redirect('unit_head_ticket_detail', ticket_id=ticket.id)
        
        return redirect('unit_head_ticket_detail', ticket_id=ticket.id)
    
    context = {
        'ticket': ticket,
        'history': history,
        'attachments': attachments,
        'time_to_close': time_to_close_str,
        'erp_id': erp_id,
        'screen_object': screen_object,
        'unit_head': unit_head,
        'unit': unit,
        'unit_head_unit_code': unit.code,
        'unit_head_unit_name': unit.full_name,
    }
    return render(request, 'unit_head/ticket_detail.html', context)


# ============================================================
# UNIT HEAD - DOWNLOAD TICKET EXCEL
# ============================================================
@login_required
@unit_head_required
def unit_head_download_ticket_excel(request, ticket_id):
    """
    Download individual ticket as Excel - only if ticket belongs to unit head's unit
    """
    unit = get_unit_head_unit(request.user)
    
    if not unit:
        messages.error(request, "Your unit head profile is not properly configured.")
        return redirect('employee_dashboard')
    
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Security: Ensure ticket belongs to unit head's unit
    if ticket.unit != unit:
        messages.error(request, "You do not have permission to download this ticket.")
        return redirect('unit_head_all_tickets')
    
    erp_id = 'Not Mapped'
    if ticket.employee_id:
        erp_mapping = ERPHolderMapping.objects.filter(
            employee__employee_id=ticket.employee_id
        ).first()
        if erp_mapping:
            erp_id = erp_mapping.erp_user_id
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Ticket_{ticket.ticket_number}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ticket {ticket.ticket_number}"
    
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    section_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    label_font = Font(name='Calibri', size=11, bold=True, color='1A2A6C')
    data_font = Font(name='Calibri', size=11, color='333333')
    
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    section_fill = PatternFill(start_color='FF6B00', end_color='FF6B00', fill_type='solid')
    label_fill = PatternFill(start_color='E8EDF5', end_color='E8EDF5', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"GPLAST TICKET DETAILS - {ticket.ticket_number}"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    row = 3
    
    # Basic Information
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "BASIC INFORMATION"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 30
    row += 1
    
    basic_info = [
        ('Ticket Number', ticket.ticket_number),
        ('Subject', ticket.subject),
        ('Description', ticket.description or ''),
        ('Priority', ticket.priority),
        ('Status', ticket.status),
        ('Error Type', ticket.error_type or 'Not Set'),
        ('Created Date', timezone.localtime(ticket.created_at).strftime('%d-%b-%Y %I:%M %p') if ticket.created_at else ''),
        ('Updated Date', ticket.updated_at.strftime('%d-%b-%Y %I:%M %p') if ticket.updated_at else ''),
    ]
    
    for label, value in basic_info:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=1).fill = label_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.cell(row=row, column=2, value=value).font = data_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
    
    row += 1
    
    # Employee Details
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "EMPLOYEE DETAILS"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 30
    row += 1
    
    emp_info = [
        ('Employee Name', ticket.employee_name),
        ('Employee ID', ticket.employee_id),
        ('ERP ID', erp_id),
        ('Mobile', ticket.mobile or ''),
        ('Email', ticket.email or ''),
        ('Unit', ticket.unit.full_name if ticket.unit else ''),
        ('Department', ticket.department.name if ticket.department else ''),
        ('Screen/Module', ticket.screen_number),
    ]
    
    for label, value in emp_info:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=1).fill = label_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.cell(row=row, column=2, value=value).font = data_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
    
    row += 1
    
    # Assignment & Status
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "ASSIGNMENT & STATUS"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 30
    row += 1
    
    assign_info = [
        ('Created By Role', ticket.created_by_role),
        ('Assigned To', ticket.assigned_person or 'Not Assigned'),
        ('Hold Reason', ticket.hold_reason or ''),
        ('Vendor Ticket', ticket.vendor_ticket_number or ''),
        ('Admin Creation Reason', ticket.admin_creation_reason or ''),
    ]
    
    for label, value in assign_info:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=1).fill = label_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.cell(row=row, column=2, value=value).font = data_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
    
    row += 1
    
    # Closing Details (if closed)
    if ticket.status == 'Closed':
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "CLOSING DETAILS"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 30
        row += 1
        
        time_to_close_str = ''
        if ticket.created_at and ticket.closed_at:
            duration = ticket.closed_at - ticket.created_at
            days = duration.days
            hours = duration.seconds // 3600
            minutes = (duration.seconds % 3600) // 60
            if days > 0:
                time_to_close_str = f"{days}d {hours}h {minutes}m"
            else:
                time_to_close_str = f"{hours}h {minutes}m"
        
        closing_info = [
            ('Closed By', ticket.closed_by or ''),
            ('Closed Date', timezone.localtime(ticket.closed_at).strftime('%d-%b-%Y %I:%M %p') if ticket.closed_at else ''),
            ('Main Error Type', ticket.main_error_type or 'N/A'),
            ('Sub Error Type', ticket.sub_error_type or 'N/A'),
            ('Closing Remarks', ticket.closing_remarks or ''),
            ('Time to Close', time_to_close_str),
        ]
        
        for label, value in closing_info:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=1).fill = label_fill
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.cell(row=row, column=2, value=value).font = data_font
            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            row += 1
        
        row += 1
    
    # Footer
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"Report generated on {timezone.now().strftime('%d-%b-%Y %I:%M %p')} | GPLAST Support System"
    ws[f'A{row}'].font = Font(name='Calibri', size=9, italic=True, color='666666')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    wb.save(response)
    return response


# ============================================================
# UNIT HEAD - REPORTS (VIEW ONLY - NO CHARTS)
# ============================================================
@login_required
@unit_head_required
def unit_head_reports(request):
    """
    Unit Head reports view - VIEW ONLY
    - Unit-specific reports with KPIs and data table only
    - NO CHARTS (removed per requirement)
    """
    unit = get_unit_head_unit(request.user)
    unit_head = get_unit_head_object(request.user)
    
    if not unit:
        messages.error(request, "Your unit head profile is not properly configured.")
        return redirect('employee_dashboard')
    
    unit_tickets = Ticket.objects.filter(unit=unit)
    
    # Get filter parameters
    status = request.GET.get('status', '')
    priority = request.GET.get('priority', '')
    main_error_type = request.GET.get('main_error_type', '').strip()
    sub_error_type = request.GET.get('sub_error_type', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    search = request.GET.get('search', '').strip()
    
    tickets_qs = unit_tickets.all()
    
    if status:
        tickets_qs = tickets_qs.filter(status=status)
    if priority:
        tickets_qs = tickets_qs.filter(priority=priority)
    if main_error_type and main_error_type != '':
        tickets_qs = tickets_qs.filter(main_error_type=main_error_type)
    if sub_error_type and sub_error_type != '' and sub_error_type != 'All':
        tickets_qs = tickets_qs.filter(sub_error_type=sub_error_type)
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            from_datetime = timezone.make_aware(
                datetime.combine(date_from_obj, datetime.min.time())
            )
            tickets_qs = tickets_qs.filter(created_at__gte=from_datetime)
        except ValueError:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            to_datetime = timezone.make_aware(
                datetime.combine(date_to_obj, datetime.max.time())
            )
            tickets_qs = tickets_qs.filter(created_at__lte=to_datetime)
        except ValueError:
            pass
    if search:
        tickets_qs = tickets_qs.filter(
            Q(ticket_number__icontains=search) |
            Q(subject__icontains=search) |
            Q(employee_name__icontains=search)
        )
    
    # KPIs
    kpis = {
        'total': unit_tickets.count(),
        'open': unit_tickets.filter(status='Open').count(),
        'assigned': unit_tickets.filter(status='Assigned').count(),
        'hold': unit_tickets.filter(status='Hold').count(),
        'escalated': unit_tickets.filter(status='Escalated').count(),
        'closed': unit_tickets.filter(status='Closed').count(),
        'critical': unit_tickets.filter(priority='Critical').count(),
    }
    
    # Pagination
    paginator = Paginator(tickets_qs, 15)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # Export
    if request.GET.get('export') == 'excel':
        return unit_head_export_filtered_tickets_excel(request, tickets_qs, unit)
    
    # Build filter query for pagination
    filter_query = request.GET.copy()
    filter_query.pop('page', None)
    
    context = {
        'page_obj': page_obj,
        'kpis': kpis,
        'unit_head': unit_head,
        'unit': unit,
        'unit_head_unit_code': unit.code,
        'unit_head_unit_name': unit.full_name,
        'status_choices': Ticket.STATUS_CHOICES,
        'priority_choices': Ticket.PRIORITY_CHOICES,
        'selected_status': status,
        'selected_priority': priority,
        'selected_main_error_type': main_error_type,
        'selected_sub_error_type': sub_error_type,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search,
        'total_filtered': tickets_qs.count(),
        'filter_query': filter_query.urlencode(),
    }
    return render(request, 'unit_head/reports.html', context)


# ============================================================
# UNIT HEAD - EXPORT CLOSED TICKETS (30 DAYS)
# ============================================================
@login_required
@unit_head_required
def unit_head_export_closed_tickets_30_days(request):
    """
    Export closed tickets from last 30 days - unit only
    """
    unit = get_unit_head_unit(request.user)
    
    if not unit:
        messages.error(request, "Your unit head profile is not properly configured.")
        return redirect('employee_dashboard')
    
    thirty_days_ago = timezone.now() - timedelta(days=30)
    
    tickets_qs = Ticket.objects.filter(
        unit=unit,
        status='Closed',
        closed_at__gte=thirty_days_ago
    ).order_by('-closed_at')
    
    current_tz = timezone.get_current_timezone()
    now_utc = timezone.now()
    if timezone.is_naive(now_utc):
        now_utc = timezone.make_aware(now_utc, timezone.utc)
    now_local = now_utc.astimezone(current_tz)
    report_time = now_local.strftime('%d-%b-%Y %I:%M:%S %p')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Closed_Tickets_{unit.code}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Build mapping of employee_id to ERP ID
    employee_ids = tickets_qs.values_list('employee_id', flat=True).distinct()
    erp_mappings = {}
    if employee_ids:
        erp_mappings_qs = ERPHolderMapping.objects.filter(
            employee__employee_id__in=employee_ids
        ).select_related('employee')
        for mapping in erp_mappings_qs:
            erp_mappings[mapping.employee.employee_id] = mapping.erp_user_id
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Closed Tickets - {unit.code}"
    
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10)
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    ws.merge_cells('A1:AA1')
    ws['A1'] = f"CLOSED TICKETS - {unit.code} - LAST 30 DAYS"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45
    
    ws.merge_cells('A2:AA2')
    ws['A2'] = f"Generated: {report_time}  |  Total Closed Tickets: {tickets_qs.count()}  |  Unit: {unit.full_name}"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    
    headers = [
        'Ticket Number', 'Status', 'Unit Code', 'Unit Name', 'Department',
        'Employee ID', 'ERP ID', 'Employee Name', 'Mobile', 'Email', 'Screen/Module',
        'Subject', 'Description', 'Priority', 'Error Type', 'Created By Role',
        'Admin Creation Reason', 'Assigned Person', 'Hold Reason',
        'Main Error Type', 'Sub Error Type', 'Closing Remarks', 'Closed By',
        'Vendor Ticket Number', 'Created At', 'Closed At', 'Time to Close', 'Escalated At'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[4].height = 30
    
    row_idx = 5
    for ticket in tickets_qs:
        created_at_local = ticket.created_at.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p') if ticket.created_at else ''
        closed_at_local = ticket.closed_at.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p') if ticket.closed_at else ''
        escalated_at_local = ticket.escalated_at.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p') if ticket.escalated_at else ''
        
        time_to_close = ''
        if ticket.created_at and ticket.closed_at:
            duration = ticket.closed_at - ticket.created_at
            days = duration.days
            hours = duration.seconds // 3600
            minutes = (duration.seconds % 3600) // 60
            time_to_close = f"{days}d {hours}h {minutes}m" if days > 0 else f"{hours}h {minutes}m"
        
        erp_id = erp_mappings.get(ticket.employee_id, 'Not Mapped')
        
        row_data = [
            ticket.ticket_number, ticket.status, ticket.unit.code if ticket.unit else '',
            ticket.unit.full_name if ticket.unit else '', ticket.department.name if ticket.department else '',
            ticket.employee_id, erp_id, ticket.employee_name, ticket.mobile or '', ticket.email or '',
            ticket.screen_number, ticket.subject, ticket.description or '', ticket.priority,
            ticket.error_type or '', ticket.created_by_role, ticket.admin_creation_reason or '',
            ticket.assigned_person or '', ticket.hold_reason or '', ticket.main_error_type or 'N/A',
            ticket.sub_error_type or 'N/A', ticket.closing_remarks or '', ticket.closed_by or '',
            ticket.vendor_ticket_number or '', created_at_local, closed_at_local, time_to_close,
            escalated_at_local,
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        row_idx += 1
    
    column_widths = {
        'A': 18, 'B': 14, 'C': 12, 'D': 25, 'E': 20, 'F': 14, 'G': 16,
        'H': 22, 'I': 16, 'J': 25, 'K': 16, 'L': 30, 'M': 40, 'N': 14,
        'O': 20, 'P': 18, 'Q': 25, 'R': 20, 'S': 20, 'T': 22, 'U': 22,
        'V': 30, 'W': 18, 'X': 18, 'Y': 22, 'Z': 22, 'AA': 16, 'AB': 22
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    wb.save(response)
    return response


# ============================================================
# UNIT HEAD - EXPORT FILTERED TICKETS EXCEL - ✅ FIXED ERP ID
# ============================================================
def unit_head_export_filtered_tickets_excel(request, tickets_qs, unit):
    """
    Export filtered tickets to Excel for Unit Head
    ✅ FIXED: Now fetches ERP ID from ERPHolderMapping model
    """
    current_tz = timezone.get_current_timezone()
    now_utc = timezone.now()
    if timezone.is_naive(now_utc):
        now_utc = timezone.make_aware(now_utc, timezone.utc)
    now_local = now_utc.astimezone(current_tz)
    report_time = now_local.strftime('%d-%b-%Y %I:%M:%S %p')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Tickets_Report_{unit.code}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # ✅ Build mapping of employee_id to ERP ID
    employee_ids = tickets_qs.values_list('employee_id', flat=True).distinct()
    erp_mappings = {}
    if employee_ids:
        erp_mappings_qs = ERPHolderMapping.objects.filter(
            employee__employee_id__in=employee_ids
        ).select_related('employee')
        for mapping in erp_mappings_qs:
            erp_mappings[mapping.employee.employee_id] = mapping.erp_user_id
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Tickets - {unit.code}"
    
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10)
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    ws.merge_cells('A1:AA1')
    ws['A1'] = f"TICKETS REPORT - {unit.code} - {unit.full_name}"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45
    
    ws.merge_cells('A2:AA2')
    ws['A2'] = f"Generated: {report_time}  |  Total Tickets: {tickets_qs.count()}"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    
    headers = [
        'Ticket Number', 'Status', 'Unit Code', 'Unit Name', 'Department',
        'Employee ID', 'ERP ID', 'Employee Name', 'Mobile', 'Email', 'Screen/Module',
        'Subject', 'Description', 'Priority', 'Error Type', 'Created By Role',
        'Assigned Person', 'Hold Reason', 'Closing Remarks', 'Closed By',
        'Vendor Ticket', 'Main Error Type', 'Sub Error Type',
        'Created At', 'Closed At', 'Time to Close', 'Escalated At'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[4].height = 30
    
    row_idx = 5
    for ticket in tickets_qs:
        created_at_local = ticket.created_at.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p') if ticket.created_at else ''
        closed_at_local = ticket.closed_at.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p') if ticket.closed_at else ''
        escalated_at_local = ticket.escalated_at.astimezone(current_tz).strftime('%d-%b-%Y %I:%M:%S %p') if ticket.escalated_at else ''
        
        time_to_close = ''
        if ticket.created_at and ticket.closed_at:
            duration = ticket.closed_at - ticket.created_at
            days = duration.days
            hours = duration.seconds // 3600
            minutes = (duration.seconds % 3600) // 60
            time_to_close = f"{days}d {hours}h {minutes}m" if days > 0 else f"{hours}h {minutes}m"
        
        # ✅ Get ERP ID from mapping
        erp_id = erp_mappings.get(ticket.employee_id, 'Not Mapped')
        
        row_data = [
            ticket.ticket_number, ticket.status, ticket.unit.code if ticket.unit else '',
            ticket.unit.full_name if ticket.unit else '', ticket.department.name if ticket.department else '',
            ticket.employee_id, erp_id, ticket.employee_name, ticket.mobile or '', ticket.email or '',
            ticket.screen_number, ticket.subject, ticket.description or '', ticket.priority,
            ticket.error_type or '', ticket.created_by_role, ticket.assigned_person or '',
            ticket.hold_reason or '', ticket.closing_remarks or '', ticket.closed_by or '',
            ticket.vendor_ticket_number or '', ticket.main_error_type or 'N/A',
            ticket.sub_error_type or 'N/A', created_at_local, closed_at_local, time_to_close,
            escalated_at_local,
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        row_idx += 1
    
    column_widths = {
        'A': 18, 'B': 14, 'C': 12, 'D': 25, 'E': 20, 'F': 14, 'G': 16,
        'H': 22, 'I': 16, 'J': 25, 'K': 16, 'L': 30, 'M': 40, 'N': 14,
        'O': 20, 'P': 18, 'Q': 20, 'R': 30, 'S': 18, 'T': 18,
        'U': 18, 'V': 22, 'W': 22, 'X': 22, 'Y': 22, 'Z': 16, 'AA': 22
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    wb.save(response)
    return response