# tickets/views/settings_actions/employees.py

"""
Employee Management - Add, Edit, Toggle, Delete, Bulk Upload, Downloads
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.db import IntegrityError
from django.contrib import messages
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
import logging

from tickets.models import EmployeeMaster, Unit, Department, SettingsAuditLog
from .settings_audit import log_settings_change
from ..utils import is_admin, get_client_ip

logger = logging.getLogger(__name__)


@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_employees(request):
    """
    Complete Employee Management Actions:
    - Add Employee
    - Edit Employee
    - Toggle Active/Inactive
    - Toggle Can Assign
    - Delete Employee (Must be deactivated first)
    - Bulk Upload (AJAX)
    POST: action, employee_id, employee_name, mobile, email, unit, department, can_assign_ticket
    Redirects to: settings_employees_page
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # ========== ADD EMPLOYEE ==========
        if action == 'add_employee':
            eid = request.POST.get('employee_id','').strip().upper()
            ename = request.POST.get('employee_name','').strip().upper()
            mob = request.POST.get('mobile','').strip()
            email = request.POST.get('email','').strip()
            uid = request.POST.get('unit')
            did = request.POST.get('department')
            
            # Only Employee ID and Name are mandatory
            if not eid:
                messages.error(request, "Employee ID is required.")
                return redirect('settings_employees_page')
            if not ename:
                messages.error(request, "Employee Name is required.")
                return redirect('settings_employees_page')
            
            # Validate mobile if provided
            if mob and (not mob.isdigit() or len(mob) != 10):
                messages.error(request, f'Mobile number must be exactly 10 digits: {mob}')
                return redirect('settings_employees_page')
            
            # Validate email if provided
            if email and ('@' not in email or '.' not in email):
                messages.error(request, f'Invalid email format: {email}')
                return redirect('settings_employees_page')
            
            # Check for duplicate employee_id (case-insensitive)
            existing_employee = EmployeeMaster.objects.filter(employee_id__iexact=eid).first()
            if existing_employee:
                messages.error(request, f'❌ Employee ID "{eid}" already exists in the database.')
                return redirect('settings_employees_page')
            
            try:
                emp = EmployeeMaster.objects.create(
                    employee_id=eid,
                    employee_name=ename,
                    mobile=mob or None,
                    email=email or None,
                    unit_id=uid or None,
                    department_id=did or None,
                    is_active=True
                )
                messages.success(request, f'✅ Employee "{eid}" added successfully.')
                
                log_settings_change(
                    request,
                    action_type='CREATE',
                    setting_type='EMPLOYEE',
                    setting_name=f"Employee: {eid} - {ename}",
                    new_value=f"ID: {eid}, Name: {ename}, Mobile: {mob or 'Not provided'}, Email: {email or 'Not provided'}",
                    change_summary=f"Added employee {eid} - {ename}",
                    remarks=f"Employee added by {request.user.username}"
                )
            except IntegrityError as e:
                logger.error(f"IntegrityError while adding employee {eid}: {str(e)}")
                messages.error(request, f'❌ Employee ID "{eid}" already exists.')
            except Exception as e:
                logger.error(f"Error adding employee {eid}: {str(e)}")
                messages.error(request, f'❌ Error adding employee: {str(e)}')
        
        # ========== BULK UPLOAD ==========
        elif action == 'bulk_upload':
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': 'Please select an Excel file.'
                    })
                messages.error(request, "Please select an Excel file.")
                return redirect('settings_employees_page')
            
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': 'Invalid file format. Only .xlsx and .xls files are supported.'
                    })
                messages.error(request, "Invalid file format. Only .xlsx and .xls files are supported.")
                return redirect('settings_employees_page')
            
            try:
                df = pd.read_excel(excel_file, dtype=str)
                
                if df.empty:
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': 'The uploaded file is empty.'
                        })
                    messages.error(request, "The uploaded file is empty.")
                    return redirect('settings_employees_page')
                
                # Required columns - Email and Mobile are optional
                required_columns = ['Employee ID', 'Employee Name']
                missing_columns = []
                for col in required_columns:
                    if col not in df.columns:
                        found = False
                        for existing_col in df.columns:
                            if existing_col.lower() == col.lower():
                                found = True
                                break
                        if not found:
                            missing_columns.append(col)
                
                if missing_columns:
                    error_msg = f"Missing required columns: {', '.join(missing_columns)}"
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': error_msg
                        })
                    messages.error(request, error_msg)
                    return redirect('settings_employees_page')
                
                validation_errors = []
                successful_rows = 0
                
                all_units = {unit.code.upper(): unit for unit in Unit.objects.filter(is_active=True)}
                all_departments = {dept.name.upper(): dept for dept in Department.objects.filter(is_active=True)}
                
                for idx, row in df.iterrows():
                    row_num = idx + 2
                    
                    eid = None
                    ename = None
                    mob = None
                    email = None
                    uc = None
                    dn = None
                    
                    for col in df.columns:
                        col_lower = col.lower()
                        if col_lower in ['employee id', 'employee_id', 'employeeid']:
                            eid = str(row.get(col, '')).strip().upper()
                        elif col_lower in ['employee name', 'employee_name', 'employeename', 'name']:
                            ename = str(row.get(col, '')).strip().upper()
                        elif col_lower in ['mobile', 'phone', 'contact']:
                            mob = str(row.get(col, '')).strip()
                        elif col_lower in ['email', 'email id', 'email_id', 'emailid']:
                            email = str(row.get(col, '')).strip().lower()
                        elif col_lower in ['unit code', 'unit_code', 'unitcode', 'unit']:
                            uc = str(row.get(col, '')).strip().upper()
                        elif col_lower in ['department', 'dept', 'department name', 'department_name']:
                            dn = str(row.get(col, '')).strip().upper()
                    
                    # Only Employee ID and Name are mandatory
                    if not eid:
                        validation_errors.append({'row': row_num, 'message': 'Employee ID is required'})
                        continue
                    if not ename:
                        validation_errors.append({'row': row_num, 'message': 'Employee Name is required'})
                        continue
                    
                    # Validate mobile if provided
                    if mob and (not mob.isdigit() or len(mob) != 10):
                        validation_errors.append({'row': row_num, 'message': f'Mobile number must be 10 digits: {mob}'})
                        continue
                    
                    # Validate email if provided
                    if email and ('@' not in email or '.' not in email):
                        validation_errors.append({'row': row_num, 'message': f'Invalid email format: {email}'})
                        continue
                    
                    if uc and uc not in all_units:
                        valid_units = ', '.join(list(all_units.keys())[:5])
                        if len(all_units) > 5:
                            valid_units += f' and {len(all_units) - 5} more'
                        validation_errors.append({'row': row_num, 'message': f'Invalid Unit Code "{uc}". Valid units: {valid_units}'})
                        continue
                    if dn and dn not in all_departments:
                        valid_depts = ', '.join(list(all_departments.keys())[:5])
                        if len(all_departments) > 5:
                            valid_depts += f' and {len(all_departments) - 5} more'
                        validation_errors.append({'row': row_num, 'message': f'Invalid Department "{dn}". Valid departments: {valid_depts}'})
                        continue
                    
                    successful_rows += 1
                
                if validation_errors:
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': f'Validation failed. Found {len(validation_errors)} error(s).',
                            'errors': validation_errors
                        })
                    messages.error(request, f'Validation failed. Found {len(validation_errors)} error(s).')
                    for err in validation_errors[:5]:
                        messages.error(request, f'Row {err["row"]}: {err["message"]}')
                    if len(validation_errors) > 5:
                        messages.error(request, f'... and {len(validation_errors) - 5} more errors.')
                    return redirect('settings_employees_page')
                
                success_count = 0
                error_count = 0
                update_count = 0
                
                for idx, row in df.iterrows():
                    try:
                        eid = None
                        ename = None
                        mob = None
                        email = None
                        uc = None
                        dn = None
                        
                        for col in df.columns:
                            col_lower = col.lower()
                            if col_lower in ['employee id', 'employee_id', 'employeeid']:
                                eid = str(row.get(col, '')).strip().upper()
                            elif col_lower in ['employee name', 'employee_name', 'employeename', 'name']:
                                ename = str(row.get(col, '')).strip().upper()
                            elif col_lower in ['mobile', 'phone', 'contact']:
                                mob = str(row.get(col, '')).strip()
                            elif col_lower in ['email', 'email id', 'email_id', 'emailid']:
                                email = str(row.get(col, '')).strip().lower()
                            elif col_lower in ['unit code', 'unit_code', 'unitcode', 'unit']:
                                uc = str(row.get(col, '')).strip().upper()
                            elif col_lower in ['department', 'dept', 'department name', 'department_name']:
                                dn = str(row.get(col, '')).strip().upper()
                        
                        if not eid or not ename:
                            error_count += 1
                            continue
                        
                        # Validate mobile if provided
                        if mob and (not mob.isdigit() or len(mob) != 10):
                            error_count += 1
                            continue
                        
                        # Validate email if provided
                        if email and ('@' not in email or '.' not in email):
                            error_count += 1
                            continue
                        
                        unit_obj = all_units.get(uc) if uc else None
                        dept_obj = all_departments.get(dn) if dn else None
                        
                        # Check if employee already exists (case-insensitive)
                        existing_emp = EmployeeMaster.objects.filter(employee_id__iexact=eid).first()
                        if existing_emp:
                            # Update existing employee
                            existing_emp.employee_name = ename
                            existing_emp.mobile = mob or None
                            existing_emp.email = email or None
                            existing_emp.unit = unit_obj
                            existing_emp.department = dept_obj
                            existing_emp.is_active = True
                            existing_emp.save()
                            update_count += 1
                            logger.info(f"Updated existing employee: {eid}")
                        else:
                            # Create new employee
                            EmployeeMaster.objects.create(
                                employee_id=eid,
                                employee_name=ename,
                                mobile=mob or None,
                                email=email or None,
                                unit=unit_obj,
                                department=dept_obj,
                                is_active=True
                            )
                            success_count += 1
                            logger.info(f"Created new employee: {eid}")
                        
                    except Exception as e:
                        error_count += 1
                        logger.error(f"Error processing row {idx+2}: {str(e)}")
                
                # Log bulk upload summary
                if success_count > 0 or update_count > 0:
                    log_settings_change(
                        request,
                        action_type='CREATE',
                        setting_type='EMPLOYEE',
                        setting_name=f"Bulk Upload: {success_count + update_count} employees",
                        new_value=f"Created {success_count} new employees, Updated {update_count} existing",
                        change_summary=f"Bulk upload processed {success_count + update_count} employees",
                        remarks=f"Bulk upload by {request.user.username}"
                    )
                
                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'message': f'✅ Created {success_count} new employees, Updated {update_count} existing. {error_count} skipped.'
                    })
                
                if success_count > 0:
                    messages.success(request, f'✅ Created {success_count} new employees.')
                if update_count > 0:
                    messages.success(request, f'✅ Updated {update_count} existing employees.')
                if error_count > 0:
                    messages.warning(request, f'⚠️ {error_count} rows were skipped due to errors.')
                
            except Exception as e:
                logger.error(f"Bulk upload error: {str(e)}")
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': f'Error processing file: {str(e)}'
                    })
                messages.error(request, f'Error processing file: {str(e)}')
        
        # ========== EDIT EMPLOYEE ==========
        elif action == 'edit_employee':
            emp = get_object_or_404(EmployeeMaster, pk=request.POST.get('emp_id'))
            
            old_id = emp.employee_id
            old_name = emp.employee_name
            old_mobile = emp.mobile
            old_email = emp.email
            old_unit = emp.unit.code if emp.unit else 'None'
            old_dept = emp.department.name if emp.department else 'None'
            old_can_assign = 'Yes' if emp.can_assign_ticket else 'No'
            
            new_eid = request.POST.get('employee_id','').strip().upper()
            new_ename = request.POST.get('employee_name','').strip().upper()
            new_mobile = request.POST.get('mobile','').strip()
            new_email = request.POST.get('email','').strip()
            new_unit = request.POST.get('unit') or None
            new_dept = request.POST.get('department') or None
            
            # Validate required fields
            if not new_eid:
                messages.error(request, "Employee ID is required.")
                return redirect('settings_employees_page')
            if not new_ename:
                messages.error(request, "Employee Name is required.")
                return redirect('settings_employees_page')
            
            # Validate mobile if provided
            if new_mobile and (not new_mobile.isdigit() or len(new_mobile) != 10):
                messages.error(request, f'Mobile number must be exactly 10 digits: {new_mobile}')
                return redirect('settings_employees_page')
            
            # Validate email if provided
            if new_email and ('@' not in new_email or '.' not in new_email):
                messages.error(request, f'Invalid email format: {new_email}')
                return redirect('settings_employees_page')
            
            # Check for duplicate employee_id if changed (case-insensitive)
            if new_eid != old_id:
                existing_employee = EmployeeMaster.objects.filter(employee_id__iexact=new_eid).exclude(pk=emp.pk).first()
                if existing_employee:
                    messages.error(request, f'❌ Employee ID "{new_eid}" already exists in the database.')
                    return redirect('settings_employees_page')
            
            emp.employee_id = new_eid
            emp.employee_name = new_ename
            emp.mobile = new_mobile or None
            emp.email = new_email or None
            emp.unit_id = new_unit
            emp.department_id = new_dept
            emp.can_assign_ticket = request.POST.get('can_assign_ticket') == 'on'
            
            try: 
                emp.save()
                messages.success(request, '✅ Employee updated successfully.')
                
                change_details = []
                if old_id != emp.employee_id:
                    change_details.append(f"ID: {old_id} → {emp.employee_id}")
                if old_name != emp.employee_name:
                    change_details.append(f"Name: {old_name} → {emp.employee_name}")
                if old_mobile != emp.mobile:
                    change_details.append(f"Mobile: {old_mobile or 'None'} → {emp.mobile or 'None'}")
                if old_email != emp.email:
                    change_details.append(f"Email: {old_email or 'None'} → {emp.email or 'None'}")
                if old_unit != (emp.unit.code if emp.unit else 'None'):
                    change_details.append(f"Unit: {old_unit} → {emp.unit.code if emp.unit else 'None'}")
                if old_dept != (emp.department.name if emp.department else 'None'):
                    change_details.append(f"Department: {old_dept} → {emp.department.name if emp.department else 'None'}")
                if old_can_assign != ('Yes' if emp.can_assign_ticket else 'No'):
                    change_details.append(f"Can Assign: {old_can_assign} → {'Yes' if emp.can_assign_ticket else 'No'}")
                
                log_settings_change(
                    request,
                    action_type='UPDATE',
                    setting_type='EMPLOYEE',
                    setting_name=f"Employee: {emp.employee_id} - {emp.employee_name}",
                    old_value=f"ID: {old_id}, Name: {old_name}, Mobile: {old_mobile or 'None'}, Email: {old_email or 'None'}",
                    new_value=f"ID: {emp.employee_id}, Name: {emp.employee_name}, Mobile: {emp.mobile or 'None'}, Email: {emp.email or 'None'}",
                    change_summary='; '.join(change_details) if change_details else 'Employee updated',
                    remarks=f"Employee updated by {request.user.username}"
                )
            except IntegrityError as e:
                logger.error(f"IntegrityError while editing employee: {str(e)}")
                messages.error(request, '❌ Employee ID already exists.')
            except Exception as e:
                logger.error(f"Error editing employee: {str(e)}")
                messages.error(request, f'❌ Error updating employee: {str(e)}')
        
        # ========== TOGGLE EMPLOYEE ACTIVE/INACTIVE ==========
        elif action == 'toggle_employee':
            emp = get_object_or_404(EmployeeMaster, pk=request.POST.get('emp_id'))
            old_status = 'Active' if emp.is_active else 'Inactive'
            emp.is_active = not emp.is_active
            emp.save()
            new_status = 'Active' if emp.is_active else 'Inactive'
            
            messages.success(request, f'✅ Employee {"activated" if emp.is_active else "deactivated"}.')
            
            log_settings_change(
                request,
                action_type='TOGGLE',
                setting_type='EMPLOYEE',
                setting_name=f"Employee: {emp.employee_id} - {emp.employee_name}",
                old_value=f"Status: {old_status}",
                new_value=f"Status: {new_status}",
                change_summary=f"Status changed from {old_status} to {new_status}",
                remarks=f"Employee toggled by {request.user.username}"
            )
        
        # ========== TOGGLE CAN ASSIGN ==========
        elif action == 'toggle_can_assign':
            emp = get_object_or_404(EmployeeMaster, pk=request.POST.get('emp_id'))
            old_value = 'Yes' if emp.can_assign_ticket else 'No'
            emp.can_assign_ticket = not emp.can_assign_ticket
            emp.save()
            new_value = 'Yes' if emp.can_assign_ticket else 'No'
            
            status = "enabled" if emp.can_assign_ticket else "disabled"
            messages.success(request, f'✅ Employee "{emp.employee_id}" assignment {status}.')
            
            log_settings_change(
                request,
                action_type='UPDATE',
                setting_type='EMPLOYEE',
                setting_name=f"Employee: {emp.employee_id} - {emp.employee_name}",
                old_value=f"Can Assign: {old_value}",
                new_value=f"Can Assign: {new_value}",
                change_summary=f"Can Assign changed from {old_value} to {new_value}",
                remarks=f"Assignment toggled by {request.user.username}"
            )
        
        # ========== DELETE EMPLOYEE - MUST BE DEACTIVATED FIRST ==========
        elif action == 'delete_employee':
            emp = get_object_or_404(EmployeeMaster, pk=request.POST.get('emp_id'))
            eid = emp.employee_id
            ename = emp.employee_name
            
            # Check if employee is active
            if emp.is_active:
                messages.error(request, f'❌ Cannot delete "{eid}" because they are still ACTIVE. Please deactivate the employee first, then try deleting again.')
                return redirect('settings_employees_page')
            
            # Employee is inactive, proceed with deletion
            emp_mobile = emp.mobile
            emp_email = emp.email
            
            # Log before deleting
            log_settings_change(
                request,
                action_type='DELETE',
                setting_type='EMPLOYEE',
                setting_name=f"Employee: {eid} - {ename}",
                old_value=f"ID: {eid}, Name: {ename}, Mobile: {emp_mobile or 'None'}, Email: {emp_email or 'None'}, Status: Inactive",
                change_summary=f"Permanently deleted employee {eid} - {ename}",
                remarks=f"Employee permanently deleted by {request.user.username} (was already deactivated)"
            )
            
            # Hard delete from database
            emp.delete()
            
            messages.success(request, f'✅ Employee "{eid}" has been permanently deleted from the database.')
    
    return redirect('settings_employees_page')


@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def download_employee_list(request):
    """
    Download complete employee list as Excel with all fields
    Includes: Employee ID, Name, Mobile, Email, Unit, Department, Status, Can Assign
    """
    emps = EmployeeMaster.objects.all().order_by('employee_id')
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename=Employee_List_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    tf = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    hf = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    df = Font(name='Calibri', size=11)
    tfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hfill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    
    ws.merge_cells('A1:I1')
    ws['A1'] = "Employee Directory"
    ws['A1'].font = tf
    ws['A1'].fill = tfill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    for ci, h in enumerate(['Employee ID', 'Name', 'Mobile', 'Email', 'Unit Code', 'Unit Name', 'Department', 'Status', 'Can Assign'], 1):
        c = ws.cell(row=3, column=ci)
        c.value = h
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 25
    
    for ri, emp in enumerate(emps, 4):
        rd = [
            emp.employee_id,
            emp.employee_name,
            emp.mobile or '',
            emp.email or '',
            emp.unit.code if emp.unit else '',
            emp.unit.full_name if emp.unit else '',
            emp.department.name if emp.department else '',
            'Active' if emp.is_active else 'Inactive',
            'Yes' if emp.can_assign_ticket else 'No'
        ]
        for ci, v in enumerate(rd, 1):
            c = ws.cell(row=ri, column=ci)
            c.value = v
            c.font = df
    
    for col in ws.columns:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(
            max(len(str(c.value or '')) for c in col if c.row > 1) + 3, 12
        )
    wb.save(resp)
    return resp


@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def download_employee_template(request):
    """
    Download Excel template for bulk employee upload
    Columns: Employee ID, Employee Name, Mobile, Email, Unit Code, Department
    Note: Mobile and Email are optional
    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Employee_Upload_Template.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Template"
    
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    headers = ['Employee ID', 'Employee Name', 'Mobile', 'Email', 'Unit Code', 'Department']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    sample_data = [
        ['EMP001', 'JOHN DOE', '9876543210', 'john.doe@company.com', 'GPL', 'Production'],
        ['EMP002', 'JANE SMITH', '9876543211', 'jane.smith@company.com', 'GPLAST', 'QA'],
        ['EMP003', 'MIKE JOHNSON', '9876543212', 'mike.johnson@company.com', 'IMD', 'Purchase'],
        ['EMP004', 'SARAH WILLIAMS', '', '', 'GPL', 'Sales'],
        ['EMP005', 'DAVID BROWN', '9876543213', '', 'GPLAST', 'Finance'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = Font(name='Calibri', size=11)
    
    note_row = len(sample_data) + 3
    note_cell = ws.cell(row=note_row, column=1)
    note_cell.value = "Mandatory fields: Employee ID, Employee Name. Mobile and Email are optional."
    note_cell.font = Font(name='Calibri', size=10, italic=True, color='FF0000')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
    
    validation_row = note_row + 1
    validation_cell = ws.cell(row=validation_row, column=1)
    validation_cell.value = "Note: Unit Code must match existing active units in the system. Department must match existing active departments."
    validation_cell.font = Font(name='Calibri', size=9, italic=True, color='666666')
    ws.merge_cells(start_row=validation_row, start_column=1, end_row=validation_row, end_column=6)
    
    column_widths = [15, 20, 15, 25, 12, 15]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    wb.save(response)
    return response