# tickets/views/settings_action/screen_mapping.py

"""
Screen Mapping - Add, Remove, Delete ERP, Export Excel, Page View, AJAX, Bulk Upload
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator
import json
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.utils import timezone
import logging

from tickets.models import ScreenMaster, ScreenMapping, ERPHolderMapping
from .settings_audit import log_settings_change
from ..utils import is_admin

logger = logging.getLogger(__name__)


# ============================================================
# SCREEN MAPPING - ADD (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def screen_mapping_add(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=400)

    screen_id = request.POST.get('screen_id', '').strip()
    erp_user_id = request.POST.get('erp_user_id', '').strip()

    if not screen_id:
        return JsonResponse({'success': False, 'message': 'Screen is required'})
    if not erp_user_id:
        return JsonResponse({'success': False, 'message': 'ERP User ID is required'})

    try:
        screen = ScreenMaster.objects.get(id=screen_id)
    except ScreenMaster.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Screen not found'})

    if ScreenMapping.objects.filter(screen=screen, erp_user_id=erp_user_id).exists():
        return JsonResponse({'success': False, 'message': f'Mapping already exists: {screen.screen_code} â†’ ERP {erp_user_id}'})

    try:
        mapping = ScreenMapping.objects.create(
            screen=screen, erp_user_id=erp_user_id, created_by=request.user.username
        )
        log_settings_change(
            request, 'CREATE', 'SCREEN',
            f'{screen.screen_code} â†’ ERP {erp_user_id}',
            new_value=f'Screen: {screen.screen_code}, ERP ID: {erp_user_id}',
            change_summary=f'Mapped screen {screen.screen_code} to ERP {erp_user_id}'
        )
        # Get employee info for response
        emp_mappings = ERPHolderMapping.objects.filter(erp_user_id=erp_user_id).select_related('employee')
        emp_names = ', '.join([m.employee.employee_name for m in emp_mappings]) or 'N/A'
        return JsonResponse({
            'success': True,
            'message': f'Screen mapped successfully',
            'mapping': {
                'id': mapping.id,
                'screen_id': screen.id,
                'screen_code': screen.screen_code,
                'screen_name': screen.screen_name,
                'erp_user_id': erp_user_id,
                'employee_names': emp_names,
            }
        })
    except Exception as e:
        logger.error(f'Error adding screen mapping: {e}')
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================================
# SCREEN MAPPING - REMOVE (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def screen_mapping_remove(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=400)

    mapping_id = request.POST.get('mapping_id', '').strip()
    if not mapping_id:
        return JsonResponse({'success': False, 'message': 'Mapping ID is required'})

    try:
        mapping = ScreenMapping.objects.select_related('screen').get(id=mapping_id)
    except ScreenMapping.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Mapping not found'})

    screen_code = mapping.screen.screen_code
    erp_user_id = mapping.erp_user_id
    mapping.delete()

    log_settings_change(
        request, 'DELETE', 'SCREEN',
        f'{screen_code} â†’ ERP {erp_user_id}',
        old_value=f'Screen: {screen_code}, ERP ID: {erp_user_id}',
        change_summary=f'Removed screen mapping: {screen_code} from ERP {erp_user_id}'
    )
    return JsonResponse({'success': True, 'message': 'Mapping removed successfully'})


# ============================================================
# SCREEN MAPPING - DELETE ALL FOR ERP ID (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def screen_mapping_delete_erp(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=400)

    erp_user_id = request.POST.get('erp_user_id', '').strip()
    if not erp_user_id:
        return JsonResponse({'success': False, 'message': 'ERP User ID is required'})

    mappings = ScreenMapping.objects.filter(erp_user_id=erp_user_id)
    count = mappings.count()

    if count == 0:
        return JsonResponse({'success': False, 'message': f'No mappings found for ERP ID "{erp_user_id}"'})

    mappings.delete()

    log_settings_change(
        request, 'DELETE', 'SCREEN',
        f'All mappings for ERP {erp_user_id}',
        old_value=f'Removed {count} mappings for ERP ID: {erp_user_id}',
        change_summary=f'Deleted all {count} screen mappings for ERP ID {erp_user_id}'
    )
    return JsonResponse({'success': True, 'message': f'Successfully deleted {count} mappings for ERP ID "{erp_user_id}"'})


# ============================================================
# SCREEN MAPPING - EXPORT EXCEL
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def screen_mapping_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.utils import timezone
    from django.http import HttpResponse

    erp_filter = request.GET.get('erp_user_id', '').strip()
    emp_filter = request.GET.get('employee_id', '').strip()

    mappings = ScreenMapping.objects.all().select_related('screen').order_by('erp_user_id', 'screen__screen_name')

    if erp_filter:
        mappings = mappings.filter(erp_user_id=erp_filter)
    if emp_filter:
        erp_ids = ERPHolderMapping.objects.filter(
            employee__employee_id__icontains=emp_filter
        ).values_list('erp_user_id', flat=True).distinct()
        mappings = mappings.filter(erp_user_id__in=erp_ids)

    current_tz = timezone.get_current_timezone()
    now_local = timezone.now().astimezone(current_tz)
    report_time = now_local.strftime('%d-%b-%Y %I:%M:%S %p')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Screen_Mapping_{now_local.strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Screen Mapping'

    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10)
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0')
    )

    ws.merge_cells('A1:G1')
    ws['A1'] = f'SCREEN MAPPING REPORT - Generated: {report_time}  |  Total: {mappings.count()}'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    headers = ['#', 'ERP User ID', 'Employee ID(s)', 'Employee Name(s)', 'Screen Code', 'Screen Name', 'Mapped On']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[3].height = 25

    # Build ERPâ†’Employee lookup
    all_erp_ids = mappings.values_list('erp_user_id', flat=True).distinct()
    erp_emp_map = {}
    for erp_id in all_erp_ids:
        emps = ERPHolderMapping.objects.filter(erp_user_id=erp_id).select_related('employee')
        erp_emp_map[erp_id] = emps

    for ri, mapping in enumerate(mappings, 1):
        emps = erp_emp_map.get(mapping.erp_user_id, [])
        emp_ids = ', '.join([e.employee.employee_id for e in emps]) or 'Not Mapped'
        emp_names = ', '.join([e.employee.employee_name for e in emps]) or 'Not Mapped'
        mapped_on = mapping.created_at.astimezone(current_tz).strftime('%d-%b-%Y %I:%M %p') if mapping.created_at else ''
        row_data = [ri, mapping.erp_user_id, emp_ids, emp_names, mapping.screen.screen_code, mapping.screen.screen_name, mapped_on]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri + 3, column=ci)
            cell.value = val
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border

    for col_letter, width in {'A': 8, 'B': 16, 'C': 22, 'D': 35, 'E': 18, 'F': 40, 'G': 22}.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(response)
    return response


# ============================================================
# SCREEN MAPPING PAGE VIEW
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def settings_screen_mapping_page(request):
    """
    Screen Mapping Settings Page - Compact Table View with Modal
    Shows ERP IDs in a table with screen count, click to view screens in modal
    """
    erp_filter = request.GET.get('erp_user_id', '').strip()
    screen_search = request.GET.get('screen_search', '').strip()

    # Get all mappings with related screen data
    mappings = ScreenMapping.objects.all().select_related('screen').order_by('erp_user_id', 'screen__screen_name')

    # Apply filters
    if erp_filter:
        mappings = mappings.filter(erp_user_id=erp_filter)
    if screen_search:
        mappings = mappings.filter(
            Q(screen__screen_code__icontains=screen_search) |
            Q(screen__screen_name__icontains=screen_search)
        )

    # Get all ERP User IDs for dropdown - from ERPHolderMapping
    all_erp_user_ids = ERPHolderMapping.objects.values_list('erp_user_id', flat=True).distinct().order_by('erp_user_id')
    
    # Get all screens for dropdown
    screens = ScreenMaster.objects.filter(is_active=True).order_by('screen_name')

    # Group mappings by ERP User ID - COMPACT DATA
    erp_data = []
    erp_screens_map = {}  # Store screens for modal view
    
    for erp_id in all_erp_user_ids:
        erp_mappings = mappings.filter(erp_user_id=erp_id)
        screen_count = erp_mappings.count()
        
        # Get screens for this ERP (for modal)
        screen_list = []
        for mapping in erp_mappings:
            screen_list.append({
                'mapping_id': mapping.id,
                'screen_code': mapping.screen.screen_code,
                'screen_name': mapping.screen.screen_name,
                'screen_type': mapping.screen.screen_type,
            })
        
        erp_screens_map[erp_id] = screen_list
        
        erp_data.append({
            'erp_user_id': erp_id,
            'screen_count': screen_count,
            'screens': screen_list,
        })

    # Paginate the ERP data (15 per page)
    paginator = Paginator(erp_data, 15)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Convert erp_screens_map to JSON for template
    erp_screens_json = json.dumps(erp_screens_map)

    context = {
        'page_obj': page_obj,
        'erp_data': page_obj,
        'erp_user_ids': all_erp_user_ids,
        'screens': screens,
        'erp_screens_map': erp_screens_map,
        'erp_screens_json': erp_screens_json,
        'erp_filter': erp_filter,
        'screen_search': screen_search,
        'total_erps': len(erp_data),
        'total_mappings': mappings.count(),
    }
    return render(request, 'admin_panel/settings_screen_mapping.html', context)


# ============================================================
# AJAX - GET SCREENS FOR ERP USER ID
# ============================================================
@login_required
def ajax_get_screens_for_erp(request):
    """Return screens mapped to a given ERP User ID (for ticket creation)"""
    erp_user_id = request.GET.get('erp_user_id', '').strip()
    if not erp_user_id:
        return JsonResponse({'success': False, 'screens': []})

    screens = ScreenMapping.objects.filter(
        erp_user_id=erp_user_id
    ).select_related('screen').order_by('screen__screen_name')

    screen_list = [{
        'id': sm.screen.id,
        'screen_code': sm.screen.screen_code,
        'screen_name': sm.screen.screen_name,
        'display': f'{sm.screen.screen_code} - {sm.screen.screen_name} ({sm.screen.get_screen_type_display()})'
    } for sm in screens]

    return JsonResponse({'success': True, 'screens': screen_list})


# ============================================================
# âœ… NEW: SCREEN MAPPING - BULK UPLOAD (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def screen_mapping_bulk_upload(request):
    """
    Bulk upload screen mappings from Excel/CSV
    Columns: ERP User ID, Screen Code
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=400)
    
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return JsonResponse({'success': False, 'message': 'Please select an Excel file.'})
    
    if not excel_file.name.endswith(('.xlsx', '.xls', '.csv')):
        return JsonResponse({'success': False, 'message': 'Invalid file format. Only .xlsx, .xls, and .csv files are supported.'})
    
    try:
        # Read the file
        if excel_file.name.endswith('.csv'):
            df = pd.read_csv(excel_file, dtype=str)
        else:
            df = pd.read_excel(excel_file, dtype=str)
        
        if df.empty:
            return JsonResponse({'success': False, 'message': 'The uploaded file is empty.'})
        
        # Find columns
        erp_column = None
        screen_column = None
        
        for col in df.columns:
            col_lower = col.strip().lower()
            if 'erp' in col_lower and ('user' in col_lower or 'id' in col_lower):
                erp_column = col
            elif 'screen' in col_lower and ('code' in col_lower or 'id' in col_lower):
                screen_column = col
        
        # If columns not found, use first two columns
        if erp_column is None:
            erp_column = df.columns[0]
        if screen_column is None:
            screen_column = df.columns[1] if len(df.columns) > 1 else df.columns[0]
        
        # Get all ERP IDs from system
        existing_erp_ids = set(ERPHolderMapping.objects.values_list('erp_user_id', flat=True))
        
        # Get all screens from Screen Master (map by screen_code)
        screens_by_code = {}
        for screen in ScreenMaster.objects.filter(is_active=True):
            screens_by_code[screen.screen_code.upper()] = screen
        
        # Get existing mappings to check duplicates (erp_id + screen_id)
        existing_mappings = set()
        for mapping in ScreenMapping.objects.all():
            existing_mappings.add((mapping.erp_user_id, mapping.screen_id))
        
        # Process rows
        added_count = 0
        skipped_count = 0
        errors = []
        added_mappings = []
        
        for idx, row in df.iterrows():
            row_num = idx + 2
            erp_id = str(row.get(erp_column, '')).strip()
            screen_code = str(row.get(screen_column, '')).strip().upper()
            
            if not erp_id:
                errors.append(f'Row {row_num}: ERP User ID is empty')
                skipped_count += 1
                continue
            
            if not screen_code:
                errors.append(f'Row {row_num}: Screen Code is empty')
                skipped_count += 1
                continue
            
            # Check if ERP ID exists
            if erp_id not in existing_erp_ids:
                errors.append(f'Row {row_num}: ERP ID "{erp_id}" not found in system')
                skipped_count += 1
                continue
            
            # Find screen by code
            screen = screens_by_code.get(screen_code)
            if not screen:
                errors.append(f'Row {row_num}: Screen Code "{screen_code}" not found in Screen Master')
                skipped_count += 1
                continue
            
            # Check if mapping already exists
            if (erp_id, screen.id) in existing_mappings:
                errors.append(f'Row {row_num}: Mapping already exists for ERP "{erp_id}" â†’ Screen "{screen_code}"')
                skipped_count += 1
                continue
            
            # Create mapping
            try:
                mapping = ScreenMapping.objects.create(
                    erp_user_id=erp_id,
                    screen=screen,
                    created_by=request.user.username
                )
                added_count += 1
                added_mappings.append(f'{erp_id} â†’ {screen_code}')
                # Add to existing mappings set
                existing_mappings.add((erp_id, screen.id))
            except Exception as e:
                errors.append(f'Row {row_num}: Error creating mapping - {str(e)}')
                skipped_count += 1
        
        # Log bulk upload
        if added_count > 0:
            log_settings_change(
                request,
                action_type='CREATE',
                setting_type='SCREEN_MAPPING',
                setting_name=f'Bulk Upload: {added_count} screen mappings',
                new_value=f'Added {added_count} mappings: {", ".join(added_mappings[:10])}{"..." if len(added_mappings) > 10 else ""}',
                change_summary=f'Bulk uploaded {added_count} screen mappings',
                remarks=f'Bulk upload by {request.user.username}'
            )
        
        message = f'Successfully added {added_count} screen mappings.'
        if skipped_count > 0:
            message += f' Skipped {skipped_count} rows with errors.'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'added_count': added_count,
            'skipped_count': skipped_count,
            'added_mappings': added_mappings[:20],
            'errors': errors[:20]
        })
        
    except Exception as e:
        logger.error(f"Bulk upload error: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error processing file: {str(e)}'})


# ============================================================
# âœ… NEW: SCREEN MAPPING - DOWNLOAD BULK UPLOAD TEMPLATE
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def screen_mapping_download_template(request):
    """
    Download Excel template for bulk screen mapping upload
    Columns: ERP User ID, Screen Code
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Screen_Mapping_Upload_Template.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Screen Mapping"
    
    # Header
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    headers = ['ERP User ID', 'Screen Code']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Sample data
    data_font = Font(name='Calibri', size=11)
    samples = [
        ['HRD1223', 'SO-001'],
        ['HRD1223', 'PO-002'],
        ['FIN001', 'INV-003'],
        ['IT9876', 'USR-004'],
        ['MKT456', 'REP-005'],
    ]
    for row_idx, row in enumerate(samples, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
    
    # Notes
    note_row = len(samples) + 3
    note_cell = ws.cell(row=note_row, column=1)
    note_cell.value = "ðŸ“Œ Add your screen mappings below. One per row."
    note_cell.font = Font(name='Calibri', size=10, italic=True, color='FF0000')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)
    
    note_row2 = len(samples) + 4
    note_cell2 = ws.cell(row=note_row2, column=1)
    note_cell2.value = "âš ï¸ ERP User ID must exist in the system. Screen Code must exist in Screen Master."
    note_cell2.font = Font(name='Calibri', size=10, italic=True, color='FF6B00')
    ws.merge_cells(start_row=note_row2, start_column=1, end_row=note_row2, end_column=2)
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    wb.save(response)
    return response
