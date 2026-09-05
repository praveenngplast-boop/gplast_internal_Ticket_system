# tickets/views/settings_action/units.py

"""
Unit and Department Settings - Add, Edit, Toggle Active/Inactive, Bulk Upload Departments
"""
from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse
from django.db import IntegrityError, transaction
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
import logging

from tickets.models import Unit, Department
from tickets.forms import UnitForm, DepartmentForm
from .settings_audit import log_settings_change
from ..utils import is_admin

logger = logging.getLogger(__name__)


@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_units(request):
    """
    Manage Units: Add, Edit, Toggle Active/Inactive
    POST: action (add/edit/toggle), unit_id, code, full_name
    Redirects to: settings_units_departments
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            form = UnitForm(request.POST)
            if form.is_valid(): 
                unit = form.save(commit=False)
                unit.created_by = request.user.username
                unit.save()
                messages.success(request, f"Unit '{unit.code}' added.")
                
                log_settings_change(
                    request,
                    action_type='CREATE',
                    setting_type='UNIT',
                    setting_name=f"Unit: {unit.code}",
                    new_value=f"Code: {unit.code}, Name: {unit.full_name}",
                    change_summary=f"Added unit '{unit.code}' - {unit.full_name}",
                    remarks=f"Unit added by {request.user.username}"
                )
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        
        elif action == 'edit':
            unit = get_object_or_404(Unit, pk=request.POST.get('unit_id'))
            old_code = unit.code
            old_name = unit.full_name
            
            form = UnitForm(request.POST, instance=unit)
            if form.is_valid(): 
                form.save()
                messages.success(request, f"Unit '{unit.code}' updated.")
                
                change_details = []
                if old_code != unit.code:
                    change_details.append(f"Code: {old_code} → {unit.code}")
                if old_name != unit.full_name:
                    change_details.append(f"Name: {old_name} → {unit.full_name}")
                
                log_settings_change(
                    request,
                    action_type='UPDATE',
                    setting_type='UNIT',
                    setting_name=f"Unit: {unit.code}",
                    old_value=f"Code: {old_code}, Name: {old_name}",
                    new_value=f"Code: {unit.code}, Name: {unit.full_name}",
                    change_summary='; '.join(change_details) if change_details else 'Unit updated',
                    remarks=f"Unit updated by {request.user.username}"
                )
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        
        elif action == 'toggle':
            unit = get_object_or_404(Unit, pk=request.POST.get('unit_id'))
            old_status = 'Active' if unit.is_active else 'Inactive'
            unit.is_active = not unit.is_active
            unit.save()
            new_status = 'Active' if unit.is_active else 'Inactive'
            
            messages.success(request, f"Unit '{unit.code}' {'activated' if unit.is_active else 'deactivated'}.")
            
            log_settings_change(
                request,
                action_type='TOGGLE',
                setting_type='UNIT',
                setting_name=f"Unit: {unit.code}",
                old_value=f"Status: {old_status}",
                new_value=f"Status: {new_status}",
                change_summary=f"Status changed from {old_status} to {new_status}",
                remarks=f"Unit toggled by {request.user.username}"
            )
    
    return redirect('settings_units_departments')


@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_departments(request):
    """
    Manage Departments: Add, Edit, Toggle Active/Inactive
    POST: action (add/edit/toggle), dept_id, name, unit
    Redirects to: settings_units_departments
    ✅ FIXED: Allows same department name in different units
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            unit_id = request.POST.get('unit')
            dept_name = request.POST.get('name', '').strip()
            
            if not unit_id or not dept_name:
                messages.error(request, "Unit and Department Name are required.")
                return redirect('settings_units_departments')
            
            # ✅ FIXED: Check per unit, NOT globally
            unit = get_object_or_404(Unit, pk=unit_id)
            
            if Department.objects.filter(unit=unit, name__iexact=dept_name).exists():
                messages.error(request, f"Department '{dept_name}' already exists in unit '{unit.code}'.")
                return redirect('settings_units_departments')
            
            try:
                dept = Department.objects.create(
                    unit=unit,
                    name=dept_name,
                    is_active=True
                )
                messages.success(request, f"Department '{dept.name}' added under {unit.code}.")
                
                log_settings_change(
                    request,
                    action_type='CREATE',
                    setting_type='DEPARTMENT',
                    setting_name=f"Department: {dept.name}",
                    new_value=f"Name: {dept.name}, Unit: {dept.unit.code}",
                    change_summary=f"Added department '{dept.name}' under unit '{dept.unit.code}'",
                    remarks=f"Department added by {request.user.username}"
                )
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
        
        elif action == 'edit':
            dept_id = request.POST.get('dept_id')
            dept_name = request.POST.get('name', '').strip()
            
            if not dept_id or not dept_name:
                messages.error(request, "Department ID and Name are required.")
                return redirect('settings_units_departments')
            
            dept = get_object_or_404(Department, pk=dept_id)
            old_name = dept.name
            old_unit = dept.unit.code
            
            # ✅ FIXED: Check per unit, excluding current instance
            if Department.objects.filter(
                unit=dept.unit, 
                name__iexact=dept_name
            ).exclude(pk=dept_id).exists():
                messages.error(request, f"Department '{dept_name}' already exists in unit '{dept.unit.code}'.")
                return redirect('settings_units_departments')
            
            dept.name = dept_name
            dept.save()
            messages.success(request, f"Department '{dept.name}' updated.")
            
            change_details = []
            if old_name != dept.name:
                change_details.append(f"Name: {old_name} → {dept.name}")
            if old_unit != dept.unit.code:
                change_details.append(f"Unit: {old_unit} → {dept.unit.code}")
            
            log_settings_change(
                request,
                action_type='UPDATE',
                setting_type='DEPARTMENT',
                setting_name=f"Department: {dept.name}",
                old_value=f"Name: {old_name}, Unit: {old_unit}",
                new_value=f"Name: {dept.name}, Unit: {dept.unit.code}",
                change_summary='; '.join(change_details) if change_details else 'Department updated',
                remarks=f"Department updated by {request.user.username}"
            )
        
        elif action == 'toggle':
            dept = get_object_or_404(Department, pk=request.POST.get('dept_id'))
            old_status = 'Active' if dept.is_active else 'Inactive'
            dept.is_active = not dept.is_active
            dept.save()
            new_status = 'Active' if dept.is_active else 'Inactive'
            
            messages.success(request, f"Department '{dept.name}' {'activated' if dept.is_active else 'deactivated'}.")
            
            log_settings_change(
                request,
                action_type='TOGGLE',
                setting_type='DEPARTMENT',
                setting_name=f"Department: {dept.name}",
                old_value=f"Status: {old_status}",
                new_value=f"Status: {new_status}",
                change_summary=f"Status changed from {old_status} to {new_status}",
                remarks=f"Department toggled by {request.user.username}"
            )
    
    return redirect('settings_units_departments')


# ============================================================
# ✅ BULK UPLOAD DEPARTMENTS - FIXED WITH PROPER ERROR HANDLING
# ============================================================
@csrf_exempt
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def departments_bulk_upload(request):
    """
    Bulk upload departments from Excel/CSV file
    Expected columns: 'Unit Code', 'Department Name'
    Maps department to existing unit based on Unit Code
    URL: /custom-admin/settings/departments/bulk-upload/
    ✅ FIXED: Allows same department name in different units
    """
    # Check if it's an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if not is_ajax:
        return JsonResponse({
            'success': False,
            'message': 'Invalid request. This endpoint only accepts AJAX requests.',
            'added_count': 0,
            'skipped_count': 0,
            'errors': ['Invalid request type.']
        }, status=400)
    
    try:
        if request.method != 'POST':
            return JsonResponse({
                'success': False,
                'message': 'Invalid method. Use POST.',
                'added_count': 0,
                'skipped_count': 0,
                'errors': ['Invalid request method. Please use POST.']
            }, status=400)
        
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return JsonResponse({
                'success': False,
                'message': 'Please select an Excel file.',
                'added_count': 0,
                'skipped_count': 0,
                'errors': ['No file selected.']
            })
        
        if not excel_file.name.endswith(('.xlsx', '.xls', '.csv')):
            return JsonResponse({
                'success': False,
                'message': 'Invalid file format. Only .xlsx, .xls, and .csv files are supported.',
                'added_count': 0,
                'skipped_count': 0,
                'errors': ['Unsupported file format.']
            })
        
        # Read the file
        try:
            if excel_file.name.endswith('.csv'):
                df = pd.read_csv(excel_file, dtype=str)
            else:
                df = pd.read_excel(excel_file, dtype=str)
        except Exception as e:
            logger.error(f"Error reading file: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'Error reading file: {str(e)}',
                'added_count': 0,
                'skipped_count': 0,
                'errors': [f'Could not read file: {str(e)}']
            })
        
        if df.empty:
            return JsonResponse({
                'success': False,
                'message': 'The uploaded file is empty.',
                'added_count': 0,
                'skipped_count': 0,
                'errors': ['File is empty.']
            })
        
        # Find columns - look for Unit Code and Department Name
        unit_code_column = None
        dept_name_column = None
        
        for col in df.columns:
            col_lower = col.strip().lower()
            if 'unit' in col_lower and ('code' in col_lower or 'id' in col_lower):
                unit_code_column = col
            if 'department' in col_lower or 'dept' in col_lower:
                if 'name' in col_lower or 'dept' in col_lower:
                    dept_name_column = col
        
        # If not found by pattern, use first two columns
        if unit_code_column is None and len(df.columns) >= 1:
            unit_code_column = df.columns[0]
        if dept_name_column is None and len(df.columns) >= 2:
            dept_name_column = df.columns[1]
        
        # If still not found, use column names based on position
        if unit_code_column is None:
            unit_code_column = df.columns[0]
        if dept_name_column is None:
            dept_name_column = df.columns[1] if len(df.columns) > 1 else df.columns[0]
        
        # Get all active units for validation
        existing_units = {unit.code: unit for unit in Unit.objects.filter(is_active=True)}
        
        added_count = 0
        skipped_count = 0
        errors = []
        added_depts = []
        skipped_depts = []
        
        with transaction.atomic():
            for index, row in df.iterrows():
                # Get values
                unit_code = str(row[unit_code_column]).strip() if pd.notna(row[unit_code_column]) else ''
                dept_name = str(row[dept_name_column]).strip() if pd.notna(row[dept_name_column]) else ''
                
                # Validate
                if not unit_code:
                    errors.append(f"Row {index + 1}: Unit Code is empty")
                    skipped_count += 1
                    continue
                
                if not dept_name:
                    errors.append(f"Row {index + 1}: Department Name is empty")
                    skipped_count += 1
                    continue
                
                # Check if unit exists
                unit = existing_units.get(unit_code)
                if not unit:
                    errors.append(f"Row {index + 1}: Unit '{unit_code}' not found")
                    skipped_count += 1
                    continue
                
                # ✅ FIXED: Check if department exists in THIS unit only
                if Department.objects.filter(unit=unit, name__iexact=dept_name).exists():
                    errors.append(f"Row {index + 1}: Department '{dept_name}' already exists in unit '{unit_code}'")
                    skipped_count += 1
                    skipped_depts.append(dept_name)
                    continue
                
                # Create department
                try:
                    dept = Department.objects.create(
                        unit=unit,
                        name=dept_name,
                        is_active=True
                    )
                    added_count += 1
                    added_depts.append(dept_name)
                except IntegrityError:
                    errors.append(f"Row {index + 1}: Database constraint error for {unit_code} - {dept_name}")
                    skipped_count += 1
                    skipped_depts.append(dept_name)
                except Exception as e:
                    errors.append(f"Row {index + 1}: Error creating department '{dept_name}': {str(e)}")
                    skipped_count += 1
        
        # Log bulk upload
        if added_count > 0:
            log_settings_change(
                request,
                action_type='CREATE',
                setting_type='DEPARTMENT',
                setting_name=f'Bulk Upload: {added_count} Departments',
                new_value=f'Added {added_count} departments',
                change_summary=f'Bulk uploaded {added_count} departments',
                remarks=f'Bulk uploaded {added_count} departments, {skipped_count} skipped by {request.user.username}'
            )
        
        message = f'Successfully added {added_count} departments.'
        if skipped_count > 0:
            message += f' Skipped {skipped_count} entries.'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'added_count': added_count,
            'skipped_count': skipped_count,
            'added_depts': added_depts[:20],
            'skipped_depts': skipped_depts[:20],
            'errors': errors[:20]
        })
        
    except Exception as e:
        logger.error(f"Bulk upload departments error: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error processing file: {str(e)}',
            'added_count': 0,
            'skipped_count': 0,
            'errors': [str(e)]
        }, status=500)


# ============================================================
# ✅ DOWNLOAD BULK UPLOAD TEMPLATE FOR DEPARTMENTS
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def departments_download_template(request):
    """
    Download Excel template for bulk department upload
    URL: /custom-admin/settings/departments/download-template/
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Department_Upload_Template.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Departments"
    
    # Styles
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_font = Font(name='Calibri', size=11)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Headers
    headers = ['Unit Code', 'Department Name']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Sample data with existing units
    existing_units = Unit.objects.filter(is_active=True).order_by('code')
    sample_data = []
    
    if existing_units.exists():
        for unit in existing_units[:5]:  # Show up to 5 sample units
            sample_data.append([unit.code, f'{unit.code} Department'])
    
    # If no units exist or less than 3 samples, add generic samples
    if len(sample_data) < 3:
        generic_samples = [
            ['HR', 'Human Resources'],
            ['IT', 'Information Technology'],
            ['FIN', 'Finance'],
        ]
        for idx, (code, name) in enumerate(generic_samples):
            if idx >= len(sample_data):
                sample_data.append([code, name])
    
    for row_idx, (unit_code, dept_name) in enumerate(sample_data, 2):
        for col_idx, value in enumerate([unit_code, dept_name], 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
    
    # Notes
    note_row = len(sample_data) + 4
    
    # Note 1 - Instructions
    note_cell1 = ws.cell(row=note_row, column=1)
    note_cell1.value = "📌 INSTRUCTIONS:"
    note_cell1.font = Font(name='Calibri', size=10, bold=True, color='1F4E79')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)
    
    note_row += 1
    note_cell2 = ws.cell(row=note_row, column=1)
    note_cell2.value = "1. Enter existing Unit Code (must match an active unit in the system)"
    note_cell2.font = Font(name='Calibri', size=9, color='333333')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)
    
    note_row += 1
    note_cell3 = ws.cell(row=note_row, column=1)
    note_cell3.value = "2. Enter Department Name (will be created under the specified unit)"
    note_cell3.font = Font(name='Calibri', size=9, color='333333')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)
    
    note_row += 1
    note_cell4 = ws.cell(row=note_row, column=1)
    note_cell4.value = "3. Departments with duplicate names within the same unit will be skipped automatically"
    note_cell4.font = Font(name='Calibri', size=9, color='333333')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)
    
    note_row += 1
    note_cell5 = ws.cell(row=note_row, column=1)
    note_cell5.value = "4. The same department name can exist in different units (e.g., HRD in DCD and IMD)"
    note_cell5.font = Font(name='Calibri', size=9, color='333333')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)
    
    note_row += 2
    note_cell6 = ws.cell(row=note_row, column=1)
    note_cell6.value = "⚠️ Unit Code must exist in the system. Departments are created as ACTIVE by default."
    note_cell6.font = Font(name='Calibri', size=9, italic=True, color='FF6B00')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)
    
    # Add active units list for reference
    note_row += 2
    note_cell7 = ws.cell(row=note_row, column=1)
    note_cell7.value = "📋 Active Units in System:"
    note_cell7.font = Font(name='Calibri', size=9, bold=True, color='1F4E79')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)
    
    note_row += 1
    unit_codes = ', '.join([unit.code for unit in existing_units]) if existing_units.exists() else 'No units found'
    note_cell8 = ws.cell(row=note_row, column=1)
    note_cell8.value = unit_codes
    note_cell8.font = Font(name='Calibri', size=9, color='666666')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    
    wb.save(response)
    return response