# tickets/views/settings_action/erp_mapping_views.py

"""
ERP User ID Mapping Views - Add, Remove, Bulk Upload, Export Excel, Unmap
âœ… FIXED: Grouped view to show employees with same ERP ID mappings
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.utils import timezone
from django.db.models import Count, Q
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging

from tickets.models import ERPHolderMapping, EmployeeMaster, SettingsAuditLog
from .settings_audit import log_settings_change
from ..utils import is_admin

logger = logging.getLogger(__name__)


# ============================================================
# ERP MAPPING - PAGE VIEW (GROUPED BY ERP ID)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def erp_mapping_page(request):
    """
    ERP User ID Mapping page - Grouped by ERP ID
    URL: /custom-admin/settings/erp-mapping/
    Shows which employees share the same ERP ID
    """
    mappings = ERPHolderMapping.objects.all().select_related('employee', 'employee__unit', 'employee__department').order_by('erp_user_id')
    
    employees = EmployeeMaster.objects.filter(is_active=True).order_by('employee_id')
    
    # ============================================================
    # GROUP MAPPINGS BY ERP USER ID
    # ============================================================
    grouped_mappings = {}
    for mapping in mappings:
        erp_id = mapping.erp_user_id
        if erp_id not in grouped_mappings:
            grouped_mappings[erp_id] = {
                'erp_user_id': erp_id,
                'employees': [],
                'count': 0,
                'is_mapped': False,
                'mapping_ids': [],  # Store mapping IDs for bulk operations
            }
        
        if mapping.employee:
            grouped_mappings[erp_id]['employees'].append({
                'mapping_id': mapping.id,
                'employee_id': mapping.employee.employee_id,
                'employee_name': mapping.employee.employee_name,
                'unit_code': mapping.employee.unit.code if mapping.employee.unit else 'â€”',
                'department_name': mapping.employee.department.name if mapping.employee.department else 'â€”',
                'is_mapped': True,
            })
            grouped_mappings[erp_id]['is_mapped'] = True
        else:
            # Unmapped entry (no employee)
            grouped_mappings[erp_id]['employees'].append({
                'mapping_id': mapping.id,
                'employee_id': None,
                'employee_name': 'Not Mapped',
                'unit_code': 'â€”',
                'department_name': 'â€”',
                'is_mapped': False,
            })
        
        grouped_mappings[erp_id]['mapping_ids'].append(mapping.id)
        grouped_mappings[erp_id]['count'] += 1
    
    # Convert to list for template
    grouped_list = list(grouped_mappings.values())
    
    # Sort: Mapped first, then unmapped, then by ERP ID
    grouped_list.sort(key=lambda x: (not x['is_mapped'], x['erp_user_id']))
    
    # Paginate grouped list
    paginator = Paginator(grouped_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    total_mappings = mappings.count()
    mapped_count = mappings.filter(employee__isnull=False).count()
    unmapped_count = mappings.filter(employee__isnull=True).count()
    unique_erp_ids = len(grouped_mappings)
    
    context = {
        'page_obj': page_obj,
        'employees': employees,
        'total_mappings': total_mappings,
        'mapped_count': mapped_count,
        'unmapped_count': unmapped_count,
        'unique_erp_ids': unique_erp_ids,
        'grouped_mappings': page_obj,  # Pass paginated grouped data
    }
    return render(request, 'admin_panel/erp_mapping.html', context)


# ============================================================
# ERP MAPPING - ADD (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def erp_mapping_add(request):
    """
    Add a single ERP User ID mapping (AJAX)
    URL: /custom-admin/settings/erp-mapping/add/
    âœ… FIXED: Allows same ERP ID to be mapped to multiple employees across different units
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=400)
    
    erp_user_id = request.POST.get('erp_user_id', '').strip()
    employee_id = request.POST.get('employee_id', '').strip()
    action = request.POST.get('action', 'add')
    
    if not erp_user_id:
        return JsonResponse({'success': False, 'message': 'ERP User ID is required'})
    
    try:
        with transaction.atomic():
            
            # ============================================================
            # ACTION: MAP ERP ID TO EMPLOYEE
            # ============================================================
            if action == 'map':
                if not employee_id:
                    return JsonResponse({'success': False, 'message': 'Employee ID is required for mapping'})
                
                # Get employee
                try:
                    employee = EmployeeMaster.objects.get(employee_id=employee_id, is_active=True)
                except EmployeeMaster.DoesNotExist:
                    return JsonResponse({'success': False, 'message': f'Employee ID "{employee_id}" not found'})
                
                # âœ… Check if this exact ERP-Employee combination already exists
                existing_mapping = ERPHolderMapping.objects.filter(
                    erp_user_id=erp_user_id,
                    employee=employee
                ).first()
                
                if existing_mapping:
                    return JsonResponse({
                        'success': False,
                        'message': f'ERP {erp_user_id} is already mapped to {employee_id} ({employee.employee_name})'
                    })
                
                # âœ… Create new mapping (allow same ERP ID for different employees)
                mapping = ERPHolderMapping.objects.create(
                    erp_user_id=erp_user_id,
                    employee=employee,
                    is_mapped=True,
                    created_by=request.user.username
                )
                
                log_settings_change(
                    request,
                    action_type='CREATE',
                    setting_type='ERP_MAPPING',
                    setting_name=f"ERP {erp_user_id} â†’ {employee_id}",
                    new_value=f"ERP: {erp_user_id}, Employee: {employee_id} - {employee.employee_name}",
                    change_summary=f"Mapped ERP {erp_user_id} to {employee_id}",
                    remarks=f"Mapped by {request.user.username}"
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'ERP {erp_user_id} mapped to {employee.employee_name} successfully!',
                    'mapping': {
                        'id': mapping.id,
                        'erp_user_id': mapping.erp_user_id,
                        'employee_id': employee.employee_id,
                        'employee_name': employee.employee_name,
                        'unit_code': employee.unit.code if employee.unit else 'â€”',
                        'department_name': employee.department.name if employee.department else 'â€”',
                        'is_mapped': True,
                    }
                })
            
            # ============================================================
            # ACTION: ADD NEW ERP ID (without employee)
            # ============================================================
            else:
                # âœ… Allow adding duplicate ERP IDs (since unique=True is removed)
                # Just create a new unmapped entry
                mapping = ERPHolderMapping.objects.create(
                    erp_user_id=erp_user_id,
                    employee=None,
                    is_mapped=False,
                    created_by=request.user.username
                )
                
                log_settings_change(
                    request,
                    action_type='CREATE',
                    setting_type='ERP_MAPPING',
                    setting_name=f"ERP {erp_user_id}",
                    new_value=f"ERP: {erp_user_id} (Not Mapped)",
                    change_summary=f"Added ERP ID: {erp_user_id}",
                    remarks=f"Added by {request.user.username}"
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'ERP ID "{erp_user_id}" added successfully!',
                    'mapping': {
                        'id': mapping.id,
                        'erp_user_id': mapping.erp_user_id,
                        'employee_id': None,
                        'employee_name': 'Not Mapped',
                        'unit_code': 'â€”',
                        'department_name': 'â€”',
                        'is_mapped': False,
                    }
                })
                
    except IntegrityError as e:
        return JsonResponse({'success': False, 'message': f'Database error: {str(e)}'})
    except Exception as e:
        logger.error(f'Error adding ERP mapping: {e}')
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================================
# ERP MAPPING - REMOVE (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def erp_mapping_remove(request):
    """
    Remove an ERP User ID mapping (AJAX)
    URL: /custom-admin/settings/erp-mapping/remove/
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=400)
    
    mapping_id = request.POST.get('mapping_id', '').strip()
    
    if not mapping_id:
        return JsonResponse({'success': False, 'message': 'Mapping ID is required'})
    
    try:
        mapping = ERPHolderMapping.objects.get(id=mapping_id)
        erp_id = mapping.erp_user_id
        employee_id = mapping.employee.employee_id if mapping.employee else 'Not Mapped'
        employee_name = mapping.employee.employee_name if mapping.employee else 'Not Mapped'
        
        mapping.delete()
        
        log_settings_change(
            request,
            action_type='DELETE',
            setting_type='ERP_MAPPING',
            setting_name=f"ERP {erp_id}",
            old_value=f"ERP: {erp_id}, Employee: {employee_id} - {employee_name}",
            change_summary=f"Removed ERP ID: {erp_id}",
            remarks=f"Removed by {request.user.username}"
        )
        
        return JsonResponse({
            'success': True,
            'message': f'ERP ID "{erp_id}" removed successfully!'
        })
    except ERPHolderMapping.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Mapping not found'}, status=404)
    except Exception as e:
        logger.error(f'Error removing ERP mapping: {e}')
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================================
# ERP MAPPING - UNMAP (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def erp_mapping_unmap(request):
    """
    Remove employee mapping from ERP ID (keep ERP ID) (AJAX)
    URL: /custom-admin/settings/erp-mapping/unmap/
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=400)
    
    mapping_id = request.POST.get('mapping_id', '').strip()
    
    if not mapping_id:
        return JsonResponse({'success': False, 'message': 'Mapping ID is required'})
    
    try:
        mapping = ERPHolderMapping.objects.get(id=mapping_id)
    except ERPHolderMapping.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'ERP ID not found'})
    
    if not mapping.employee:
        return JsonResponse({'success': False, 'message': 'ERP ID is already unmapped'})
    
    erp_user_id = mapping.erp_user_id
    employee_id = mapping.employee.employee_id
    employee_name = mapping.employee.employee_name
    
    try:
        with transaction.atomic():
            mapping.employee = None
            mapping.is_mapped = False
            mapping.save()
            
            log_settings_change(
                request,
                action_type='UPDATE',
                setting_type='ERP_MAPPING',
                setting_name=f"ERP {erp_user_id}",
                old_value=f"ERP: {erp_user_id}, Employee: {employee_id} - {employee_name}",
                new_value=f"ERP: {erp_user_id} (Not Mapped)",
                change_summary=f"Unmapped ERP {erp_user_id} from {employee_id}",
                remarks=f"Unmapped by {request.user.username}"
            )
            
            return JsonResponse({
                'success': True,
                'message': f'ERP ID "{erp_user_id}" unmapped from {employee_id} ({employee_name}) successfully!'
            })
    except Exception as e:
        logger.error(f'Error unmapping ERP: {e}')
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================================
# ERP MAPPING - UNMAP ALL MAPPINGS FOR ERP ID (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def erp_mapping_unmap_all(request):
    """
    Remove all employee mappings from an ERP ID (keep ERP ID entries)
    URL: /custom-admin/settings/erp-mapping/unmap-all/
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=400)
    
    erp_user_id = request.POST.get('erp_user_id', '').strip()
    
    if not erp_user_id:
        return JsonResponse({'success': False, 'message': 'ERP User ID is required'})
    
    try:
        mappings = ERPHolderMapping.objects.filter(erp_user_id=erp_user_id)
        
        if not mappings.exists():
            return JsonResponse({'success': False, 'message': f'ERP ID "{erp_user_id}" not found'})
        
        updated_count = 0
        with transaction.atomic():
            for mapping in mappings:
                if mapping.employee:
                    mapping.employee = None
                    mapping.is_mapped = False
                    mapping.save()
                    updated_count += 1
        
        log_settings_change(
            request,
            action_type='UPDATE',
            setting_type='ERP_MAPPING',
            setting_name=f"ERP {erp_user_id}",
            old_value=f"ERP: {erp_user_id} (Mapped to {updated_count} employees)",
            new_value=f"ERP: {erp_user_id} (Not Mapped)",
            change_summary=f"Unmapped all employees from ERP {erp_user_id}",
            remarks=f"Unmapped all by {request.user.username}"
        )
        
        return JsonResponse({
            'success': True,
            'message': f'All {updated_count} employee(s) unmapped from ERP ID "{erp_user_id}" successfully!'
        })
    except Exception as e:
        logger.error(f'Error unmapping all ERP: {e}')
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================================
# ERP MAPPING - DELETE ALL MAPPINGS FOR ERP ID (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def erp_mapping_delete_all(request):
    """
    Delete all mappings for an ERP ID
    URL: /custom-admin/settings/erp-mapping/delete-all/
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=400)
    
    erp_user_id = request.POST.get('erp_user_id', '').strip()
    
    if not erp_user_id:
        return JsonResponse({'success': False, 'message': 'ERP User ID is required'})
    
    try:
        mappings = ERPHolderMapping.objects.filter(erp_user_id=erp_user_id)
        
        if not mappings.exists():
            return JsonResponse({'success': False, 'message': f'ERP ID "{erp_user_id}" not found'})
        
        count = mappings.count()
        mappings.delete()
        
        log_settings_change(
            request,
            action_type='DELETE',
            setting_type='ERP_MAPPING',
            setting_name=f"ERP {erp_user_id}",
            old_value=f"ERP: {erp_user_id} ({count} mappings)",
            change_summary=f"Deleted all {count} mappings for ERP {erp_user_id}",
            remarks=f"Deleted all by {request.user.username}"
        )
        
        return JsonResponse({
            'success': True,
            'message': f'All {count} mapping(s) for ERP ID "{erp_user_id}" deleted successfully!'
        })
    except Exception as e:
        logger.error(f'Error deleting all ERP mappings: {e}')
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================================
# ERP MAPPING - EXPORT EXCEL
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def erp_mapping_export_excel(request):
    """
    Export all ERP mappings to Excel
    URL: /custom-admin/settings/erp-mapping/export-excel/
    """
    mappings = ERPHolderMapping.objects.all().select_related('employee', 'employee__unit', 'employee__department').order_by('erp_user_id')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ERP_Mappings_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ERP Mappings"
    
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10)
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    current_tz = timezone.get_current_timezone()
    now_local = timezone.now().astimezone(current_tz)
    report_time = now_local.strftime('%d-%b-%Y %I:%M:%S %p')
    
    ws.merge_cells('A1:G1')
    ws['A1'] = f"ERP USER ID MAPPINGS - Generated: {report_time}  |  Total: {mappings.count()}"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    headers = ['#', 'ERP User ID', 'Employee ID', 'Employee Name', 'Unit', 'Department', 'Status']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[3].height = 30
    
    for row_idx, mapping in enumerate(mappings, 4):
        if mapping.employee:
            row_data = [
                row_idx - 3,
                mapping.erp_user_id,
                mapping.employee.employee_id,
                mapping.employee.employee_name,
                mapping.employee.unit.code if mapping.employee.unit else 'â€”',
                mapping.employee.department.name if mapping.employee.department else 'â€”',
                'Mapped'
            ]
        else:
            row_data = [
                row_idx - 3,
                mapping.erp_user_id,
                'â€”',
                'Not Mapped',
                'â€”',
                'â€”',
                'Unmapped'
            ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
            
            if col_idx == 7:
                if value == 'Mapped':
                    cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
                    cell.font = Font(name='Calibri', size=10, bold=True, color='22C55E')
                else:
                    cell.fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
                    cell.font = Font(name='Calibri', size=10, bold=True, color='F59E0B')
    
    column_widths = {'A': 8, 'B': 18, 'C': 18, 'D': 30, 'E': 20, 'F': 25, 'G': 15}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    wb.save(response)
    return response


# ============================================================
# ERP MAPPING - BULK UPLOAD (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def erp_mapping_bulk_upload(request):
    """
    Bulk upload ONLY ERP IDs from Excel/CSV - NO format validation
    Creates ERP IDs WITHOUT employee mapping
    URL: /custom-admin/settings/erp-mapping/bulk-upload/
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=400)
    
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return JsonResponse({'success': False, 'message': 'Please select an Excel file.'})
    
    if not excel_file.name.endswith(('.xlsx', '.xls', '.csv')):
        return JsonResponse({'success': False, 'message': 'Invalid file format. Only .xlsx, .xls, and .csv files are supported.'})
    
    try:
        if excel_file.name.endswith('.csv'):
            df = pd.read_csv(excel_file, dtype=str)
        else:
            df = pd.read_excel(excel_file, dtype=str)
        
        if df.empty:
            return JsonResponse({'success': False, 'message': 'The uploaded file is empty.'})
        
        # Find ERP column
        erp_column = None
        for col in df.columns:
            col_lower = col.strip().lower()
            if 'erp' in col_lower or 'user' in col_lower or 'id' in col_lower:
                erp_column = col
                break
        
        if erp_column is None:
            erp_column = df.columns[0]
        
        erp_ids = []
        for value in df[erp_column]:
            if pd.notna(value):
                erp_id = str(value).strip()
                if erp_id:
                    erp_ids.append(erp_id)
        
        if not erp_ids:
            return JsonResponse({'success': False, 'message': 'No valid ERP IDs found in the file.'})
        
        # âœ… Allow duplicate ERP IDs - no uniqueness check
        added_count = 0
        errors = []
        added_erp_ids = []
        
        for erp_id in erp_ids:
            try:
                mapping = ERPHolderMapping.objects.create(
                    erp_user_id=erp_id,
                    employee=None,
                    is_mapped=False,
                    created_by=request.user.username
                )
                added_erp_ids.append(erp_id)
                added_count += 1
            except Exception as e:
                errors.append(f'Error adding ERP ID "{erp_id}": {str(e)}')
        
        if added_count > 0:
            log_settings_change(
                request,
                action_type='CREATE',
                setting_type='ERP_MAPPING',
                setting_name=f'Bulk Upload: {added_count} ERP IDs',
                new_value=f'Added {added_count} ERP IDs',
                change_summary=f'Bulk uploaded {added_count} ERP IDs',
                remarks=f'Bulk uploaded {added_count} ERP IDs by {request.user.username}'
            )
        
        message = f'Successfully added {added_count} ERP IDs.'
        if errors:
            message += f' {len(errors)} entries had errors.'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'added_count': added_count,
            'skipped_count': len(errors),
            'added_erp_ids': added_erp_ids[:20],
            'errors': errors[:20]
        })
        
    except Exception as e:
        logger.error(f"Bulk upload error: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error processing file: {str(e)}'})


# ============================================================
# ERP MAPPING - DOWNLOAD BULK UPLOAD TEMPLATE
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def erp_mapping_download_template(request):
    """
    Download Excel template for bulk ERP ID upload
    URL: /custom-admin/settings/erp-mapping/download-template/
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=ERP_Upload_Template.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ERP IDs"
    
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    cell = ws.cell(row=1, column=1)
    cell.value = "ERP User ID"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    
    data_font = Font(name='Calibri', size=11)
    samples = ['HRD1223', 'FIN001', 'IT9876', 'MKT456', 'OPS789']
    for row_idx, sample in enumerate(samples, 2):
        cell = ws.cell(row=row_idx, column=1)
        cell.value = sample
        cell.font = data_font
    
    note_row = len(samples) + 3
    note_cell = ws.cell(row=note_row, column=1)
    note_cell.value = "Add your ERP IDs below. One per row. Any format accepted."
    note_cell.font = Font(name='Calibri', size=10, italic=True, color='FF0000')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=1)
    
    note_row2 = len(samples) + 4
    note_cell2 = ws.cell(row=note_row2, column=1)
    note_cell2.value = "Same ERP ID can be used by multiple employees across different units."
    note_cell2.font = Font(name='Calibri', size=10, italic=True, color='FF6B00')
    ws.merge_cells(start_row=note_row2, start_column=1, end_row=note_row2, end_column=1)
    
    ws.column_dimensions['A'].width = 25
    
    wb.save(response)
    return response


# ============================================================
# ERP MAPPING - LIST (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def erp_mapping_list(request):
    """
    Get ERP mappings list (AJAX)
    URL: /ajax/get-erp-mappings/
    """
    mappings = ERPHolderMapping.objects.all().select_related('employee', 'employee__unit', 'employee__department').order_by('erp_user_id')
    
    data = []
    for mapping in mappings:
        data.append({
            'id': mapping.id,
            'erp_user_id': mapping.erp_user_id,
            'employee_id': mapping.employee.employee_id if mapping.employee else None,
            'employee_name': mapping.employee.employee_name if mapping.employee else 'Not Mapped',
            'unit_code': mapping.employee.unit.code if mapping.employee and mapping.employee.unit else '',
            'department_name': mapping.employee.department.name if mapping.employee and mapping.employee.department else '',
            'is_mapped': mapping.is_mapped,
        })
    
    return JsonResponse({
        'success': True,
        'mappings': data,
        'count': len(data)
    })


# ============================================================
# ERP MAPPING - SEARCH EMPLOYEES (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='login')
def erp_mapping_search_employees(request):
    """
    Search employees by ID or Name (AJAX)
    URL: /ajax/search-employees/
    """
    from django.db import models
    
    search = request.GET.get('search', '').strip()
    
    employees = EmployeeMaster.objects.filter(is_active=True)
    if search:
        employees = employees.filter(
            models.Q(employee_id__icontains=search) |
            models.Q(employee_name__icontains=search)
        )
    
    employees = employees.order_by('employee_id')[:20]
    
    data = []
    for emp in employees:
        data.append({
            'employee_id': emp.employee_id,
            'employee_name': emp.employee_name,
            'unit_code': emp.unit.code if emp.unit else '',
            'department_name': emp.department.name if emp.department else '',
        })
    
    return JsonResponse({
        'success': True,
        'employees': data
    })
