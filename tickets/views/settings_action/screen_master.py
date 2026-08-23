# tickets/views/settings_actions/screen_master.py

"""
Screen Master - Add, Edit, Delete, Download Excel, Download Template, Bulk Upload
"""
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging

from tickets.models import ScreenMaster
from .settings_audit import log_settings_change
from ..utils import is_admin

logger = logging.getLogger(__name__)


# ============================================================
# SCREEN MASTER - ADD
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def screen_master_add(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=400)

    screen_name = request.POST.get('screen_name', '').strip()
    screen_code = request.POST.get('screen_code', '').strip().upper()
    screen_type = request.POST.get('screen_type', 'ALL').strip()

    if not screen_name:
        return JsonResponse({'success': False, 'message': 'Screen Name is required'})
    if not screen_code:
        return JsonResponse({'success': False, 'message': 'Screen Code is required'})

    # Duplicate check
    if ScreenMaster.objects.filter(screen_name__iexact=screen_name).exists():
        return JsonResponse({'success': False, 'message': f'Screen Name "{screen_name}" already exists'})
    if ScreenMaster.objects.filter(screen_code__iexact=screen_code).exists():
        return JsonResponse({'success': False, 'message': f'Screen Code "{screen_code}" already exists'})

    try:
        screen = ScreenMaster.objects.create(
            screen_name=screen_name,
            screen_code=screen_code,
            screen_type=screen_type,
            created_by=request.user.username
        )
        log_settings_change(
            request, 'CREATE', 'SCREEN',
            f'{screen_code} - {screen_name}',
            new_value=f'Code: {screen_code}, Name: {screen_name}, Type: {screen_type}',
            change_summary=f'Added screen: {screen_code} - {screen_name}'
        )
        return JsonResponse({
            'success': True,
            'message': f'Screen "{screen_name}" added successfully',
            'screen': {
                'id': screen.id,
                'screen_name': screen.screen_name,
                'screen_code': screen.screen_code,
                'screen_type': screen.get_screen_type_display()
            }
        })
    except Exception as e:
        logger.error(f'Error adding screen: {e}')
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================================
# SCREEN MASTER - EDIT
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def screen_master_edit(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=400)

    screen_id = request.POST.get('screen_id', '').strip()
    screen_name = request.POST.get('screen_name', '').strip()
    screen_code = request.POST.get('screen_code', '').strip().upper()
    screen_type = request.POST.get('screen_type', 'ALL').strip()

    if not screen_id:
        return JsonResponse({'success': False, 'message': 'Screen ID is required'})
    if not screen_name:
        return JsonResponse({'success': False, 'message': 'Screen Name is required'})
    if not screen_code:
        return JsonResponse({'success': False, 'message': 'Screen Code is required'})

    try:
        screen = ScreenMaster.objects.get(id=screen_id)
    except ScreenMaster.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Screen not found'})

    # Duplicate check (exclude self)
    if ScreenMaster.objects.filter(screen_name__iexact=screen_name).exclude(id=screen_id).exists():
        return JsonResponse({'success': False, 'message': f'Screen Name "{screen_name}" already exists'})
    if ScreenMaster.objects.filter(screen_code__iexact=screen_code).exclude(id=screen_id).exists():
        return JsonResponse({'success': False, 'message': f'Screen Code "{screen_code}" already exists'})

    old_value = f'Code: {screen.screen_code}, Name: {screen.screen_name}, Type: {screen.screen_type}'
    screen.screen_name = screen_name
    screen.screen_code = screen_code
    screen.screen_type = screen_type
    screen.save()

    log_settings_change(
        request, 'UPDATE', 'SCREEN',
        f'{screen_code} - {screen_name}',
        old_value=old_value,
        new_value=f'Code: {screen_code}, Name: {screen_name}, Type: {screen_type}',
        change_summary=f'Updated screen: {screen_code} - {screen_name}'
    )
    return JsonResponse({
        'success': True,
        'message': f'Screen updated successfully',
        'screen': {
            'id': screen.id,
            'screen_name': screen.screen_name,
            'screen_code': screen.screen_code,
            'screen_type': screen.get_screen_type_display()
        }
    })


# ============================================================
# SCREEN MASTER - DELETE
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def screen_master_delete(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=400)

    screen_id = request.POST.get('screen_id', '').strip()
    if not screen_id:
        return JsonResponse({'success': False, 'message': 'Screen ID is required'})

    try:
        screen = ScreenMaster.objects.get(id=screen_id)
    except ScreenMaster.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Screen not found'})

    old_value = f'Code: {screen.screen_code}, Name: {screen.screen_name}'
    mapping_count = screen.screen_mappings.count()
    screen_name = screen.screen_name
    screen_code = screen.screen_code
    screen.delete()

    log_settings_change(
        request, 'DELETE', 'SCREEN',
        f'{screen_code} - {screen_name}',
        old_value=old_value,
        change_summary=f'Deleted screen: {screen_code} - {screen_name} (had {mapping_count} mappings)'
    )
    return JsonResponse({'success': True, 'message': f'Screen "{screen_name}" deleted successfully'})


# ============================================================
# SCREEN MASTER - DOWNLOAD EXCEL
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def screen_master_download_excel(request):
    screens = ScreenMaster.objects.all().order_by('screen_name')

    current_tz = timezone.get_current_timezone()
    now_local = timezone.now().astimezone(current_tz)
    report_time = now_local.strftime('%d-%b-%Y %I:%M:%S %p')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Screen_Master_{now_local.strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Screen Master'

    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10)
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0')
    )

    ws.merge_cells('A1:F1')
    ws['A1'] = f'SCREEN MASTER LIST - Generated: {report_time}  |  Total: {screens.count()}'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    headers = ['#', 'Screen Code', 'Screen Name', 'Screen Type', 'Status', 'Created At']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[3].height = 25

    for ri, screen in enumerate(screens, 1):
        created_local = screen.created_at.astimezone(current_tz).strftime('%d-%b-%Y %I:%M %p') if screen.created_at else ''
        row_data = [
            ri,
            screen.screen_code,
            screen.screen_name,
            screen.get_screen_type_display(),
            'Active' if screen.is_active else 'Inactive',
            created_local
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri + 3, column=ci)
            cell.value = val
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border

    for col_letter, width in {'A': 8, 'B': 18, 'C': 40, 'D': 18, 'E': 12, 'F': 22}.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(response)
    return response


# ============================================================
# SCREEN MASTER - DOWNLOAD TEMPLATE
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def screen_master_download_template(request):
    """
    Download Excel template for bulk screen upload
    Columns: Screen Code, Screen Name, Screen Type
    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Screen_Upload_Template.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Screen Template'
    
    # Headers
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    headers = ['Screen Code', 'Screen Name', 'Screen Type']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Sample data
    samples = [
        ['SO-001', 'Sales Order Entry', 'ENTRY'],
        ['PO-001', 'Purchase Order Entry', 'ENTRY'],
        ['INV-001', 'Inventory Report', 'QUERY'],
        ['USR-001', 'User Management', 'CONFIGURATION'],
        ['GEN-001', 'General Dashboard', 'ALL'],
    ]
    for row_idx, row in enumerate(samples, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = Font(name='Calibri', size=11)
    
    # Notes
    note_row = len(samples) + 3
    note_cell = ws.cell(row=note_row, column=1)
    note_cell.value = "Mandatory: Screen Code, Screen Name, Screen Type (ALL/ENTRY/CONFIGURATION/QUERY)"
    note_cell.font = Font(name='Calibri', size=10, italic=True, color='FF0000')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)
    
    # Column widths
    for col in range(1, 4):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    
    wb.save(response)
    return response


# ============================================================
# SCREEN MASTER - BULK UPLOAD
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def screen_master_bulk_upload(request):
    """
    Bulk upload screens from Excel
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=400)
    
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return JsonResponse({'success': False, 'message': 'Please select an Excel file.'})
    
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({'success': False, 'message': 'Invalid file format. Only .xlsx and .xls files are supported.'})
    
    try:
        df = pd.read_excel(excel_file, dtype=str)
        
        if df.empty:
            return JsonResponse({'success': False, 'message': 'The uploaded file is empty.'})
        
        # Find required columns
        required_columns = ['Screen Code', 'Screen Name', 'Screen Type']
        missing_columns = []
        for col in required_columns:
            found = False
            for existing_col in df.columns:
                if existing_col.strip().lower() == col.lower():
                    found = True
                    break
            if not found:
                missing_columns.append(col)
        
        if missing_columns:
            return JsonResponse({
                'success': False,
                'message': f'Missing required columns: {", ".join(missing_columns)}'
            })
        
        added_count = 0
        skipped_count = 0
        errors = []
        valid_types = ['ALL', 'ENTRY', 'CONFIGURATION', 'QUERY']
        
        for idx, row in df.iterrows():
            row_num = idx + 2
            
            # Find columns
            screen_code = None
            screen_name = None
            screen_type = 'ALL'
            
            for col in df.columns:
                col_lower = col.strip().lower()
                if 'screen code' in col_lower or 'code' in col_lower:
                    screen_code = str(row.get(col, '')).strip()
                elif 'screen name' in col_lower or 'name' in col_lower:
                    screen_name = str(row.get(col, '')).strip()
                elif 'screen type' in col_lower or 'type' in col_lower:
                    screen_type = str(row.get(col, '')).strip().upper()
            
            if not screen_code or not screen_name:
                errors.append(f'Row {row_num}: Screen Code and Screen Name are required')
                skipped_count += 1
                continue
            
            if screen_type not in valid_types:
                errors.append(f'Row {row_num}: Invalid Screen Type "{screen_type}". Use ALL/ENTRY/CONFIGURATION/QUERY')
                skipped_count += 1
                continue
            
            if ScreenMaster.objects.filter(screen_code__iexact=screen_code).exists():
                errors.append(f'Row {row_num}: Screen Code "{screen_code}" already exists')
                skipped_count += 1
                continue
            
            if ScreenMaster.objects.filter(screen_name__iexact=screen_name).exists():
                errors.append(f'Row {row_num}: Screen Name "{screen_name}" already exists')
                skipped_count += 1
                continue
            
            try:
                ScreenMaster.objects.create(
                    screen_code=screen_code.upper(),
                    screen_name=screen_name,
                    screen_type=screen_type,
                    created_by=request.user.username
                )
                added_count += 1
            except Exception as e:
                errors.append(f'Row {row_num}: {str(e)}')
                skipped_count += 1
        
        # Log the bulk upload
        if added_count > 0:
            log_settings_change(
                request, 'CREATE', 'SCREEN',
                f'Bulk Upload: {added_count} screens',
                new_value=f'Added {added_count} screens, Skipped {skipped_count}',
                change_summary=f'Bulk uploaded {added_count} screens'
            )
        
        message = f'Successfully added {added_count} screens.'
        if skipped_count > 0:
            message += f' Skipped {skipped_count} rows with errors.'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'added_count': added_count,
            'skipped_count': skipped_count,
            'errors': errors[:10]  # Show first 10 errors
        })
        
    except Exception as e:
        logger.error(f"Bulk upload error: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error processing file: {str(e)}'})