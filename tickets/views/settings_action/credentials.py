# tickets/views/settings_action/credentials.py

"""
Credentials Management - Add, Edit, Toggle, Delete, Download, Bulk Upload
"""
from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db import transaction, IntegrityError
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging

from tickets.models import DepartmentCredential, Unit, Department
from .settings_audit import log_settings_change
from ..utils import is_admin

logger = logging.getLogger(__name__)


@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_credentials(request):
    """
    Manage Department Credentials:
    - Add Credential
    - Edit Credential
    - Toggle Active/Inactive
    - Delete Credential
    POST: action (add_credential, edit_credential, toggle_credential, delete_credential)
    Redirects to: settings_credentials_page
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # ========== ADD CREDENTIAL ==========
        if action == 'add_credential':
            uid = request.POST.get('unit')
            did = request.POST.get('department')
            uname = request.POST.get('username', '').strip()
            pwd = request.POST.get('password', '').strip()
            
            if not all([uid, did, uname, pwd]):
                messages.error(request, "All fields are required.")
                return redirect('settings_credentials_page')
            
            # Check if credential already exists for this unit and department
            if DepartmentCredential.objects.filter(unit_id=uid, department_id=did).exists():
                messages.error(request, "Credential already exists for this department.")
                return redirect('settings_credentials_page')
            
            try:
                # Get unit and department for display
                unit = Unit.objects.get(pk=uid)
                dept = Department.objects.get(pk=did)
                
                # Create credential
                cred = DepartmentCredential.objects.create(
                    unit=unit,
                    department=dept,
                    username=uname,
                    password=pwd,
                    is_active=True
                )
                
                # Create/Update Django User
                user, created = User.objects.get_or_create(
                    username=uname,
                    defaults={'is_staff': False}
                )
                if created:
                    user.set_password(pwd)
                    user.save()
                
                messages.success(request, f'Credential for {unit.code} - {dept.name} added successfully!')
                
                log_settings_change(
                    request,
                    action_type='CREATE',
                    setting_type='CREDENTIAL',
                    setting_name=f"Credential: {unit.code} - {dept.name}",
                    old_value=None,
                    new_value=f"Username: {uname}, Unit: {unit.code}, Department: {dept.name}",
                    change_summary=f"Added credential for {unit.code} - {dept.name}",
                    remarks=f"Credential added by {request.user.username}"
                )
            except Unit.DoesNotExist:
                messages.error(request, "Selected unit does not exist.")
            except Department.DoesNotExist:
                messages.error(request, "Selected department does not exist.")
            except Exception as ex:
                messages.error(request, f'Error: {ex}')
        
        # ========== EDIT CREDENTIAL ==========
        elif action == 'edit_credential':
            cred_id = request.POST.get('cred_id')
            if not cred_id:
                messages.error(request, "Credential ID is required.")
                return redirect('settings_credentials_page')
                
            cred = get_object_or_404(DepartmentCredential, pk=cred_id)
            old_username = cred.username
            old_password = cred.password
            nu = request.POST.get('username', '').strip()
            np = request.POST.get('password', '').strip()
            
            if not nu:
                messages.error(request, "Username is required.")
                return redirect('settings_credentials_page')
            
            try:
                # Update credential
                cred.username = nu
                if np:
                    cred.password = np
                cred.save()
                
                # Update Django User
                user = User.objects.filter(username=old_username).first()
                if user:
                    if old_username != nu:
                        user.username = nu
                    if np:
                        user.set_password(np)
                    user.save()
                elif not User.objects.filter(username=nu).exists():
                    User.objects.create_user(username=nu, password=np or cred.password, is_staff=False)
                
                messages.success(request, 'Credential updated successfully!')
                
                change_details = []
                if old_username != nu:
                    change_details.append(f"Username: {old_username} → {nu}")
                if np:
                    change_details.append("Password changed")
                
                log_settings_change(
                    request,
                    action_type='UPDATE',
                    setting_type='CREDENTIAL',
                    setting_name=f"Credential: {cred.unit.code} - {cred.department.name}",
                    old_value=f"Username: {old_username}",
                    new_value=f"Username: {nu}",
                    change_summary='; '.join(change_details) if change_details else 'Credential updated',
                    remarks=f"Credential updated by {request.user.username}"
                )
            except Exception as ex:
                messages.error(request, f'Error: {ex}')
        
        # ========== TOGGLE CREDENTIAL ==========
        elif action == 'toggle_credential':
            cred_id = request.POST.get('cred_id')
            if not cred_id:
                messages.error(request, "Credential ID is required.")
                return redirect('settings_credentials_page')
                
            cred = get_object_or_404(DepartmentCredential, pk=cred_id)
            old_status = 'Active' if cred.is_active else 'Inactive'
            cred.is_active = not cred.is_active
            cred.save()
            new_status = 'Active' if cred.is_active else 'Inactive'
            
            # Update Django User status
            user = User.objects.filter(username=cred.username).first()
            if user:
                user.is_active = cred.is_active
                user.save()
            
            messages.success(request, f'Credential {"activated" if cred.is_active else "deactivated"}.')
            
            log_settings_change(
                request,
                action_type='TOGGLE',
                setting_type='CREDENTIAL',
                setting_name=f"Credential: {cred.unit.code} - {cred.department.name}",
                old_value=f"Status: {old_status}",
                new_value=f"Status: {new_status}",
                change_summary=f"Status changed from {old_status} to {new_status}",
                remarks=f"Credential toggled by {request.user.username}"
            )
        
        # ========== DELETE CREDENTIAL ==========
        elif action == 'delete_credential':
            cred_id = request.POST.get('cred_id')
            if not cred_id:
                messages.error(request, "Credential ID is required.")
                return redirect('settings_credentials_page')
                
            cred = get_object_or_404(DepartmentCredential, pk=cred_id)
            info = f'{cred.unit.code} - {cred.department.name}'
            uname = cred.username
            
            # Deactivate Django User
            user = User.objects.filter(username=uname).first()
            if user:
                user.is_active = False
                user.save()
            
            # Delete credential
            cred.delete()
            messages.success(request, f'Credential for {info} deleted.')
            
            log_settings_change(
                request,
                action_type='DELETE',
                setting_type='CREDENTIAL',
                setting_name=f"Credential: {info}",
                old_value=f"Username: {uname}",
                new_value=None,
                change_summary=f"Deleted credential for {info}",
                remarks=f"Credential deleted by {request.user.username}"
            )
        
        # ========== UNKNOWN ACTION ==========
        else:
            messages.warning(request, f"Unknown action: {action}")
    
    return redirect('settings_credentials_page')


@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def download_credentials(request):
    """
    Download all department credentials as Excel
    Includes: Unit Code, Unit Name, Department, Username, Password, Status
    """
    creds = DepartmentCredential.objects.all().select_related('unit', 'department').order_by('unit__code', 'department__name')
    
    if not creds.exists():
        messages.warning(request, "No credentials available to download.")
        return redirect('settings_credentials_page')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Credentials_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credentials"
    
    # Define styles
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=11)
    
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "Department Credentials - GPLAST"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    # Subtitle with date
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Generated on: {timezone.now().strftime('%d-%b-%Y %I:%M:%S %p')}  |  Total Credentials: {creds.count()}"
    ws['A2'].font = Font(name='Calibri', size=9, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    
    # Headers
    headers = ['Unit Code', 'Unit Name', 'Department', 'Username', 'Password', 'Status']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[4].height = 30
    
    # Data rows
    for row_idx, cred in enumerate(creds, 5):
        row_data = [
            cred.unit.code,
            cred.unit.full_name,
            cred.department.name,
            cred.username,
            cred.password,
            'Active' if cred.is_active else 'Inactive'
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
            
            # Color status cell
            if col_idx == 6:
                if cred.is_active:
                    cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
                    cell.font = Font(name='Calibri', size=11, bold=True, color='22C55E')
                else:
                    cell.fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
                    cell.font = Font(name='Calibri', size=11, bold=True, color='EF4444')
    
    # Set column widths
    column_widths = {'A': 15, 'B': 35, 'C': 30, 'D': 25, 'E': 25, 'F': 15}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    wb.save(response)
    return response


# ============================================================
# ✅ BULK UPLOAD CREDENTIALS - FIXED WITH AJAX HANDLING
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def credentials_bulk_upload(request):
    """
    Bulk upload department credentials from Excel/CSV file
    Expected columns: 'Unit Code', 'Department Name', 'Username', 'Password'
    URL: /custom-admin/settings/credentials/bulk-upload/
    ✅ FIXED: Always returns JSON for AJAX requests
    """
    # Check if it's an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    # Always return JSON for AJAX requests
    if is_ajax:
        try:
            if request.method != 'POST':
                return JsonResponse({
                    'success': False, 
                    'message': 'Invalid method. Use POST.',
                    'added_count': 0,
                    'skipped_count': 0,
                    'errors': ['Invalid request method. Please use POST.']
                }, status=400)
            
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                return JsonResponse({
                    'success': False, 
                    'message': 'Please select an Excel file.',
                    'added_count': 0,
                    'skipped_count': 0,
                    'errors': ['No file selected.']
                })
            
            if not excel_file.name.endswith(('.xlsx', '.xls', '.csv')):
                return JsonResponse({
                    'success': False, 
                    'message': 'Invalid file format. Only .xlsx, .xls, and .csv files are supported.',
                    'added_count': 0,
                    'skipped_count': 0,
                    'errors': ['Unsupported file format.']
                })
            
            # Read the file
            try:
                if excel_file.name.endswith('.csv'):
                    df = pd.read_csv(excel_file, dtype=str)
                else:
                    df = pd.read_excel(excel_file, dtype=str)
            except Exception as e:
                logger.error(f"Error reading file: {str(e)}")
                return JsonResponse({
                    'success': False,
                    'message': f'Error reading file: {str(e)}',
                    'added_count': 0,
                    'skipped_count': 0,
                    'errors': [f'Could not read file: {str(e)}']
                })
            
            if df.empty:
                return JsonResponse({
                    'success': False, 
                    'message': 'The uploaded file is empty.',
                    'added_count': 0,
                    'skipped_count': 0,
                    'errors': ['File is empty.']
                })
            
            # Find columns - look for Unit Code, Department Name, Username, Password
            unit_code_column = None
            dept_name_column = None
            username_column = None
            password_column = None
            
            for col in df.columns:
                col_lower = col.strip().lower()
                if 'unit' in col_lower and ('code' in col_lower or 'id' in col_lower):
                    unit_code_column = col
                if 'department' in col_lower or 'dept' in col_lower:
                    if 'name' in col_lower or 'dept' in col_lower:
                        dept_name_column = col
                if 'user' in col_lower and 'name' in col_lower:
                    username_column = col
                if 'password' in col_lower or 'pwd' in col_lower:
                    password_column = col
            
            # If not found by pattern, use column names based on position
            if unit_code_column is None and len(df.columns) >= 1:
                unit_code_column = df.columns[0]
            if dept_name_column is None and len(df.columns) >= 2:
                dept_name_column = df.columns[1]
            if username_column is None and len(df.columns) >= 3:
                username_column = df.columns[2]
            if password_column is None and len(df.columns) >= 4:
                password_column = df.columns[3]
            
            # Get all active units for validation
            units = {unit.code: unit for unit in Unit.objects.filter(is_active=True)}
            
            added_count = 0
            skipped_count = 0
            errors = []
            added_creds = []
            skipped_creds = []
            
            with transaction.atomic():
                for index, row in df.iterrows():
                    # Get values
                    unit_code = str(row[unit_code_column]).strip() if pd.notna(row[unit_code_column]) else ''
                    dept_name = str(row[dept_name_column]).strip() if pd.notna(row[dept_name_column]) else ''
                    username = str(row[username_column]).strip() if pd.notna(row[username_column]) else ''
                    password = str(row[password_column]).strip() if pd.notna(row[password_column]) else ''
                    
                    # Validate
                    if not unit_code:
                        errors.append(f"Row {index + 1}: Unit Code is empty")
                        skipped_count += 1
                        continue
                    
                    if not dept_name:
                        errors.append(f"Row {index + 1}: Department Name is empty")
                        skipped_count += 1
                        continue
                    
                    if not username:
                        errors.append(f"Row {index + 1}: Username is empty")
                        skipped_count += 1
                        continue
                    
                    if not password:
                        errors.append(f"Row {index + 1}: Password is empty")
                        skipped_count += 1
                        continue
                    
                    # Check if unit exists
                    unit = units.get(unit_code)
                    if not unit:
                        errors.append(f"Row {index + 1}: Unit '{unit_code}' not found")
                        skipped_count += 1
                        continue
                    
                    # Find department by name within the unit
                    department = Department.objects.filter(
                        unit=unit,
                        name__iexact=dept_name,
                        is_active=True
                    ).first()
                    
                    if not department:
                        errors.append(f"Row {index + 1}: Department '{dept_name}' not found in unit '{unit_code}'")
                        skipped_count += 1
                        continue
                    
                    # Check if credential already exists for this unit and department
                    if DepartmentCredential.objects.filter(unit=unit, department=department).exists():
                        errors.append(f"Row {index + 1}: Credential already exists for {unit_code} - {dept_name}")
                        skipped_count += 1
                        skipped_creds.append(f"{unit_code} - {dept_name}")
                        continue
                    
                    # Create credential
                    try:
                        cred = DepartmentCredential.objects.create(
                            unit=unit,
                            department=department,
                            username=username,
                            password=password,
                            is_active=True
                        )
                        
                        # Create/Update Django User
                        user, created = User.objects.get_or_create(
                            username=username,
                            defaults={'is_staff': False}
                        )
                        if created:
                            user.set_password(password)
                            user.save()
                        
                        added_count += 1
                        added_creds.append(f"{unit_code} - {dept_name}")
                        
                    except IntegrityError:
                        errors.append(f"Row {index + 1}: Database constraint error for {unit_code} - {dept_name}")
                        skipped_count += 1
                        skipped_creds.append(f"{unit_code} - {dept_name}")
                    except Exception as e:
                        errors.append(f"Row {index + 1}: Error creating credential '{username}': {str(e)}")
                        skipped_count += 1
            
            # Log bulk upload
            if added_count > 0:
                log_settings_change(
                    request,
                    action_type='CREATE',
                    setting_type='CREDENTIAL',
                    setting_name=f'Bulk Upload: {added_count} Credentials',
                    new_value=f'Added {added_count} credentials',
                    change_summary=f'Bulk uploaded {added_count} department credentials',
                    remarks=f'Bulk uploaded {added_count} credentials, {skipped_count} skipped by {request.user.username}'
                )
            
            message = f'Successfully added {added_count} credentials.'
            if skipped_count > 0:
                message += f' Skipped {skipped_count} entries.'
            
            return JsonResponse({
                'success': True,
                'message': message,
                'added_count': added_count,
                'skipped_count': skipped_count,
                'added_creds': added_creds[:20],
                'skipped_creds': skipped_creds[:20],
                'errors': errors[:20]
            })
            
        except Exception as e:
            logger.error(f"Bulk upload credentials error: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'Error processing file: {str(e)}',
                'added_count': 0,
                'skipped_count': 0,
                'errors': [str(e)]
            }, status=500)
    
    # Non-AJAX request - return error
    return JsonResponse({
        'success': False,
        'message': 'Invalid request. This endpoint only accepts AJAX requests.',
        'added_count': 0,
        'skipped_count': 0,
        'errors': ['Invalid request type.']
    }, status=400)


# ============================================================
# ✅ DOWNLOAD BULK UPLOAD TEMPLATE FOR CREDENTIALS
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def credentials_download_template(request):
    """
    Download Excel template for bulk credential upload
    URL: /custom-admin/settings/credentials/download-template/
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Credentials_Upload_Template.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credentials"
    
    # Styles
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_font = Font(name='Calibri', size=11)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Headers
    headers = ['Unit Code', 'Department Name', 'Username', 'Password']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Sample data with existing units and departments
    existing_units = Unit.objects.filter(is_active=True).order_by('code')
    sample_data = []
    
    if existing_units.exists():
        for unit in existing_units[:3]:
            depts = Department.objects.filter(unit=unit, is_active=True)[:2]
            for dept in depts:
                sample_data.append([
                    unit.code,
                    dept.name,
                    f'{unit.code.lower()}_{dept.name.lower().replace(" ", "_")}',
                    'temp123'
                ])
    
    # If no data, add generic samples
    if len(sample_data) < 3:
        generic_samples = [
            ['HR', 'Human Resources', 'hr_user', 'hr123'],
            ['IT', 'Information Technology', 'it_user', 'it123'],
            ['FIN', 'Finance', 'fin_user', 'fin123'],
        ]
        for idx, sample in enumerate(generic_samples):
            if idx >= len(sample_data):
                sample_data.append(sample)
    
    for row_idx, (unit_code, dept_name, username, password) in enumerate(sample_data, 2):
        for col_idx, value in enumerate([unit_code, dept_name, username, password], 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
    
    # Notes
    note_row = len(sample_data) + 4
    
    # Note 1 - Instructions
    note_cell1 = ws.cell(row=note_row, column=1)
    note_cell1.value = "📌 INSTRUCTIONS:"
    note_cell1.font = Font(name='Calibri', size=10, bold=True, color='1F4E79')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    
    note_row += 1
    note_cell2 = ws.cell(row=note_row, column=1)
    note_cell2.value = "1. Enter existing Unit Code (must match an active unit in the system)"
    note_cell2.font = Font(name='Calibri', size=9, color='333333')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    
    note_row += 1
    note_cell3 = ws.cell(row=note_row, column=1)
    note_cell3.value = "2. Enter Department Name (must exist under the specified unit)"
    note_cell3.font = Font(name='Calibri', size=9, color='333333')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    
    note_row += 1
    note_cell4 = ws.cell(row=note_row, column=1)
    note_cell4.value = "3. Enter Username (will be used as login credential)"
    note_cell4.font = Font(name='Calibri', size=9, color='333333')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    
    note_row += 1
    note_cell5 = ws.cell(row=note_row, column=1)
    note_cell5.value = "4. Enter Password (minimum 6 characters recommended)"
    note_cell5.font = Font(name='Calibri', size=9, color='333333')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    
    note_row += 2
    note_cell6 = ws.cell(row=note_row, column=1)
    note_cell6.value = "⚠️ Unit Code and Department Name must exist in the system. Credentials are created as ACTIVE by default."
    note_cell6.font = Font(name='Calibri', size=9, italic=True, color='FF6B00')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    
    # Add active units list for reference
    note_row += 2
    note_cell7 = ws.cell(row=note_row, column=1)
    note_cell7.value = "📋 Active Units in System:"
    note_cell7.font = Font(name='Calibri', size=9, bold=True, color='1F4E79')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    
    note_row += 1
    unit_codes = ', '.join([unit.code for unit in existing_units]) if existing_units.exists() else 'No units found'
    note_cell8 = ws.cell(row=note_row, column=1)
    note_cell8.value = unit_codes
    note_cell8.font = Font(name='Calibri', size=9, color='666666')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    
    wb.save(response)
    return response