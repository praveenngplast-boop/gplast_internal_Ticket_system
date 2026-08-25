"""Admin-only full database export."""

import json
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.db.models import FileField
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Font, PatternFill

from tickets import models as ticket_models
from .utils import is_admin


BACKUP_MODELS = [
    ticket_models.Unit,
    ticket_models.UnitHead,
    ticket_models.Department,
    ticket_models.AdminContact,
    ticket_models.AdminNotificationEmail,
    ticket_models.EmailSchedule,
    ticket_models.EmployeeMaster,
    ticket_models.DepartmentCredential,
    ticket_models.Ticket,
    ticket_models.TicketHistory,
    ticket_models.ReopenAttachment,
    ticket_models.SettingsAuditLog,
    ticket_models.ERPHolderMapping,
    ticket_models.ScreenMaster,
    ticket_models.ScreenMapping,
    User,
]


def _excel_value(value):
    if value is None:
        return ''
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=True, default=str)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _sheet_name(model, used_names):
    name = model.__name__[:31]
    candidate = name
    suffix = 1
    while candidate in used_names:
        suffix_text = f'_{suffix}'
        candidate = f'{name[:31 - len(suffix_text)]}{suffix_text}'
        suffix += 1
    used_names.add(candidate)
    return candidate


def _write_model_sheet(workbook, model, used_names):
    worksheet = workbook.create_sheet(_sheet_name(model, used_names))
    fields = list(model._meta.concrete_fields)
    many_to_many_fields = list(model._meta.many_to_many)
    headers = [field.name for field in fields] + [field.name for field in many_to_many_fields]
    worksheet.append(headers)

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    worksheet.freeze_panes = 'A2'

    for record in model.objects.all().iterator():
        row = []
        for field in fields:
            value = getattr(record, field.attname)
            if isinstance(field, FileField):
                value = str(value or '')
            row.append(_excel_value(value))
        for field in many_to_many_fields:
            row.append(', '.join(str(value) for value in getattr(record, field.name).values_list('pk', flat=True)))
        worksheet.append(row)

    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max((len(str(cell.value or '')) for cell in column_cells[:100]), default=0)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def download_full_backup(request):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    used_names = set()

    for model in BACKUP_MODELS:
        _write_model_sheet(workbook, model, used_names)

    generated_at = timezone.localtime(timezone.now())
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        'attachment; filename=GPLAST_Full_Backup_'
        f'{generated_at.strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
    workbook.save(response)
    return response
