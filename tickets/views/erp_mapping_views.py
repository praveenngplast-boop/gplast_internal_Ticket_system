# tickets/views/erp_mapping_views.py

"""
ERP User ID Mapping Views
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging

from tickets.models import (
    EmployeeMaster,
    ERPHolderMapping,
    SettingsAuditLog
)
from .utils import is_admin

logger = logging.getLogger(__name__)


# ============================================================
# ERP USER ID MAPPING - MAIN PAGE
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def erp_mapping_page(request):
    """
    ERP User ID Mapping Settings Page
    - View all mappings
    - Add manual mapping
    """
    mappings = ERPHolderMapping.objects.all().select_related('employee').order_by('erp_user_id', 'employee__employee_id')
    
    # Get all ERP User IDs for dropdown
    erp_user_ids = ERPHolderMapping.objects.values_list('erp_user_id', flat=True).distinct().order_by('erp_user_id')
    
    # Get all employees for manual mapping dropdown
    employees = EmployeeMaster.objects.filter(is_active=True).order_by('employee_id')
    
    # Pagination
    paginator = Paginator(mappings, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'mappings': page_obj,
        'erp_user_ids': erp_user_ids,
        'employees': employees,
        'total_mappings': mappings.count(),
    }
    return render(request, 'admin_panel/settings_erp_mapping.html', context)


# ============================================================
# ADD MANUAL MAPPING (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def erp_mapping_add(request):
    """
    Add a new ERP User ID to Employee mapping manually (AJAX)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)
    
    erp_user_id = request.POST.get('erp_user_id', '').strip()
    employee_id = request.POST.get('employee_id', '').strip()
    
    if not erp_user_id:
        return JsonResponse({'success': False, 'message': 'ERP User ID is required'})
    
    if not employee_id:
        return JsonResponse({'success': False, 'message': 'Employee ID is required'})
    
    # Check if employee exists
    try:
        employee = EmployeeMaster.objects.get(employee_id=employee_id, is_active=True)
    except EmployeeMaster.DoesNotExist:
        return JsonResponse({'success': False, 'message': f'Employee ID "{employee_id}" not found'})
    
    # Check if mapping already exists
    existing = ERPHolderMapping.objects.filter(
        erp_user_id=erp_user_id,
        employee=employee
    ).first()
    
    if existing:
        return JsonResponse({
            'success': False,
            'message': f'Mapping already exists: ERP {erp_user_id} → {employee_id} ({employee.employee_name})'
        })
    
    # Create mapping
    try:
        with transaction.atomic():
            mapping = ERPHolderMapping.objects.create(
                erp_user_id=erp_user_id,
                employee=employee
            )
            
            # Log audit
            SettingsAuditLog.objects.create(
                action_type='CREATE',
                setting_type='ERP_MAPPING',
                setting_name=f'ERP {erp_user_id} → {employee_id}',
                new_value=f'ERP User ID: {erp_user_id}, Employee: {employee_id} ({employee.employee_name})',
                change_summary=f'Added ERP mapping: {erp_user_id} → {employee_id}',
                performed_by_name=request.user.username,
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Mapping added: ERP {erp_user_id} → {employee_id} ({employee.employee_name})',
                'mapping': {
                    'id': mapping.id,
                    'erp_user_id': mapping.erp_user_id,
                    'employee_id': employee.employee_id,
                    'employee_name': employee.employee_name,
                    'unit': employee.unit.code if employee.unit else '',
                    'department': employee.department.name if employee.department else '',
                }
            })
    except Exception as e:
        logger.error(f"Error adding ERP mapping: {e}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================================
# REMOVE MAPPING (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def erp_mapping_remove(request):
    """
    Remove an ERP User ID to Employee mapping (AJAX)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)
    
    mapping_id = request.POST.get('mapping_id', '').strip()
    
    if not mapping_id:
        return JsonResponse({'success': False, 'message': 'Mapping ID is required'})
    
    try:
        mapping = ERPHolderMapping.objects.get(id=mapping_id)
    except ERPHolderMapping.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Mapping not found'})
    
    # Store details for audit log
    erp_user_id = mapping.erp_user_id
    employee_id = mapping.employee.employee_id
    employee_name = mapping.employee.employee_name
    
    try:
        with transaction.atomic():
            # Log audit before deleting
            SettingsAuditLog.objects.create(
                action_type='DELETE',
                setting_type='ERP_MAPPING',
                setting_name=f'ERP {erp_user_id} → {employee_id}',
                old_value=f'ERP User ID: {erp_user_id}, Employee: {employee_id} ({employee_name})',
                change_summary=f'Removed ERP mapping: {erp_user_id} → {employee_id}',
                performed_by_name=request.user.username,
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
            
            mapping.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Mapping removed: ERP {erp_user_id} → {employee_id} ({employee_name})'
            })
    except Exception as e:
        logger.error(f"Error removing ERP mapping: {e}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================================
# GET MAPPINGS FOR DROPDOWN / AUTOCOMPLETE (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def erp_mapping_list(request):
    """
    Get all ERP mappings or filter by ERP User ID (AJAX)
    """
    erp_user_id = request.GET.get('erp_user_id', '').strip()
    
    mappings = ERPHolderMapping.objects.all().select_related('employee')
    
    if erp_user_id:
        mappings = mappings.filter(erp_user_id=erp_user_id)
    
    mapping_list = []
    for mapping in mappings:
        mapping_list.append({
            'id': mapping.id,
            'erp_user_id': mapping.erp_user_id,
            'employee_id': mapping.employee.employee_id,
            'employee_name': mapping.employee.employee_name,
            'unit': mapping.employee.unit.code if mapping.employee.unit else '',
            'department': mapping.employee.department.name if mapping.employee.department else '',
        })
    
    return JsonResponse({
        'success': True,
        'mappings': mapping_list,
        'count': len(mapping_list)
    })


# ============================================================
# SEARCH EMPLOYEES FOR DROPDOWN (AJAX)
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def erp_mapping_search_employees(request):
    """
    Search employees by ID or Name for dropdown (AJAX)
    """
    search_term = request.GET.get('q', '').strip()
    
    employees = EmployeeMaster.objects.filter(is_active=True)
    
    if search_term:
        employees = employees.filter(
            Q(employee_id__icontains=search_term) |
            Q(employee_name__icontains=search_term)
        )
    
    employees = employees[:20]
    
    employee_list = []
    for emp in employees:
        employee_list.append({
            'employee_id': emp.employee_id,
            'employee_name': emp.employee_name,
            'unit': emp.unit.code if emp.unit else '',
            'department': emp.department.name if emp.department else '',
        })
    
    return JsonResponse({
        'success': True,
        'employees': employee_list
    })


# ============================================================
# EXPORT MAPPINGS TO EXCEL
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def erp_mapping_export_excel(request):
    """
    Export all ERP User ID mappings to Excel
    """
    mappings = ERPHolderMapping.objects.all().select_related('employee').order_by('erp_user_id', 'employee__employee_id')
    
    current_tz = timezone.get_current_timezone()
    now_utc = timezone.now()
    if timezone.is_naive(now_utc):
        now_utc = timezone.make_aware(now_utc, timezone.utc)
    now_local = now_utc.astimezone(current_tz)
    report_time = now_local.strftime('%d-%b-%Y %I:%M:%S %p')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ERP_Mappings_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ERP Mappings"
    
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10)
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    ws.merge_cells('A1:G1')
    ws['A1'] = f"ERP USER ID MAPPINGS"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Generated: {report_time}  |  Total Mappings: {mappings.count()}"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    
    headers = ['#', 'ERP User ID', 'Employee ID', 'Employee Name', 'Mobile', 'Unit', 'Department']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[4].height = 30
    
    row_idx = 5
    for idx, mapping in enumerate(mappings, 1):
        row_data = [
            idx,
            mapping.erp_user_id,
            mapping.employee.employee_id,
            mapping.employee.employee_name,
            mapping.employee.mobile,
            mapping.employee.unit.code if mapping.employee.unit else '',
            mapping.employee.department.name if mapping.employee.department else '',
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        row_idx += 1
    
    column_widths = {'A': 8, 'B': 18, 'C': 18, 'D': 30, 'E': 16, 'F': 20, 'G': 25}
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    wb.save(response)
    return response