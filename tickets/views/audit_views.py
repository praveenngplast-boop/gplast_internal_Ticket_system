"""Views for the settings audit log and its Excel export."""

from datetime import datetime, time, timedelta

import openpyxl
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl.styles import Alignment, Font, PatternFill

from tickets.models import SettingsAuditLog

from .utils import is_admin


DATE_FORMAT = '%Y-%m-%d'


def _parse_date(value):
    try:
        return datetime.strptime(value, DATE_FORMAT).date()
    except (TypeError, ValueError):
        return None


def _filtered_audit_logs(request):
    logs = SettingsAuditLog.objects.all().order_by('-created_at')
    action_type = request.GET.get('action', '').strip()
    setting_type = request.GET.get('setting_type', '').strip()
    performed_by = request.GET.get('performed_by', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    search = request.GET.get('search', '').strip()

    if action_type:
        logs = logs.filter(action_type=action_type)
    if setting_type:
        logs = logs.filter(setting_type=setting_type)
    if performed_by:
        logs = logs.filter(performed_by_name__icontains=performed_by)

    current_tz = timezone.get_current_timezone()
    parsed_from = _parse_date(date_from)
    parsed_to = _parse_date(date_to)
    date_error = None

    if date_from and not parsed_from:
        date_error = 'Please enter a valid Date From value.'
    elif date_to and not parsed_to:
        date_error = 'Please enter a valid Date To value.'
    elif parsed_from and parsed_to and parsed_to < parsed_from:
        date_error = 'Date To cannot be earlier than Date From.'

    if date_error:
        return logs.none(), date_error

    if parsed_from:
        start = timezone.make_aware(datetime.combine(parsed_from, time.min), current_tz)
        logs = logs.filter(created_at__gte=start)

    if parsed_to:
        # Use an exclusive next-day boundary to include every time on date_to.
        next_day = parsed_to + timedelta(days=1)
        end = timezone.make_aware(datetime.combine(next_day, time.min), current_tz)
        logs = logs.filter(created_at__lt=end)

    if search:
        logs = logs.filter(
            Q(setting_name__icontains=search) |
            Q(change_summary__icontains=search) |
            Q(performed_by_name__icontains=search)
        )

    return logs, None


@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_audit_log(request):
    logs, date_error = _filtered_audit_logs(request)
    page_obj = Paginator(logs, 10).get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'admins': SettingsAuditLog.objects.values_list('performed_by_name', flat=True).distinct(),
        'action_types': SettingsAuditLog.ACTION_TYPES,
        'setting_types': SettingsAuditLog.SETTING_TYPES,
        'selected_action': request.GET.get('action', '').strip(),
        'selected_setting_type': request.GET.get('setting_type', '').strip(),
        'selected_performed_by': request.GET.get('performed_by', '').strip(),
        'date_from': request.GET.get('date_from', '').strip(),
        'date_to': request.GET.get('date_to', '').strip(),
        'date_error': date_error,
        'search_query': request.GET.get('search', '').strip(),
    }
    return render(request, 'admin_panel/audit_log.html', context)


@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def download_audit_log_excel(request):
    logs, date_error = _filtered_audit_logs(request)
    generated_at = timezone.localtime(timezone.now())
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename=Audit_Log_{generated_at.strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Audit Log'
    headers = [
        'ID', 'Action', 'Setting Type', 'Setting Name', 'Old Value', 'New Value',
        'Change Summary', 'Performed By', 'IP Address', 'Remarks', 'Created At',
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    current_tz = timezone.get_current_timezone()
    for log in logs:
        created_at = log.created_at
        if created_at and timezone.is_naive(created_at):
            created_at = timezone.make_aware(created_at, timezone.utc)
        created_at = created_at.astimezone(current_tz) if created_at else None
        worksheet.append([
            log.id,
            log.get_action_type_display(),
            log.get_setting_type_display(),
            log.setting_name,
            log.old_value or '',
            log.new_value or '',
            log.change_summary or '',
            log.get_performed_by_display(),
            log.ip_address or 'N/A',
            log.remarks or '',
            created_at.strftime('%d-%b-%Y %I:%M:%S %p') if created_at else '',
        ])

    workbook.save(response)
    return response