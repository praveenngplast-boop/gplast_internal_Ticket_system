# tickets/views/settings_views.py

"""
Settings Views (GET) - All settings page views
- Settings Dashboard
- Units & Departments
- Communication
- Employees
- Credentials
- Dept Employees
- Audit Logs
- Audit Log Excel Download
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from tickets.models import (
    Unit, Department, AdminContact, AdminNotificationEmail, 
    EmployeeMaster, DepartmentCredential, SettingsAuditLog
)
from tickets.forms import UnitForm, DepartmentForm, AdminContactForm, AdminNotificationEmailForm

from .utils import (
    is_admin,
    _get_contact_data,
    _get_employee_directory_data,
    _get_credentials_data,
)


# ============================================================
# SETTINGS DASHBOARD
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_page(request):
    """
    Settings dashboard/index page with navigation cards
    URL: /admin/settings/
    """
    return render(request, 'admin_panel/settings_index.html')


# ============================================================
# UNITS & DEPARTMENTS
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_units_departments(request):
    """
    Manage Units and Departments
    URL: /admin/settings/units-departments/
    """
    context = {
        'unit_form': UnitForm(),
        'dept_form': DepartmentForm(),
        'units': Unit.objects.all().order_by('code'),
        'departments': Department.objects.all().order_by('unit__code', 'name'),
    }
    return render(request, 'admin_panel/units_departments.html', context)


# ============================================================
# COMMUNICATION SETTINGS
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_communication(request):
    """
    Manage Helpdesk Contact and Notification Emails
    URL: /admin/settings/communication/
    """
    contact_obj = _get_contact_data()
    context = {
        'contact': contact_obj,
        'contact_form': AdminContactForm(instance=contact_obj),
        'email_form': AdminNotificationEmailForm(),
        'emails': AdminNotificationEmail.objects.all().order_by('-created_at'),
    }
    return render(request, 'admin_panel/communication.html', context)


# ============================================================
# EMPLOYEE MANAGEMENT
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_employees_page(request):
    """
    Complete employee management with Can Assign toggle
    URL: /admin/settings/employees/
    """
    employees_qs, emp_search = _get_employee_directory_data(request)
    context = {
        'employees': employees_qs,
        'emp_search': emp_search,
        'all_units': Unit.objects.filter(is_active=True).order_by('code'),
        'departments': Department.objects.all().order_by('unit__code', 'name'),
    }
    return render(request, 'admin_panel/employees.html', context)


# ============================================================
# DEPARTMENT CREDENTIALS
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_credentials_page(request):
    """
    Manage department credentials
    URL: /admin/settings/credentials/
    """
    all_credentials, credentials_by_unit = _get_credentials_data()
    context = {
        'all_units': Unit.objects.filter(is_active=True).order_by('code'),
        'credentials': all_credentials,
        'credentials_by_unit': credentials_by_unit,
        'employee_users': User.objects.filter(is_staff=False, is_active=True).order_by('username'),
    }
    return render(request, 'admin_panel/credentials.html', context)


# ============================================================
# DEPARTMENT-WISE EMPLOYEE VIEW
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_dept_employees(request):
    """
    View employees organized by department (tree view)
    URL: /admin/settings/dept-employees/
    """
    all_credentials, credentials_by_unit = _get_credentials_data()
    context = {
        'all_units': Unit.objects.filter(is_active=True).order_by('code'),
        'credentials_by_unit': credentials_by_unit,
    }
    return render(request, 'admin_panel/dept_employees.html', context)


# ============================================================
# SETTINGS AUDIT LOG
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def settings_audit_log(request):
    """
    View settings audit logs with filters
    URL: /admin/settings/audit/
    """
    logs = SettingsAuditLog.objects.all()
    
    # Filters
    action_type = request.GET.get('action', '')
    setting_type = request.GET.get('setting_type', '')
    performed_by = request.GET.get('performed_by', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search = request.GET.get('search', '')
    
    if action_type:
        logs = logs.filter(action_type=action_type)
    if setting_type:
        logs = logs.filter(setting_type=setting_type)
    if performed_by:
        logs = logs.filter(performed_by_name__icontains=performed_by)
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            logs = logs.filter(created_at__date__gte=date_from_obj.date())
        except:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            logs = logs.filter(created_at__date__lte=date_to_obj.date())
        except:
            pass
    if search:
        logs = logs.filter(
            Q(setting_name__icontains=search) |
            Q(change_summary__icontains=search) |
            Q(performed_by_name__icontains=search)
        )
    
    # Pagination - 10 logs per page
    paginator = Paginator(logs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get unique admins for filter dropdown
    admins = SettingsAuditLog.objects.values_list('performed_by_name', flat=True).distinct()
    
    context = {
        'page_obj': page_obj,
        'admins': admins,
        'action_types': SettingsAuditLog.ACTION_TYPES,
        'setting_types': SettingsAuditLog.SETTING_TYPES,
        'selected_action': action_type,
        'selected_setting_type': setting_type,
        'selected_performed_by': performed_by,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search,
    }
    return render(request, 'admin_panel/audit_log.html', context)


# ============================================================
# SETTINGS AUDIT LOG - EXCEL DOWNLOAD
# ============================================================
@login_required
@user_passes_test(is_admin, login_url='tickets:login')
def download_audit_log_excel(request):
    """
    Export filtered audit logs to Excel
    URL: /admin/settings/audit/download-excel/
    """
    logs = SettingsAuditLog.objects.all()
    
    # Apply filters (same as settings_audit_log)
    action_type = request.GET.get('action', '')
    setting_type = request.GET.get('setting_type', '')
    performed_by = request.GET.get('performed_by', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search = request.GET.get('search', '')
    
    if action_type:
        logs = logs.filter(action_type=action_type)
    if setting_type:
        logs = logs.filter(setting_type=setting_type)
    if performed_by:
        logs = logs.filter(performed_by_name__icontains=performed_by)
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            logs = logs.filter(created_at__date__gte=date_from_obj.date())
        except:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            logs = logs.filter(created_at__date__lte=date_to_obj.date())
        except:
            pass
    if search:
        logs = logs.filter(
            Q(setting_name__icontains=search) |
            Q(change_summary__icontains=search) |
            Q(performed_by_name__icontains=search)
        )
    
    # Create Excel response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Audit_Log_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Log"
    
    # ========== STYLES ==========
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10)
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Get the current timezone from Django settings
    current_tz = timezone.get_current_timezone()
    
    # Convert current time to local timezone for header
    now_utc = timezone.now()
    if timezone.is_naive(now_utc):
        now_utc = timezone.make_aware(now_utc, timezone.utc)
    now_local = now_utc.astimezone(current_tz)
    report_time = now_local.strftime('%d-%b-%Y %I:%M:%S %p')
    
    # ========== TITLE WITH REPORT TIME AND ENTRIES COUNT ==========
    ws.merge_cells('A1:K1')
    ws['A1'] = f"GPLAST SETTINGS AUDIT LOG - Generated: {report_time}  |  Total Entries: {logs.count()}"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45
    
    # ========== HEADERS ==========
    headers = ['ID', 'Action', 'Setting Type', 'Setting Name', 'Old Value', 'New Value', 
               'Change Summary', 'Performed By', 'IP Address', 'Remarks', 'Created At']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[3].height = 30
    
    # ========== DATA ==========
    row_idx = 4
    
    for log in logs:
        ws.cell(row=row_idx, column=1, value=log.id).font = data_font
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_idx, column=1).border = thin_border
        
        ws.cell(row=row_idx, column=2, value=log.get_action_type_display()).font = data_font
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row_idx, column=2).border = thin_border
        
        ws.cell(row=row_idx, column=3, value=log.get_setting_type_display()).font = data_font
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row_idx, column=3).border = thin_border
        
        ws.cell(row=row_idx, column=4, value=log.setting_name).font = data_font
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row_idx, column=4).border = thin_border
        
        ws.cell(row=row_idx, column=5, value=log.old_value or '').font = data_font
        ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=row_idx, column=5).border = thin_border
        
        ws.cell(row=row_idx, column=6, value=log.new_value or '').font = data_font
        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=row_idx, column=6).border = thin_border
        
        ws.cell(row=row_idx, column=7, value=log.change_summary or '').font = data_font
        ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=row_idx, column=7).border = thin_border
        
        ws.cell(row=row_idx, column=8, value=log.performed_by_name).font = data_font
        ws.cell(row=row_idx, column=8).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row_idx, column=8).border = thin_border
        
        ws.cell(row=row_idx, column=9, value=log.ip_address or 'N/A').font = data_font
        ws.cell(row=row_idx, column=9).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row_idx, column=9).border = thin_border
        
        ws.cell(row=row_idx, column=10, value=log.remarks or '').font = data_font
        ws.cell(row=row_idx, column=10).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.cell(row=row_idx, column=10).border = thin_border
        
        # ========== Convert UTC to Local Timezone ==========
        if log.created_at:
            if timezone.is_naive(log.created_at):
                utc_time = timezone.make_aware(log.created_at, timezone.utc)
            else:
                utc_time = log.created_at
            local_time = utc_time.astimezone(current_tz)
            formatted_time = local_time.strftime('%d-%b-%Y %I:%M:%S %p')
        else:
            formatted_time = ''
        
        ws.cell(row=row_idx, column=11, value=formatted_time).font = data_font
        ws.cell(row=row_idx, column=11).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_idx, column=11).border = thin_border
        
        row_idx += 1
    
    # ========== SET COLUMN WIDTHS ==========
    column_widths = {
        'A': 10,   # ID
        'B': 20,   # Action
        'C': 20,   # Setting Type
        'D': 30,   # Setting Name
        'E': 35,   # Old Value
        'F': 35,   # New Value
        'G': 40,   # Change Summary
        'H': 22,   # Performed By
        'I': 18,   # IP Address
        'J': 35,   # Remarks
        'K': 25,   # Created At
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    wb.save(response)
    return response